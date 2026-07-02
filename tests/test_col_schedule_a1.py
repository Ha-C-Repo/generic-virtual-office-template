"""Tests for the A1 schedule-QTY reader and the export honest-null.

A1 corrects the SP183 defect where base-plate-schedule HSS rows landed
in the beam set, and it splits two cases that must never be conflated:

- A column or footing schedule WITH a quantity column gives a summed
  COL count, cross-checked against plan callouts (a disagreement is a
  CONFLICT, never resolved silently).
- A base-plate or column-size schedule with NO quantity column names
  column TYPES, not a member count. It produces no count row; the count
  comes from the foundation plan and Ivan verifies it. A row-count of
  types (4 types on a 16-column building) is never emitted as a total.

The end-to-end test pairs that honest null with Engine B's AREA rows.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from takeoff_pipeline import census, export_xlsx, validate_takeoff
from takeoff_pipeline.census import classify_hit


# -- classify_hit ------------------------------------------------------------

def test_base_plate_and_column_size_schedules_promote_shapes_to_col():
    assert classify_hit(
        "BEAM", "S4.00 FOUNDATION DETAILS BASE PLATE SCHEDULE") == "COL"
    assert classify_hit("BEAM", "BASE PLATE SCHEDULE") == "COL"
    assert classify_hit("BEAM", "COLUMN SIZE SCHEDULE") == "COL"
    # A PL entry in a base plate schedule is the base plate, not the
    # column: it stays PLATE.
    assert classify_hit("PLATE", "S4.00 BASE PLATE SCHEDULE") == "PLATE"
    assert classify_hit("ANCH", "BASE PLATE SCHEDULE") == "ANCH"


def test_existing_classify_hit_paths_unchanged():
    assert classify_hit("BEAM", "S4.00 COLUMN SCHEDULE") == "COL"
    assert classify_hit("BEAM", "FOOTING SCHEDULE") == "COL"
    assert classify_hit("BEAM", "S2.00 FOUNDATION PLAN") == "BEAM"
    assert classify_hit("PLATE", "S2.00 FOUNDATION PLAN") == "PLATE"
    assert classify_hit("", "") == "MISC"


def test_framing_plan_beams_never_reclassified_to_col():
    # The structural guarantee: census passes an EMPTY source for plan
    # callouts (census.run_census plan sweep), so a plan beam can never
    # reach the schedule-promotion branches. Promotion is scoped to
    # schedule-table headings only; a source-string collision cannot
    # silently turn framing-plan beams into columns.
    assert classify_hit("BEAM", "") == "BEAM"
    assert classify_hit("BEAM", "S3.00 ROOF FRAMING PLAN") == "BEAM"
    # A beam schedule, or any source carrying SCHEDULE without a
    # column/base-plate signal, stays BEAM.
    assert classify_hit("BEAM", "S3.00 BEAM SCHEDULE") == "BEAM"
    assert classify_hit(
        "BEAM", "S3.00 ROOF FRAMING PLAN BEAM SCHEDULE") == "BEAM"


# -- census persistence helper -----------------------------------------------

def _hit(designation, item_class, source_kind, primary_source, qty=None,
         confidence="high", sheet="S4.00", bbox="[1.0, 1.0, 2.0, 2.0]"):
    return {"job": "J", "designation": designation,
            "item_class": item_class, "sheet": sheet, "bbox": bbox,
            "raw_text": designation, "source_kind": source_kind,
            "primary_source": primary_source, "confidence": confidence,
            "conflict_group": None, "qty": qty, "created_at": "t"}


def _persist(tmp_path, hits):
    """Write hits the way census.run_census does: compute conflicts
    (which also marks the hits), then insert both tables."""
    db = tmp_path / "census.db"
    census.init_db(db)
    conflicts = census._find_conflicts("J", hits, "t")
    c = census._conn(db)
    try:
        c.executemany(
            "INSERT INTO census_hits (job, designation, item_class, sheet,"
            " bbox, raw_text, source_kind, primary_source, confidence,"
            " conflict_group, qty, created_at) VALUES"
            " (:job, :designation, :item_class, :sheet, :bbox, :raw_text,"
            "  :source_kind, :primary_source, :confidence,"
            "  :conflict_group, :qty, :created_at)", hits)
        c.executemany(
            "INSERT INTO conflicts (job, conflict_group, designation,"
            " item_class, schedule_qty, plan_qty, schedule_source,"
            " plan_source, note, created_at) VALUES"
            " (:job, :conflict_group, :designation, :item_class,"
            "  :schedule_qty, :plan_qty, :schedule_source, :plan_source,"
            "  :note, :created_at)", conflicts)
        c.commit()
    finally:
        c.close()
    return db


# -- type-only schedule: no fabricated count ---------------------------------

def test_type_only_col_schedule_emits_no_count_row(tmp_path):
    db = _persist(tmp_path, [
        _hit("HSS 6x6", "COL", "SCHEDULE", "S4.00 BASE PLATE SCHEDULE"),
        _hit("HSS 8x8", "COL", "SCHEDULE", "S4.00 BASE PLATE SCHEDULE"),
        _hit("HSS 10x10", "COL", "SCHEDULE", "S4.00 BASE PLATE SCHEDULE"),
    ])
    rows = export_xlsx._census_rows("J", db)
    assert not [r for r in rows if r["item_class"] == "COL"]
    types = export_xlsx.uncounted_col_types("J", db)
    assert len(types) == 3
    assert any("HSS 6x6" in t for t in types)
    assert all("BASE PLATE SCHEDULE" in t for t in types)


def test_type_only_plate_schedule_emits_no_count_row(tmp_path):
    # The identical defect for base plates: 4 PL sizes on a schedule are
    # types, not a count. No per-type count row; the base-plate count is
    # derived one per verified column downstream (P29).
    db = _persist(tmp_path, [
        _hit('PL3/8" X 11"', "PLATE", "SCHEDULE",
             "S4.00 BASE PLATE SCHEDULE"),
        _hit('PL3/8" X 13"', "PLATE", "SCHEDULE",
             "S4.00 BASE PLATE SCHEDULE"),
    ])
    rows = export_xlsx._census_rows("J", db)
    assert not [r for r in rows if r["item_class"] == "PLATE"]
    types = export_xlsx.uncounted_plate_types("J", db)
    assert len(types) == 2
    assert any("PL3/8" in t for t in types)


def test_plate_plan_callouts_are_counted_not_suppressed(tmp_path):
    # A plate called out on a detail/plan IS a real count and must not
    # be suppressed. require_no_plan blocks suppression here.
    db = _persist(tmp_path, [
        _hit("PL1/2", "PLATE", "PLAN", "S5.02 ROOF FRAMING DETAILS",
             confidence="medium", sheet="S5.02"),
        _hit("PL1/2", "PLATE", "PLAN", "S5.02 ROOF FRAMING DETAILS",
             confidence="medium", sheet="S5.02"),
    ])
    rows = export_xlsx._census_rows("J", db)
    plate = [r for r in rows if r["item_class"] == "PLATE"]
    assert len(plate) == 1
    assert plate[0]["qty"] == 2.0  # two plan callouts, a real count
    assert export_xlsx.uncounted_plate_types("J", db) == []


# -- real QTY column: a count stands -----------------------------------------

def test_col_schedule_with_qty_produces_count(tmp_path):
    db = _persist(tmp_path, [
        _hit("HSS8X8X1/4", "COL", "SCHEDULE", "S4.00 COLUMN SCHEDULE",
             qty=16.0),
    ])
    rows = export_xlsx._census_rows("J", db)
    col = [r for r in rows if r["item_class"] == "COL"]
    assert len(col) == 1
    assert col[0]["qty"] == 16.0
    assert col[0]["mode"] == "COUNT" and col[0]["unit"] == "EA"
    assert export_xlsx.uncounted_col_types("J", db) == []


def test_col_count_with_disagreeing_plan_is_conflict(tmp_path):
    hits = [_hit("HSS8X8X1/4", "COL", "SCHEDULE", "S4.00 COLUMN SCHEDULE",
                 qty=16.0)]
    hits += [_hit("HSS8X8X1/4", "BEAM", "PLAN", "S2.00 FOUNDATION PLAN",
                  confidence="medium", sheet="S2.00")
             for _ in range(14)]
    db = _persist(tmp_path, hits)
    rows = export_xlsx._census_rows("J", db)
    col = [r for r in rows if r["item_class"] == "COL"]
    assert len(col) == 1
    assert col[0]["qty"] == 16.0  # schedule is primary (P30)
    assert col[0]["confidence"] == "low"
    assert col[0]["notes"].startswith("CONFLICT:")
    assert "16" in col[0]["notes"] and "14" in col[0]["notes"]


# -- end to end: A1 honest null plus Engine B AREA rows ----------------------

def _grid_rows():
    base = {"secondary_source": "", "confidence": "medium",
            "bbox": "[10.0, 20.0, 30.0, 40.0]", "notes": "grid footprint"}
    return [
        {"item_class": "DECK", "designation": "ROOF DECK", "mode": "AREA",
         "qty": 61950, "unit": "SF",
         "primary_source": "S3.00 ROOF FRAMING PLAN", "sheet": "S3.00",
         **base},
        {"item_class": "MISC", "designation": "BUILDING GROSS SF",
         "mode": "AREA", "qty": 61950, "unit": "SF",
         "primary_source": "S2.00 FOUNDATION PLAN", "sheet": "S2.00",
         **base},
    ]


def test_build_workbook_integrates_a1_and_engine_b(tmp_path):
    db = _persist(tmp_path, [
        _hit("HSS 6x6", "COL", "SCHEDULE", "S4.00 BASE PLATE SCHEDULE"),
        _hit("HSS 10x10", "COL", "SCHEDULE", "S4.00 BASE PLATE SCHEDULE"),
    ])
    out = tmp_path / "out"
    info = export_xlsx.export("J", db_path=db, out_dir=out,
                              extra_rows=_grid_rows())
    assert info["status"] == "UNSTAMPED_VALID", info
    assert info["grid_rows"] == 2
    assert len(info["col_types_uncounted"]) == 2

    import openpyxl
    wb = openpyxl.load_workbook(info["unstamped"])
    t = wb["TAKEOFF"]
    classes = [str(t.cell(row=r, column=1).value).split("-")[0]
               for r in range(3, t.max_row + 1)
               if t.cell(row=r, column=1).value]
    assert "COL" not in classes  # no fabricated column count
    assert "DECK" in classes and "MISC" in classes  # the SF rows landed
    report = validate_takeoff.validate_file(info["unstamped"])
    assert report["hard_fails"] == []
