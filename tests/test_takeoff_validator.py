"""Tests for export_xlsx.py and validate_takeoff.py (Prompt 5).

Covers the four mandated validator paths (missing mode, missing
primary_source, smuggled pricing column, conflict warning) plus the
13.2 hash round-trip, the typed-constant weight gate, token reference
checks, the AISC lb_per_ft cross-check, and the full build-validate-
stamp pipeline against a synthetic census db.
"""

import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from takeoff_pipeline import export_xlsx, validate_takeoff
from takeoff_pipeline.takeoff_hash import compute_takeoff_hash

H = export_xlsx.TAKEOFF_HEADERS


def _base_workbook():
    """A minimal VALID four-sheet workbook: one COUNT row, one LINEAR
    row with formula weights and a GROUP link, one CONFLICT row, one
    DECK AREA row with an OPENING child."""
    import openpyxl

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "TAKEOFF"
    t.append([])  # row 1: metadata, written only by the stamp
    t.append(list(H))
    t.append(["COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
              "S4.00 COLUMN SCHEDULE", "S2.00 FOUNDATION PLAN", "high",
              "S4.00", "[100.0, 200.0, 140.0, 210.0]", ""])
    t.append(["BEAM-001", "W12X26", "LINEAR", 240, "LF",
              "S3.00 ROOF FRAMING PLAN", "", "medium", "S3.00",
              "[300.0, 400.0, 360.0, 410.0]", "GROUP: COL-001"])
    t.cell(row=4, column=13, value="=D4*L4")
    t.cell(row=4, column=14, value="=M4/2000")
    t.cell(row=4, column=15,
           value="AISC:bridge/aisc_validator.py:W12X26")
    t.append(["ANCH-001", "ANCHOR RODS 3/4 IN", "COUNT", 64, "EA",
              "S4.00 COLUMN SCHEDULE + TYPICAL DETAILS", "S2.00 plan",
              "low", "S4.00", "MANUAL",
              "CONFLICT: column schedule 64 EA vs takeoff record 160 "
              "EA. RFI candidate."])
    t.append(["DECK-001", "TYPE B ROOF DECK", "AREA", 61621, "SF",
              "S3.00 ROOF FRAMING PLAN", "", "medium", "S3.00",
              "MANUAL", ""])
    t.append(["DECK-002", "OPENING", "AREA", 64, "SF",
              "S3.00 ROOF FRAMING PLAN", "", "medium", "S3.00",
              "MANUAL", "OPENING: DECK-001 RTU opening"])

    boq = wb.create_sheet("BOQ")
    boq.append(list(export_xlsx.BOQ_HEADERS))
    bom = wb.create_sheet("BOM")
    bom.append(list(export_xlsx.BOM_HEADERS))
    ps = wb.create_sheet("PRICING_SCHEDULE")
    ps.append(list(export_xlsx.PS_HEADERS))
    ps.append(["PS-COL", "COL measured counts", "COL-001",
               "=SUM(TAKEOFF!D3)", "EA"])
    ps.append(["PS-BEAM", "BEAM derived tonnage", "BEAM-001",
               "=SUM(TAKEOFF!N4)", "TON"])
    return wb


@pytest.fixture
def valid_path(tmp_path):
    path = tmp_path / "takeoff.xlsx"
    _base_workbook().save(str(path))
    return path


def _fails(path):
    return validate_takeoff.validate_file(path)["hard_fails"]


def test_valid_file_passes_with_conflict_warning(valid_path):
    report = validate_takeoff.validate_file(valid_path)
    assert report["hard_fails"] == []
    assert len(report["conflicts"]) == 1
    row, iid, desig, note = report["conflicts"][0]
    assert iid == "ANCH-001"
    assert note.startswith("CONFLICT:")
    assert any(iid == "ANCH-001" for _, iid, _ in report["low_rows"])


def test_missing_mode_hard_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=3, column=3).value = None
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("row 3" in f and "missing mode" in f for f in fails)


def test_missing_primary_source_hard_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=4, column=6).value = None
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("row 4" in f and "missing primary_source" in f
               for f in fails)


def test_pricing_column_smuggled_in_header(tmp_path):
    wb = _base_workbook()
    wb["BOQ"].cell(row=1, column=9, value="unit cost")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("pricing token 'cost'" in f and "BOQ" in f for f in fails)


def test_pricing_tokens_in_data_cells(tmp_path):
    wb = _base_workbook()
    t = wb["TAKEOFF"]
    t.cell(row=3, column=11, value="confirm $1,200 with GC")
    t.cell(row=5, column=7, value="cost basis per office")
    # Words that merely contain a token must pass (section 12).
    t.cell(row=6, column=11, value="accurate separate strategy")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("dollar character" in f and "row 3" in f for f in fails)
    assert any("whole-word pricing token 'cost'" in f and "row 5" in f
               for f in fails)
    assert not any("row 6" in f for f in fails)


def test_conflict_row_must_be_low_confidence(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=5, column=8, value="high")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("CONFLICT row with confidence 'high'" in f
               for f in _fails(p))


def test_unit_must_match_mode(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=3, column=5, value="LF")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("does not match mode COUNT" in f for f in _fails(p))


def test_typed_weight_constant_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=4, column=14, value=3.12)
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("typed constant" in f and "tons" in f for f in _fails(p))


def test_lb_per_ft_on_count_row_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=3, column=12, value=26.0)
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("lb_per_ft populated on a COUNT row" in f
               for f in _fails(p))


def test_group_and_opening_reference_checks(tmp_path):
    wb = _base_workbook()
    t = wb["TAKEOFF"]
    t.cell(row=7, column=11, value="OPENING: COL-001")  # not a DECK row
    t.cell(row=4, column=11, value="GROUP: BEAM-009")  # absent id
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("OPENING: parent 'COL-001' is not a DECK row" in f
               for f in fails)
    assert any("GROUP: cites 'BEAM-009'" in f for f in fails)


def test_extra_worksheet_fails(tmp_path):
    wb = _base_workbook()
    wb.create_sheet("SCRATCH")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("'SCRATCH' is not one of the four named sheets" in f
               for f in _fails(p))


def test_pricing_schedule_lb_unit_excluded(tmp_path):
    wb = _base_workbook()
    wb["PRICING_SCHEDULE"].cell(row=2, column=5, value="LB")
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("LB is intentionally excluded" in f for f in _fails(p))


def test_aisc_lb_per_ft_cross_check(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=4, column=12, value=99.0)  # W12X26 is 26
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    report = validate_takeoff.validate_file(p)
    mismatch = [f for f in report["hard_fails"]
                if "differs from bridge/aisc_validator.py" in f]
    skipped = [w for w in report["warnings"]
               if "could not be cross-checked" in w]
    # In the repo the check runs and fails; in a stripped environment
    # the skip must be reported, never silent.
    assert mismatch or skipped
    if mismatch:
        assert "W12X26" in mismatch[0]


def test_hash_stamp_roundtrip_and_tamper_detection(valid_path, tmp_path):
    import openpyxl

    stamped = export_xlsx.stamp_workbook(valid_path, "TESTJOB",
                                         out_dir=tmp_path)
    assert stamped["export_number"] == "R1"
    final = Path(stamped["path"])
    assert final.name == f"TESTJOB_TAKEOFF_R1_{stamped['hash'][:12]}.xlsx"

    report = validate_takeoff.validate_file(final)
    assert report["hard_fails"] == []

    # The stamp must not change its own hash (row 1 is excluded).
    wb = openpyxl.load_workbook(str(final))
    assert compute_takeoff_hash(wb["TAKEOFF"]) == stamped["hash"]

    # Any post-stamp edit must be caught (13.4: immutable once stamped).
    wb["TAKEOFF"].cell(row=3, column=4, value=17)
    tampered = tmp_path / "tampered.xlsx"
    wb.save(str(tampered))
    fails = _fails(tampered)
    assert any("does not match the recomputed TAKEOFF hash" in f
               for f in fails)


def test_export_numbers_are_monotonic(valid_path, tmp_path):
    s1 = export_xlsx.stamp_workbook(valid_path, "JOBX", out_dir=tmp_path)
    s2 = export_xlsx.stamp_workbook(valid_path, "JOBX", out_dir=tmp_path)
    assert s1["export_number"] == "R1"
    assert s2["export_number"] == "R2"
    # Same TAKEOFF contents, two exports, two distinct filenames.
    assert Path(s1["path"]).name != Path(s2["path"]).name


def test_manual_rows_confidence_never_defaulted(tmp_path):
    csv_path = tmp_path / "manual.csv"
    csv_path.write_text(
        "item_class,designation,mode,qty,unit,primary_source,"
        "secondary_source,confidence,sheet,bbox,notes\n"
        "DECK,TYPE B ROOF DECK,AREA,61621,SF,S3.00 ROOF FRAMING PLAN,"
        ",,S3.00,,from Ivan email 2026-05-25\n",
        encoding="utf-8")
    with pytest.raises(ValueError, match="never defaulted"):
        export_xlsx.load_manual_rows(csv_path)


def _synthetic_census(tmp_path):
    db = tmp_path / "census.db"
    c = sqlite3.connect(str(db))
    c.executescript("""
        CREATE TABLE census_hits (
            id INTEGER PRIMARY KEY AUTOINCREMENT, job TEXT,
            designation TEXT, item_class TEXT, sheet TEXT, bbox TEXT,
            raw_text TEXT, source_kind TEXT, primary_source TEXT,
            confidence TEXT, conflict_group TEXT, qty REAL,
            created_at TEXT);
        CREATE TABLE conflicts (
            id INTEGER PRIMARY KEY AUTOINCREMENT, job TEXT,
            conflict_group TEXT, designation TEXT, item_class TEXT,
            schedule_qty REAL, plan_qty REAL, schedule_source TEXT,
            plan_source TEXT, note TEXT, created_at TEXT);
    """)
    rows = [
        # Schedule-backed column with qty: exact basis.
        ("J", "HSS8X8X1/4", "COL", "S4.00", "[10.0, 10.0, 20.0, 20.0]",
         "HSS8X8X1/4", "SCHEDULE", "S4.00 COLUMN SCHEDULE", "high",
         None, 16.0, "t"),
        # Joist with a schedule (P30 primary) plus plan callouts.
        ("J", "28K7", "JST", "S5.01", "[5.0, 5.0, 9.0, 9.0]",
         "28K7", "SCHEDULE", "S5.01 JOIST SCHEDULE", "high", None,
         63.0, "t"),
        ("J", "28K7", "JST", "S3.00", "[30.0, 10.0, 40.0, 20.0]",
         "28K7", "PLAN", "S3.00 ROOF FRAMING PLAN", "medium", None,
         None, "t"),
        # Plan-only beam on a framing plan: approx basis, stays BEAM.
        ("J", "W16X26", "BEAM", "S3.00", "[42.0, 10.0, 52.0, 20.0]",
         "W16X26", "PLAN", "S3.00 ROOF FRAMING PLAN", "medium", None,
         None, "t"),
        ("J", "W16X26", "BEAM", "S3.00", "[62.0, 10.0, 72.0, 20.0]",
         "W16X26", "PLAN", "S3.00 ROOF FRAMING PLAN", "medium", None,
         None, "t"),
        # Shape evidence OUTSIDE a framing plan: Appendix A unresolved,
        # the exporter must reclass it MISC or it fails section 5.
        ("J", "HSS3X3X1/4", "BEAM", "S5.02",
         "[80.0, 10.0, 90.0, 20.0]", "HSS3X3X1/4", "PLAN",
         "S5.02 LADDER DETAILS", "medium", None, None, "t"),
        # Conflicted anchor: schedule 4 vs plan 2 callouts.
        ("J", "ANCHOR RODS", "ANCH", "S4.00",
         "[70.0, 10.0, 80.0, 20.0]", "ANCHOR RODS", "SCHEDULE",
         "S4.00 COLUMN SCHEDULE", "low", "J:ANCH:ANCHORRODS", 4.0, "t"),
        ("J", "ANCHOR RODS", "ANCH", "S2.00",
         "[90.0, 10.0, 95.0, 20.0]", "ANCHOR RODS", "PLAN",
         "S2.00 TYPICAL DETAILS", "low", "J:ANCH:ANCHORRODS", None,
         "t"),
        # Attribute evidence: must never become a TAKEOFF row.
        ("J", "CAMBER", "MISC", "S1.00", "[1.0, 1.0, 2.0, 2.0]",
         "NO CAMBER REQUIRED", "PLAN", "S1.00 GENERAL NOTES", "low",
         None, None, "t"),
    ]
    c.executemany("INSERT INTO census_hits (job, designation, "
                  "item_class, sheet, bbox, raw_text, source_kind, "
                  "primary_source, confidence, conflict_group, qty, "
                  "created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", rows)
    c.execute("INSERT INTO conflicts (job, conflict_group, designation,"
              " item_class, schedule_qty, plan_qty, schedule_source,"
              " plan_source, note, created_at) VALUES "
              "(?,?,?,?,?,?,?,?,?,?)",
              ("J", "J:ANCH:ANCHORRODS", "ANCHOR RODS", "ANCH", 4.0,
               2.0, "S4.00 COLUMN SCHEDULE", "S2.00 TYPICAL DETAILS",
               "CONFLICT: S4.00 COLUMN SCHEDULE 4 EA vs S2.00 TYPICAL "
               "DETAILS plan callouts 2 EA. RFI candidate. Never "
               "resolved silently.", "t"))
    c.commit()
    c.close()
    return db


def _manual_csv(tmp_path):
    manual = tmp_path / "manual.csv"
    manual.write_text(
        "item_class,designation,mode,qty,unit,primary_source,"
        "secondary_source,confidence,sheet,bbox,notes\n"
        "BEAM,W12X26,LINEAR,240,LF,S3.00 ROOF FRAMING PLAN,,medium,"
        "S3.00,MANUAL,stick length from grid dims\n"
        "DECK,TYPE B ROOF DECK,AREA,61621,SF,S3.00 ROOF FRAMING PLAN,"
        ",medium,S3.00,MANUAL,from Ivan email 2026-05-25\n"
        "DECK,OPENING,AREA,64,SF,S3.00 ROOF FRAMING PLAN,,medium,"
        "S3.00,MANUAL,OPENING: DECK-002 RTU opening\n",
        encoding="utf-8")
    return manual


def test_full_export_pipeline(tmp_path):
    """The 13.1 flow as the operator runs it: build and validate
    (UNSTAMPED_VALID), fill lb_per_ft from the AISC validator output,
    then stamp_file. Stamping never happens over blank weights."""
    import openpyxl

    db = _synthetic_census(tmp_path)
    out = tmp_path / "out"
    info = export_xlsx.export("J", db_path=db,
                              manual_csv=_manual_csv(tmp_path),
                              out_dir=out)
    assert info["status"] == "UNSTAMPED_VALID", info
    assert info["conflict_rows"] == 1
    assert any("TON line" in d for d in info["spec_deviations"])

    # Operator step 2: fill lb_per_ft on the LINEAR row from
    # bridge/aisc_validator.py output (W12X26 = 26.0 lb/ft).
    wb = openpyxl.load_workbook(info["unstamped"])
    t = wb["TAKEOFF"]
    rows = {t.cell(row=r, column=2).value: r
            for r in range(3, t.max_row + 1)}
    beam_r = rows["W12X26"]
    assert t.cell(row=beam_r, column=12).value is None
    t.cell(row=beam_r, column=12).value = 26.0
    wb.save(info["unstamped"])

    stamped = export_xlsx.stamp_file(info["unstamped"], "J", out)
    assert stamped["status"] == "STAMPED", stamped

    wb = openpyxl.load_workbook(stamped["path"])
    t = wb["TAKEOFF"]
    assert t.cell(row=1, column=1).value == "TAKEOFF_SCHEMA_V2"
    assert t.cell(row=1, column=6).value == "NONE"
    rows = {t.cell(row=r, column=2).value: r
            for r in range(3, t.max_row + 1)}
    assert "CAMBER" not in rows  # attribute evidence excluded
    anch_r = rows["ANCHOR RODS"]
    assert t.cell(row=anch_r, column=4).value == 4  # schedule primary
    assert str(t.cell(row=anch_r, column=11).value).startswith(
        "CONFLICT:")
    jst_r = rows["28K7"]
    assert t.cell(row=jst_r, column=4).value == 63  # schedule qty
    assert "JOIST SCHEDULE" in str(t.cell(row=jst_r, column=6).value)
    w16_r = rows["W16X26"]
    assert t.cell(row=w16_r, column=4).value == 2  # plan callout count
    assert "QTY_BASIS: approx" in str(t.cell(row=w16_r, column=11).value)
    hss3_r = rows["HSS3X3X1/4"]
    assert str(t.cell(row=hss3_r, column=1).value).startswith("MISC-")
    beam_r = rows["W12X26"]
    assert t.cell(row=beam_r, column=13).value == f"=D{beam_r}*L{beam_r}"
    assert t.cell(row=beam_r, column=12).value == 26.0

    # The COMPLETE pricing schedule, pinned: opening rows form no line
    # of their own; the parent deck line cites them in source_refs.
    ps = wb["PRICING_SCHEDULE"]
    ps_lines = [tuple(ps.cell(row=r, column=c).value for c in (1, 5))
                for r in range(2, ps.max_row + 1)]
    assert ps_lines == [
        ("PS-COL", "EA"), ("PS-BEAM", "TON"), ("PS-BEAM-2", "EA"),
        ("PS-JST", "EA"), ("PS-DECK", "SF"), ("PS-ANCH", "EA"),
        ("PS-MISC", "EA"),
    ]
    deck_line = next(r for r in range(2, ps.max_row + 1)
                     if ps.cell(row=r, column=1).value == "PS-DECK")
    refs = str(ps.cell(row=deck_line, column=3).value)
    assert "DECK-002" in refs and "DECK-001" in refs
    # qty sums only the gross deck row, a single contiguous range.
    assert str(ps.cell(row=deck_line, column=4).value).count(":") == 1

    report = validate_takeoff.validate_file(stamped["path"])
    assert report["hard_fails"] == []
    assert len(report["conflicts"]) == 1


def test_export_rows_are_reproducible(tmp_path):
    """Section 1: the same census reproduces the same takeoff, even
    when sqlite hands hits back in a different insertion order."""
    (tmp_path / "a").mkdir(exist_ok=True)
    db1 = _synthetic_census(tmp_path / "a")
    rows1 = export_xlsx._census_rows("J", db1)

    (tmp_path / "b").mkdir(exist_ok=True)
    db2 = tmp_path / "b" / "census.db"
    src = sqlite3.connect(str(db1))
    dst = sqlite3.connect(str(db2))
    src.backup(dst)
    dst.execute("CREATE TABLE t AS SELECT * FROM census_hits "
                "ORDER BY id DESC")
    dst.execute("DELETE FROM census_hits")
    dst.execute("INSERT INTO census_hits (job, designation, item_class,"
                " sheet, bbox, raw_text, source_kind, primary_source,"
                " confidence, conflict_group, qty, created_at)"
                " SELECT job, designation, item_class, sheet, bbox,"
                " raw_text, source_kind, primary_source, confidence,"
                " conflict_group, qty, created_at FROM t")
    dst.commit()
    src.close()
    dst.close()
    rows2 = export_xlsx._census_rows("J", db2)
    assert rows1 == rows2


def test_two_schedule_sources_disagreeing_become_conflict(tmp_path):
    db = tmp_path / "census.db"
    c = sqlite3.connect(str(db))
    c.executescript("""
        CREATE TABLE census_hits (id INTEGER PRIMARY KEY AUTOINCREMENT,
            job TEXT, designation TEXT, item_class TEXT, sheet TEXT,
            bbox TEXT, raw_text TEXT, source_kind TEXT,
            primary_source TEXT, confidence TEXT, conflict_group TEXT,
            qty REAL, created_at TEXT);
        CREATE TABLE conflicts (id INTEGER PRIMARY KEY AUTOINCREMENT,
            job TEXT, conflict_group TEXT, designation TEXT,
            item_class TEXT, schedule_qty REAL, plan_qty REAL,
            schedule_source TEXT, plan_source TEXT, note TEXT,
            created_at TEXT);
    """)
    c.executemany(
        "INSERT INTO census_hits (job, designation, item_class, sheet,"
        " bbox, raw_text, source_kind, primary_source, confidence,"
        " conflict_group, qty, created_at) VALUES (?,?,?,?,?,?,?,?,?,"
        "?,?,?)",
        [("J", "HSS8X8X1/4", "COL", "S0.1", "[1.0, 1.0, 2.0, 2.0]",
          "x", "SCHEDULE", "S0.1 COLUMN SCHEDULE", "high", None, 10.0,
          "t"),
         ("J", "HSS8X8X1/4", "COL", "S0.2", "[1.0, 1.0, 2.0, 2.0]",
          "x", "SCHEDULE", "S0.2 FOOTING SCHEDULE", "high", None, 12.0,
          "t")])
    c.commit()
    c.close()
    rows = export_xlsx._census_rows("J", db)
    assert len(rows) == 1
    row = rows[0]
    assert row["qty"] == 10.0  # first source in reading order, never 22
    assert row["confidence"] == "low"
    assert row["notes"].startswith("CONFLICT:")
    assert "10" in row["notes"] and "12" in row["notes"]


def test_plan_hits_on_two_sheets_are_two_rows(tmp_path):
    db = tmp_path / "census.db"
    c = sqlite3.connect(str(db))
    c.executescript("""
        CREATE TABLE census_hits (id INTEGER PRIMARY KEY AUTOINCREMENT,
            job TEXT, designation TEXT, item_class TEXT, sheet TEXT,
            bbox TEXT, raw_text TEXT, source_kind TEXT,
            primary_source TEXT, confidence TEXT, conflict_group TEXT,
            qty REAL, created_at TEXT);
        CREATE TABLE conflicts (id INTEGER PRIMARY KEY AUTOINCREMENT,
            job TEXT, conflict_group TEXT, designation TEXT,
            item_class TEXT, schedule_qty REAL, plan_qty REAL,
            schedule_source TEXT, plan_source TEXT, note TEXT,
            created_at TEXT);
    """)
    hits = [("J", "W16X26", "BEAM", "S3.00", "[1.0, 1.0, 2.0, 2.0]",
             "x", "PLAN", "S3.00 ROOF FRAMING PLAN", "medium", None,
             None, "t")] * 2 + \
           [("J", "W16X26", "BEAM", "S3.01", "[1.0, 1.0, 2.0, 2.0]",
             "x", "PLAN", "S3.01 ROOF FRAMING PLAN", "medium", None,
             None, "t")] * 3
    c.executemany(
        "INSERT INTO census_hits (job, designation, item_class, sheet,"
        " bbox, raw_text, source_kind, primary_source, confidence,"
        " conflict_group, qty, created_at) VALUES (?,?,?,?,?,?,?,?,?,"
        "?,?,?)", hits)
    c.commit()
    c.close()
    rows = export_xlsx._census_rows("J", db)
    assert len(rows) == 2  # Appendix A: one row per primary_source
    by_src = {r["primary_source"]: r["qty"] for r in rows}
    assert by_src["S3.00 ROOF FRAMING PLAN"] == 2.0
    assert by_src["S3.01 ROOF FRAMING PLAN"] == 3.0


def test_disguised_constant_formula_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=4, column=14).value = "=3.12"
    bom = wb["BOM"]
    bom.append(["BM-1", "COL-001", "HSS8X8X1/4", "col", 16, "EA",
                None, None, "=5.0", None, "", ""])
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("row 4" in f and "tons" in f
               and "cell-referencing" in f for f in fails)
    assert any(f.startswith("BOM row 2: weight_lb") for f in fails)


def test_bom_derived_without_formula_ref_fails(tmp_path):
    wb = _base_workbook()
    wb["BOM"].append(["BM-1", "COL-001", "HSS8X8X1/4", "col", 16,
                      "EA", None, None, "=E2*H2", "=I2/2000", "", ""])
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("BOM row 2: derived value without formula_ref" in f
               for f in _fails(p))


def test_stamp_removal_is_caught(valid_path, tmp_path):
    import openpyxl

    stamped = export_xlsx.stamp_workbook(valid_path, "TAMPER",
                                         out_dir=tmp_path)
    wb = openpyxl.load_workbook(stamped["path"])
    wb["TAKEOFF"].cell(row=1, column=1).value = None  # remove A1
    wb["TAKEOFF"].cell(row=3, column=4).value = 99  # tamper qty
    p = tmp_path / "cleared.xlsx"
    wb.save(str(p))
    fails = _fails(p)
    assert any("stamp damaged or removed" in f for f in fails)
    # Same with the whole row 1 cleared but the stamped FILENAME kept.
    wb2 = openpyxl.load_workbook(stamped["path"])
    for c in range(1, 7):
        wb2["TAKEOFF"].cell(row=1, column=c).value = None
    wb2["TAKEOFF"].cell(row=3, column=4).value = 99
    wb2.save(stamped["path"])
    fails2 = _fails(stamped["path"])
    assert any("stamp damaged or removed" in f for f in fails2)


def test_content_beyond_layout_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=3, column=16).value = "smuggled"
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("beyond the sheet's 15-column layout" in f
               for f in _fails(p))


def test_pricing_token_in_metadata_row_fails(tmp_path):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=1, column=7).value = "$500 owed"
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    assert any("row 1" in f and "dollar character" in f
               for f in _fails(p))


def test_assembly_crosscheck_warns(tmp_path):
    wb = _base_workbook()
    wb["BOQ"].append(["BQ-1", "ANCH-001", "asm-col-base", "MATERIAL",
                      "anchor rods from assembly", 72, "EA", "", ""])
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    report = validate_takeoff.validate_file(p)
    assert report["hard_fails"] == []
    assert any("disagrees with measured ANCH row ANCH-001" in w
               for w in report["warnings"])


def test_aisc_skip_is_reported_not_silent(tmp_path, monkeypatch):
    wb = _base_workbook()
    wb["TAKEOFF"].cell(row=4, column=12).value = 26.0
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    monkeypatch.setitem(sys.modules, "bridge.aisc_validator", None)
    monkeypatch.setitem(sys.modules, "bridge", None)
    report = validate_takeoff.validate_file(p)
    assert any("could not be cross-checked" in w
               for w in report["warnings"])
    assert not any("differs from" in f for f in report["hard_fails"])


def test_cli_exit_codes(valid_path, tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(sys, "argv", ["validate_takeoff"])
    assert validate_takeoff.main() == 2

    monkeypatch.setattr(sys, "argv",
                        ["validate_takeoff", str(valid_path)])
    assert validate_takeoff.main() == 0
    out = capsys.readouterr().out
    assert "RFI candidates" in out and "ANCH-001" in out

    import openpyxl
    wb = openpyxl.load_workbook(str(valid_path))
    wb["TAKEOFF"].cell(row=3, column=3).value = None
    bad = tmp_path / "bad.xlsx"
    wb.save(str(bad))
    monkeypatch.setattr(sys, "argv", ["validate_takeoff", str(bad)])
    assert validate_takeoff.main() == 1


def test_export_register_survives_archived_files(valid_path, tmp_path):
    s1 = export_xlsx.stamp_workbook(valid_path, "ARCH", out_dir=tmp_path)
    archive = tmp_path / "archive"
    archive.mkdir()
    Path(s1["path"]).rename(archive / Path(s1["path"]).name)
    s2 = export_xlsx.stamp_workbook(valid_path, "ARCH", out_dir=tmp_path)
    assert s2["export_number"] == "R2"  # never reissues R1


def test_stamp_refuses_unknown_assembly_version(tmp_path):
    wb = _base_workbook()
    wb["BOQ"].append(["BQ-1", "COL-001", "asm", "LABOR", "fit and weld",
                      12, "HR", "", ""])
    p = tmp_path / "x.xlsx"
    wb.save(str(p))
    with pytest.raises(ValueError, match="F1=NONE is reserved"):
        export_xlsx.stamp_workbook(p, "ASM", out_dir=tmp_path)


def test_sheet_natural_order():
    assert export_xlsx._sheet_key("S2.1") < export_xlsx._sheet_key(
        "S10.1")
