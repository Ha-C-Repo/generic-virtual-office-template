"""Tests for takeoff_pipeline/shop_log.py (PC3, Prompt 11).

All expected values are hand-computed from the fixture entries so the
rollup arithmetic is verified, not trusted. Entries are written through
the module's own writers so the P17 guards and the SF-in-pieces_done
convention are exercised on the same path the CLI uses. The real data/
stores are never touched.

Week fixture: week ending Sunday 2026-06-14, so the window is
2026-06-08 (Monday) through 2026-06-14 inclusive.
"""

import csv
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from takeoff_pipeline.shop_log import (
    _week_window,
    build_daily_sheet,
    ensure_schema,
    log_milestone,
    log_production,
    weekly_rollup,
    write_rollup_csv,
)

WEEK_ENDING = "2026-06-14"
# Synthetic id on purpose: it must match no real folder under
# "Awarded Projects/", so the writers' best-effort baseline discovery
# stays inert in tests.
PROJECT = "NC-TEST-PC3-001"

WBS_HEADER = ["WBS Line", "Scope", "Cost Code", "Progress Type",
              "Planned Units", "Unit", "Planned Hours", "Budget Cost",
              "Start", "End"]

WBS_ROWS = [
    ["FAB-S1", "Fab sequence 1", "fab", "production", 100, "tons",
     1000, 375000, date(2026, 6, 1), date(2026, 7, 10)],
    ["SD-01", "Shop drawings", "shop drawings", "milestone", "", "",
     200, 20000, date(2026, 6, 1), date(2026, 6, 20)],
    ["DECK-R1", "Roof deck", "roof deck", "production", 5000, "SF",
     200, 30000, date(2026, 6, 8), date(2026, 6, 30)],
    ["ERECT-S1", "Erection sequence 1", "erection", "production", 80,
     "tons", 600, 77600, date(2026, 6, 15), date(2026, 7, 20)],
]


def _write_baseline(path: Path, flag_text="BASELINE") -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    budget = wb.active
    budget.title = "Budget"
    budget.append(["Asian City Plaza cost baseline"])
    if flag_text:
        budget.append([flag_text, date(2026, 5, 1)])
    wbs = wb.create_sheet("WBS")
    wbs.append(WBS_HEADER)
    for row in WBS_ROWS:
        wbs.append(row)
    wb.save(str(path))
    return path


def _seed_entries(db: Path):
    """Hand-computable fixture entries, written through the real writers.

    FAB-S1 (tons): 2026-06-03 40 h 4 t (before week); 2026-06-09 50 h
    5.5 t; 2026-06-14 10 h 1 t (Sunday boundary, in week); 2026-06-15
    99 h 9 t (after week, excluded).
      to date: 10.5 t, 100 h.  week: 6.5 t, 60 h.

    SD-01 (milestone): 2026-06-02 credit 20, 30 h; 2026-06-10 credit
    75, 20 h.  to date credit 75, week advance 55, hours 50 / 20.

    DECK-R1 (SF): 2026-06-08 (Monday boundary) 16 h, 1200 SF.

    EXTRA-1 (not in baseline): 2026-06-09 8 h, 12 pieces.
    """
    rows = [
        ("Mario", "FAB-S1", {"hours": 40, "tons": 4,
                             "entry_date": "2026-06-03"}),
        ("Mario", "FAB-S1", {"hours": 50, "tons": 5.5,
                             "entry_date": "2026-06-09"}),
        ("Mario", "FAB-S1", {"hours": 10, "tons": 1,
                             "entry_date": "2026-06-14"}),
        ("Mario", "FAB-S1", {"hours": 99, "tons": 9,
                             "entry_date": "2026-06-15"}),
        ("Mario", "DECK-R1", {"hours": 16, "sf": 1200,
                              "entry_date": "2026-06-08"}),
        ("Temp", "EXTRA-1", {"hours": 8, "pieces": 12,
                             "entry_date": "2026-06-09"}),
    ]
    for person, wbs, kw in rows:
        r = log_production(person, wbs, project=PROJECT, db_path=db, **kw)
        assert "error" not in r, r
    for credit, hours, day in ((20, 30, "2026-06-02"), (75, 20,
                                                        "2026-06-10")):
        r = log_milestone("Joseph", "SD-01", credit, hours=hours,
                          entry_date=day, project=PROJECT, db_path=db)
        assert "error" not in r, r


@pytest.fixture
def sources(tmp_path):
    db = tmp_path / "shop_floor.db"
    ensure_schema(db)
    _seed_entries(db)
    return {"db_path": db,
            "baseline_path": _write_baseline(tmp_path / "budget.xlsx")}


def _by_line(roll):
    return {r["wbs_line"]: r for r in roll["rows"]}


# -- week window -------------------------------------------------------------


def test_week_window_explicit_end():
    start, end = _week_window(WEEK_ENDING)
    assert start == date(2026, 6, 8)
    assert end == date(2026, 6, 14)


# -- rollup math -------------------------------------------------------------


def test_production_rollup_math(sources):
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    fab = _by_line(roll)["FAB-S1"]
    assert fab["units_done_todate"] == 10.5
    assert fab["units_done_week"] == 6.5
    assert fab["hours_todate"] == 100.0
    assert fab["hours_week"] == 60.0
    assert fab["planned_units"] == 100
    assert fab["planned_hours"] == 1000
    assert fab["pct_units_complete"] == 0.105
    assert fab["pct_hours_used"] == 0.1
    assert fab["unit"] == "tons"
    assert fab["in_baseline"] is True


def test_week_boundaries(sources):
    """Sunday week end is inclusive, the Monday after is out, the Monday
    start is in."""
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    lines = _by_line(roll)
    # FAB-S1 week sum includes the 06-14 Sunday entry (1 t, 10 h) and
    # excludes the 06-15 entry (9 t, 99 h): hand-computed above.
    assert lines["FAB-S1"]["units_done_week"] == 6.5
    # DECK-R1 entry sits exactly on the Monday window start.
    assert lines["DECK-R1"]["units_done_week"] == 1200.0
    assert lines["DECK-R1"]["hours_week"] == 16.0


def test_milestone_credit_is_level_not_flow(sources):
    """To-date credit is the highest step reached, not a sum; the week
    column is the advance inside the week (75 - 20 = 55)."""
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    sd = _by_line(roll)["SD-01"]
    assert sd["units_done_todate"] == 75.0
    assert sd["units_done_week"] == 55.0
    assert sd["pct_units_complete"] == 0.75
    assert sd["planned_units"] == 100.0
    assert sd["unit"] == "credit"
    assert sd["hours_todate"] == 50.0
    assert sd["hours_week"] == 20.0
    assert sd["pct_hours_used"] == 0.25


def test_sf_line_uses_pieces_done_with_note(sources):
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    deck = _by_line(roll)["DECK-R1"]
    assert deck["units_done_todate"] == 1200.0
    assert deck["pct_units_complete"] == 0.24
    assert "pieces_done" in deck["note"]


def test_baseline_line_with_no_progress_rolls_zero(sources):
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    erect = _by_line(roll)["ERECT-S1"]
    assert erect["units_done_todate"] == 0.0
    assert erect["hours_todate"] == 0.0
    assert erect["pct_units_complete"] == 0.0
    assert erect["planned_units"] == 80


def test_extra_line_not_in_baseline(sources):
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    extra = _by_line(roll)["EXTRA-1"]
    assert extra["in_baseline"] is False
    assert extra["planned_units"] is None
    assert extra["pct_units_complete"] is None
    assert extra["units_done_week"] == 12.0
    assert extra["unit"] == "pieces"
    assert "not in baseline" in extra["note"]


def test_rollup_rejects_bad_week_ending(sources):
    roll = weekly_rollup(week_ending="June ninth",
                         db_path=sources["db_path"])
    assert "error" in roll


def test_rollup_warns_on_multi_project_blend(sources):
    r = log_production("Mario", "FAB-S1", hours=5, tons=2,
                       entry_date="2026-06-09", project="OTHER-JOB",
                       db_path=sources["db_path"])
    assert "error" not in r
    roll = weekly_rollup(week_ending=WEEK_ENDING,
                         db_path=sources["db_path"])
    assert any("blended" in w for w in roll["warnings"])


def test_db_path_agrees_with_pc4_reader():
    from bridge.project_controls import _progress_db_path
    from takeoff_pipeline.shop_log import _db_path
    assert _db_path() == _progress_db_path()


def test_rollup_without_baseline(sources):
    roll = weekly_rollup(week_ending=WEEK_ENDING,
                         db_path=sources["db_path"])
    assert any("no PC1 baseline" in w for w in roll["warnings"])
    fab = _by_line(roll)["FAB-S1"]
    assert fab["planned_units"] is None
    assert fab["in_baseline"] is False
    # Without a baseline the unit is inferred from what was logged.
    assert fab["unit"] == "tons"
    assert fab["units_done_todate"] == 10.5
    assert fab["hours_todate"] == 100.0


def test_unfrozen_baseline_marked_draft(sources, tmp_path):
    draft = _write_baseline(tmp_path / "draft.xlsx", flag_text="")
    roll = weekly_rollup(week_ending=WEEK_ENDING,
                         db_path=sources["db_path"], baseline_path=draft)
    assert any("draft" in w for w in roll["warnings"])
    # Planned values still report; the rollup is a report, not a claim.
    assert _by_line(roll)["FAB-S1"]["planned_units"] == 100


def test_project_filter_excludes_other_projects(sources):
    r = log_production("Mario", "FAB-S1", hours=5, tons=2,
                       entry_date="2026-06-09", project="OTHER-JOB",
                       db_path=sources["db_path"])
    assert "error" not in r
    roll = weekly_rollup(week_ending=WEEK_ENDING, project=PROJECT,
                         db_path=sources["db_path"],
                         baseline_path=sources["baseline_path"])
    assert _by_line(roll)["FAB-S1"]["units_done_todate"] == 10.5


# -- CSV export ---------------------------------------------------------------


def test_csv_export(sources, tmp_path):
    roll = weekly_rollup(week_ending=WEEK_ENDING, **sources)
    out = write_rollup_csv(roll, tmp_path / "rollup.csv")
    with open(out, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0][:4] == ["week_start", "2026-06-08",
                           "week_ending", "2026-06-14"]
    assert rows[1][0] == "wbs_line"
    body = {r[0]: r for r in rows[2:]}
    header = rows[1]
    fab = dict(zip(header, body["FAB-S1"]))
    assert fab["units_done_todate"] == "10.5"
    assert fab["hours_week"] == "60.0"
    extra = dict(zip(header, body["EXTRA-1"]))
    # None never becomes 0 in the CSV: blank means no plan, not zero plan.
    assert extra["planned_units"] == ""
    assert extra["pct_units_complete"] == ""


# -- write-layer guards (P17) -------------------------------------------------


def test_milestone_credit_validation(tmp_path):
    db = tmp_path / "s.db"
    r = log_milestone("Joseph", "SD-01", 50, entry_date="2026-06-09",
                      db_path=db)
    assert "error" in r
    assert "rule-of-credit" in r["error"]


def test_p17_mixing_guard_both_directions(tmp_path):
    db = tmp_path / "s.db"
    ok = log_production("Mario", "FAB-S1", hours=8, tons=2,
                        entry_date="2026-06-09", project="JOB-T",
                        db_path=db)
    assert "error" not in ok
    mixed = log_milestone("Joseph", "FAB-S1", 20, entry_date="2026-06-09",
                          project="JOB-T", db_path=db)
    assert "error" in mixed and "P17" in mixed["error"]

    ok = log_milestone("Joseph", "SD-01", 20, entry_date="2026-06-09",
                       project="JOB-T", db_path=db)
    assert "error" not in ok
    mixed = log_production("Mario", "SD-01", hours=8, pieces=3,
                           entry_date="2026-06-09", project="JOB-T",
                           db_path=db)
    assert "error" in mixed and "P17" in mixed["error"]


def test_p17_guard_is_project_scoped(tmp_path):
    """The P15 WBS template reuses line ids across projects; JOB-A's
    production history on a line must not reject JOB-B's milestone use
    of the same id."""
    db = tmp_path / "s.db"
    ok = log_production("Mario", "MISC-01", hours=8, pieces=2,
                        entry_date="2026-06-09", project="JOB-A",
                        db_path=db)
    assert "error" not in ok
    other = log_milestone("Joseph", "MISC-01", 20, entry_date="2026-06-09",
                          project="JOB-B", db_path=db)
    assert "error" not in other
    same = log_milestone("Joseph", "MISC-01", 20, entry_date="2026-06-09",
                         project="JOB-A", db_path=db)
    assert "error" in same and "P17" in same["error"]


def test_milestone_prior_credit_is_project_scoped(tmp_path):
    db = tmp_path / "s.db"
    log_milestone("Joseph", "SD-01", 75, entry_date="2026-06-09",
                  project="JOB-A", db_path=db)
    fresh = log_milestone("Joseph", "SD-01", 20, entry_date="2026-06-10",
                          project="JOB-B", db_path=db)
    assert "error" not in fresh
    assert "note" not in fresh


def test_pieces_and_sf_collision_rejected(tmp_path):
    r = log_production("Mario", "DECK-R1", hours=8, pieces=10, sf=500,
                       entry_date="2026-06-09", db_path=tmp_path / "s.db")
    assert "error" in r


def test_negative_and_empty_entries_rejected(tmp_path):
    db = tmp_path / "s.db"
    assert "error" in log_production("Mario", "FAB-S1", hours=-1,
                                     entry_date="2026-06-09", db_path=db)
    assert "error" in log_production("Mario", "FAB-S1", tons=-2,
                                     entry_date="2026-06-09", db_path=db)
    assert "error" in log_production("Mario", "FAB-S1",
                                     entry_date="2026-06-09", db_path=db)
    assert "error" in log_production("", "FAB-S1", hours=8,
                                     entry_date="2026-06-09", db_path=db)
    assert "error" in log_production("Mario", "FAB-S1", hours=8,
                                     entry_date="June ninth", db_path=db)


def test_credit_regression_noted_not_rejected(tmp_path):
    db = tmp_path / "s.db"
    log_milestone("Joseph", "SD-01", 75, entry_date="2026-06-09",
                  project="JOB-T", db_path=db)
    r = log_milestone("Joseph", "SD-01", 20, entry_date="2026-06-10",
                      project="JOB-T", db_path=db)
    assert "error" not in r
    assert "highest credit" in r.get("note", "")


# -- project attribution ------------------------------------------------------


def test_entry_requires_project_or_sticky_default(tmp_path):
    """PC4 excludes unattributed rows from every metric, so an entry
    with no project and no default is rejected up front; the sticky
    default then carries attribution at five fields a day."""
    import sqlite3

    from takeoff_pipeline.shop_log import set_default_project

    db = tmp_path / "s.db"
    bare = log_production("Mario", "FAB-S1", hours=8, tons=2,
                          entry_date="2026-06-09", db_path=db)
    assert "error" in bare and "project" in bare["error"]

    assert "error" not in set_default_project(PROJECT, db_path=db)
    ok = log_production("Mario", "FAB-S1", hours=8, tons=2,
                        entry_date="2026-06-09", db_path=db)
    assert "error" not in ok
    c = sqlite3.connect(str(db))
    stored = c.execute("SELECT project FROM progress_log").fetchone()[0]
    c.close()
    assert stored == PROJECT


# -- baseline guard at entry time ----------------------------------------------


def test_baseline_type_guard_at_entry(sources, tmp_path):
    db = tmp_path / "guard.db"
    wrong = log_production("Mario", "SD-01", hours=4, pieces=1,
                           entry_date="2026-06-09", project="JOB-T",
                           db_path=db,
                           baseline_path=sources["baseline_path"])
    assert "error" in wrong and "P17" in wrong["error"]
    wrong = log_milestone("Joseph", "FAB-S1", 20, entry_date="2026-06-09",
                          project="JOB-T", db_path=db,
                          baseline_path=sources["baseline_path"])
    assert "error" in wrong and "P17" in wrong["error"]


def test_unit_mismatch_noted_at_entry(sources, tmp_path):
    """Pieces on a tons line are stored but PC4 will not read them;
    the entry says so instead of staying silent."""
    db = tmp_path / "guard.db"
    r = log_production("Mario", "FAB-S1", hours=8, pieces=5,
                       entry_date="2026-06-09", project="JOB-T",
                       db_path=db,
                       baseline_path=sources["baseline_path"])
    assert "error" not in r
    assert "tons_done" in r.get("note", "")


# -- PC4 compatibility --------------------------------------------------------


def test_pc4_reads_shop_log_output(sources):
    """End to end: rows written by the PC3 writers feed spi_cpi without
    error and with the unit conventions intact (tons from tons_done,
    credit from pieces_done)."""
    from bridge.project_controls import spi_cpi
    r = spi_cpi(PROJECT, db_path=sources["db_path"],
                baseline_path=sources["baseline_path"],
                as_of=WEEK_ENDING)
    assert "error" not in r, r
    lines = {m["wbs_line"]: m for m in r["lines"]}
    assert lines["FAB-S1"]["units_done"] == 10.5
    assert lines["FAB-S1"]["actual_hours"] == 100.0
    assert lines["SD-01"]["units_done"] == 75.0
    assert lines["DECK-R1"]["units_done"] == 1200.0


# -- daily sheet ---------------------------------------------------------------


def test_daily_sheet_sections_never_mix(sources, tmp_path):
    out = tmp_path / "sheet.xlsx"
    r = build_daily_sheet(sheet_date="2026-06-12", out_path=out,
                          baseline_path=sources["baseline_path"])
    assert "error" not in r, r
    assert Path(r["path"]).exists()
    assert r["prefilled_lines"] == 4

    import openpyxl
    ws = openpyxl.load_workbook(str(out)).active
    text = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    joined = "\n".join(text)
    prod_at = joined.index("SECTION A - PRODUCTION")
    mile_at = joined.index("SECTION B - MILESTONES")
    assert prod_at < mile_at
    # Production lines land in section A, milestone lines in section B.
    assert prod_at < joined.index("FAB-S1") < mile_at
    assert joined.index("SD-01") > mile_at
    # Rule of credit is printed for the shop.
    assert "20" in joined and "75" in joined and "100" in joined


def test_daily_sheet_without_baseline(tmp_path):
    r = build_daily_sheet(sheet_date="2026-06-12",
                          out_path=tmp_path / "blank.xlsx")
    assert "error" not in r, r
    assert r["prefilled_lines"] == 0
