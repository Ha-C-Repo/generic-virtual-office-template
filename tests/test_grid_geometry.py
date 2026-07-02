"""Tests for grid_geometry.py (Engine B, the SF calculator).

Pure-function tests pin the mirrored-pair detection and the AREA-row
shape with stub pages, so they never depend on a PDF. Two SP183
regressions run against the live set: building and deck SF must land
within the 2 percent (value) and 15 percent (value_approx) gates of
Ivan's verified 61,621 SF, and the emitted AREA rows must pass the
section 14 validator. The SP183 set has no printed SF figure anywhere,
so this is the only path to those quantities.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from takeoff_pipeline import census, export_xlsx, grid_geometry, \
    validate_takeoff

REPO = Path(__file__).resolve().parent.parent
SP183_PDF = (REPO / "Bids To Estimate" / "06 South Park 183" / "drawings"
             / "Extracted pages from 2026-04-22 Issue For Pricing"
             " - Building 1.pdf")
SP183_VERIFIED_SF = 61621


class _StubPage:
    """Minimal page exposing get_text('words'). words are
    (x0, y0, x1, y1, text, ...) like PyMuPDF."""

    def __init__(self, words):
        self._words = words

    def get_text(self, kind):
        assert kind == "words"
        return list(self._words)


def _w(x0, y0, text, w=38, h=20):
    return (x0, y0, x0 + w, y0 + h, text, 0, 0, 0)


# -- mirrored-pair detection -------------------------------------------------

def test_mirrored_pairs_find_both_axes():
    # 413' printed above and below the plan (shared X, separated Y);
    # 150' printed left and right (shared Y, separated X); a smaller
    # 60' bay dimension mirrored vertically must not beat 150'.
    page = _StubPage([
        _w(1080, 180, "413'"), _w(1080, 1110, "413'"),
        _w(140, 575, "150'"), _w(2010, 575, "150'"),
        _w(170, 740, "60'"), _w(2000, 740, "60'"),
    ])
    dims = grid_geometry.find_overall_dimensions(page)
    assert dims["length_ft"] == 413
    assert dims["width_ft"] == 150
    assert dims["rectangular"] is True
    assert dims["flags"] == []


def test_single_axis_falls_back_to_low_confidence():
    # 413' mirrored horizontally, but 150' appears only once: the width
    # axis has no mirrored pair, so it falls back and the result is
    # flagged non-rectangular, never silently treated as clean.
    page = _StubPage([
        _w(1080, 180, "413'"), _w(1080, 1110, "413'"),
        _w(140, 575, "150'"),
    ])
    dims = grid_geometry.find_overall_dimensions(page)
    assert dims["length_ft"] == 413
    assert dims["width_ft"] == 150
    assert dims["rectangular"] is False
    assert any("width axis" in f for f in dims["flags"])


def test_too_few_overall_tokens_returns_none():
    assert grid_geometry.find_overall_dimensions(
        _StubPage([_w(1080, 180, "413'")])) is None
    # Sub-floor tokens (bay dims, detail dims) are not overall.
    assert grid_geometry.find_overall_dimensions(
        _StubPage([_w(10, 10, "12'"), _w(20, 20, "8'")])) is None


# -- AREA-row shape ----------------------------------------------------------

def _measurement(area=61950):
    box = {"length_ft": 413, "width_ft": 150, "area_sf": area,
           "confidence": "medium", "rectangular": True,
           "bbox": "[10.0, 20.0, 30.0, 40.0]", "flags": []}
    return {
        "deck": {**box, "sheet": "S3.00",
                 "primary_source": "S3.00 ROOF FRAMING PLAN"},
        "building": {**box, "sheet": "S2.00",
                     "primary_source": "S2.00 FOUNDATION PLAN"},
        "cross_check": {"status": "agree", "diff_pct": 0.0},
    }


def test_to_takeoff_rows_shape():
    rows = grid_geometry.to_takeoff_rows(_measurement())
    by_class = {r["item_class"]: r for r in rows}
    deck = by_class["DECK"]
    assert deck["designation"] == "ROOF DECK"
    assert deck["mode"] == "AREA" and deck["unit"] == "SF"
    assert deck["qty"] == 61950
    # Section 5: the DECK primary_source must carry FRAMING PLAN.
    assert "FRAMING PLAN" in deck["primary_source"].upper()
    building = by_class["MISC"]
    assert building["designation"] == "BUILDING GROSS SF"
    assert building["mode"] == "AREA" and building["unit"] == "SF"


def test_non_rectangular_row_is_flagged_and_low():
    m = _measurement()
    m["deck"]["rectangular"] = False
    m["deck"]["confidence"] = "low"
    rows = grid_geometry.to_takeoff_rows(m)
    deck = next(r for r in rows if r["item_class"] == "DECK")
    assert deck["confidence"] == "low"
    assert "non-rectangular: verify" in deck["notes"]


# -- validator accepts the AREA rows -----------------------------------------

def _empty_census(tmp_path):
    db = tmp_path / "census.db"
    census.init_db(db)
    return db


def test_area_rows_pass_validator(tmp_path):
    """The mode fix plus the extra_rows merge: AREA rows reach the
    workbook carrying AREA, and the section 14 gate passes them."""
    rows = grid_geometry.to_takeoff_rows(_measurement())
    db = _empty_census(tmp_path)
    out = tmp_path / "out"
    info = export_xlsx.export("SF", db_path=db, out_dir=out,
                              extra_rows=rows)
    assert info["status"] == "UNSTAMPED_VALID", info
    assert info["grid_rows"] == 2

    import openpyxl
    wb = openpyxl.load_workbook(info["unstamped"])
    t = wb["TAKEOFF"]
    modes = {t.cell(row=r, column=2).value: t.cell(row=r, column=3).value
             for r in range(3, t.max_row + 1)}
    assert modes["ROOF DECK"] == "AREA"
    assert modes["BUILDING GROSS SF"] == "AREA"
    report = validate_takeoff.validate_file(info["unstamped"])
    assert report["hard_fails"] == []


# -- SP183 regression --------------------------------------------------------

def test_deck_source_conforms_to_section_5():
    # A framing sheet whose title is not the contiguous phrase still
    # yields a section-5-conforming DECK primary_source, so the row is
    # never blocked by the validator (Finding 1).
    assert "FRAMING PLAN" in grid_geometry._conforming_deck_source(
        "S3.00", "S3.00 ROOF FRAMING - LOW ROOF PLAN").upper()
    # An already-conforming source is left intact.
    assert grid_geometry._conforming_deck_source(
        "S3.00", "S3.00 ROOF FRAMING PLAN") == "S3.00 ROOF FRAMING PLAN"


def test_building_sf_not_answered_with_deck_footprint(tmp_path):
    # Finding 2: a missing building footprint must never borrow the
    # deck area (a silent wrong-source number at 2 pct tolerance).
    from takeoff_pipeline import score_spike

    db = _empty_census(tmp_path)
    sheet_ctx = {"foundation_sheets": [], "plan_sheets": [],
                 "notes_sheets": []}
    grid = {"deck": {"area_sf": 61950, "length_ft": 413, "width_ft": 150,
                     "sheet": "S3.00", "confidence": "medium"},
            "building": None}
    building = score_spike._match_row(
        {"designation": "BUILDING SF", "item_class": "area"}, [], "J",
        db, sheet_ctx, grid)
    assert building["count"] != 61950 and not building["found"]
    deck = score_spike._match_row(
        {"designation": "DECK SF", "item_class": "deck"}, [], "J", db,
        sheet_ctx, grid)
    assert deck["found"] and deck["count"] == 61950


@pytest.mark.skipif(not SP183_PDF.exists(), reason="SP183 set not present")
def test_sp183_building_and_deck_sf_within_tolerance():
    m = grid_geometry.measure(SP183_PDF)
    assert m["building"] and m["deck"], m
    gate = {"building": 2.0, "deck": 15.0}
    for sec, tol in gate.items():
        area = m[sec]["area_sf"]
        diff = abs(area - SP183_VERIFIED_SF) / SP183_VERIFIED_SF * 100.0
        assert diff <= tol, (sec, area, f"{diff:.2f} pct")
    assert "FRAMING PLAN" in m["deck"]["primary_source"].upper()


@pytest.mark.skipif(not SP183_PDF.exists(), reason="SP183 set not present")
def test_sp183_area_rows_validate(tmp_path):
    rows = grid_geometry.to_takeoff_rows(grid_geometry.measure(SP183_PDF))
    assert len(rows) == 2
    db = _empty_census(tmp_path)
    info = export_xlsx.export("SP183", db_path=db,
                              out_dir=tmp_path / "out", extra_rows=rows)
    assert info["status"] == "UNSTAMPED_VALID", info
