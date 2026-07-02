"""Tests for apply_assemblies.py (Prompt 9, assembly library v2).

Covers the two mandated paths plus the governance edges:
- openings deduct-and-add per F4 (over-threshold openings deducted
  from the deck area AND generating framed steel adds; at-or-under
  ignored but reported)
- the rounding rules (round UP to declared increments, waste before
  rounding, exact multiples stay put, float artifacts do not buy an
  extra increment)
- measured-class cross-check rule (section 9), GROUP suppression
  (3.3), subcontract stream emission, pricing-token rejection at
  definition load (P10), stamped-file refusal (13.4), ambiguity and
  unassigned reporting, and an end-to-end run with the SHIPPED
  definitions that must leave validate_takeoff.py with zero hard
  fails.

Unit-level math tests pin SYNTHETIC definitions with clean numbers so
seed-factor diffs in the shipped library never break them; the
end-to-end test asserts structure, not factor values.
"""

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from takeoff_pipeline import apply_assemblies, validate_takeoff
from takeoff_pipeline.export_xlsx import (
    BOM_HEADERS, BOQ_HEADERS, PS_HEADERS, TAKEOFF_HEADERS)


# -- builders ----------------------------------------------------------------

def _workbook(rows):
    """A four-sheet workbook. rows: list of 11-value TAKEOFF tuples,
    optionally followed by (weight_formula, tons_formula, formula_ref)
    handled by the caller via cell writes."""
    import openpyxl

    wb = openpyxl.Workbook()
    t = wb.active
    t.title = "TAKEOFF"
    t.append([])  # metadata row, written only by the stamp
    t.append(list(TAKEOFF_HEADERS))
    for row in rows:
        t.append(list(row))
    boq = wb.create_sheet("BOQ")
    boq.append(list(BOQ_HEADERS))
    bom = wb.create_sheet("BOM")
    bom.append(list(BOM_HEADERS))
    ps = wb.create_sheet("PRICING_SCHEDULE")
    ps.append(list(PS_HEADERS))
    return wb


def _row(item_id, designation, mode, qty, unit, primary, notes=""):
    return (item_id, designation, mode, qty, unit, primary, "",
            "medium", "S3.00", "MANUAL", notes)


def _read_sheet(path, name, headers):
    import openpyxl

    ws = openpyxl.load_workbook(str(path))[name]
    out = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value
                  for c in range(1, len(headers) + 1)]
        if all(v is None for v in values):
            continue
        out.append(dict(zip(headers, values)))
    return out


def _by_ref(rows, fragment):
    return [r for r in rows if fragment in (r["formula_ref"] or "")]


_DECK_DEF = {
    "schema": "ASSEMBLY_DEF_V2",
    "assembly_id": "deck_test",
    "description": "synthetic roof deck for math pinning",
    "provenance": "TEST",
    "applies_to": {"item_class": "DECK", "mode": "AREA",
                   "designation_includes": ["TEST DECK"],
                   "designation_excludes": ["OPENING"]},
    "openings": {
        "deduct_threshold_sf": 16.0,
        "geometry_model": "square",
        "framed_adds": [
            {"component_id": "frame_angle",
             "description": "opening perimeter frame angle",
             "designation": "L3X3X1/4", "basis": "perimeter",
             "pieces_per_opening": 1, "unit": "LF",
             "waste_factor": 0.05,
             "round": {"direction": "up", "increment": 1.0},
             "aisc": True},
            {"component_id": "header_angle",
             "description": "opening header angle",
             "designation": "L4X4X1/4", "basis": "span",
             "pieces_per_opening": 2, "unit": "LF",
             "waste_factor": 0.05,
             "round": {"direction": "up", "increment": 1.0},
             "aisc": True},
        ],
        "labor_hours": [
            {"component_id": "opening_frame_install",
             "description": "fit and install opening frame",
             "hours_per_opening": 2.0,
             "round": {"direction": "up", "increment": 0.25}},
        ],
    },
    "streams": {
        "materials": [
            {"component_id": "deck_sheets",
             "description": "deck sheets, 100 SF per sheet",
             "designation": None, "designation_from_driver": True,
             "unit": "EA", "per_sheet_sf": 100.0,
             "waste_factor": 0.05,
             "round": {"direction": "up", "increment": 1.0},
             "cross_check_class": None, "length_lf": None,
             "aisc": False},
            {"component_id": "deck_fasteners",
             "description": "deck fastening per net SF",
             "designation": "DECK FASTENERS",
             "designation_from_driver": False,
             "unit": "EA", "qty_per_unit": 0.5,
             "waste_factor": 0.05,
             "round": {"direction": "up", "increment": 50.0},
             "cross_check_class": None, "length_lf": None,
             "aisc": False},
        ],
        "labor_hours": [
            {"component_id": "deck_install",
             "description": "deck placement hours per net SF",
             "hours_per_unit": 0.01,
             "round": {"direction": "up", "increment": 0.25}},
        ],
        "equipment_time": [
            {"component_id": "crane_bundles",
             "description": "crane minutes per net SF",
             "unit": "MIN", "time_per_unit": 0.005,
             "round": {"direction": "up", "increment": 5.0}},
        ],
        "subcontract": None,
    },
    "notes": "test",
}

_COL_DEF = {
    "schema": "ASSEMBLY_DEF_V2",
    "assembly_id": "col_test",
    "description": "synthetic column base for cross-check pinning",
    "provenance": "TEST",
    "applies_to": {"item_class": "COL", "mode": "COUNT",
                   "designation_includes": [],
                   "designation_excludes": []},
    "openings": None,
    "streams": {
        "materials": [
            {"component_id": "anchor_rods",
             "description": "anchor rods, four per column",
             "designation": "ANCHOR ROD",
             "designation_from_driver": False,
             "unit": "EA", "qty_per_unit": 4.0, "waste_factor": 0.0,
             "round": {"direction": "up", "increment": 1.0},
             "cross_check_class": "ANCH", "length_lf": None,
             "aisc": False},
            {"component_id": "base_plate",
             "description": "column base plate, one per column",
             "designation": "BASE PL",
             "designation_from_driver": False,
             "unit": "EA", "qty_per_unit": 1.0, "waste_factor": 0.0,
             "round": {"direction": "up", "increment": 1.0},
             "cross_check_class": "PLATE", "length_lf": None,
             "aisc": False},
        ],
        "labor_hours": [
            {"component_id": "setting_labor",
             "description": "column setting hours per column",
             "hours_per_unit": 1.0,
             "round": {"direction": "up", "increment": 0.25}},
        ],
        "equipment_time": [
            {"component_id": "crane_set",
             "description": "crane minutes per column set",
             "unit": "MIN", "time_per_unit": 20.0,
             "round": {"direction": "up", "increment": 5.0}},
        ],
        "subcontract": None,
    },
    "notes": "test",
}

_BEAM_DEF = {
    "schema": "ASSEMBLY_DEF_V2",
    "assembly_id": "beam_test",
    "description": "synthetic beam sticks for rounding pinning",
    "provenance": "TEST",
    "applies_to": {"item_class": "BEAM", "mode": "LINEAR",
                   "designation_includes": [],
                   "designation_excludes": []},
    "openings": None,
    "streams": {
        "materials": [
            {"component_id": "stick_material",
             "description": "beam stick LF after cut allowance",
             "designation": None, "designation_from_driver": True,
             "unit": "LF", "qty_per_unit": 1.0, "waste_factor": 0.05,
             "round": {"direction": "up", "increment": 5.0},
             "cross_check_class": None, "length_lf": None,
             "aisc": True},
        ],
        "labor_hours": [
            {"component_id": "fab_labor",
             "description": "shop hours per measured LF",
             "hours_per_unit": 0.1,
             "round": {"direction": "up", "increment": 0.25}},
        ],
        "equipment_time": [
            {"component_id": "crane_erect",
             "description": "crane minutes per measured LF",
             "unit": "MIN", "time_per_unit": 0.4,
             "round": {"direction": "up", "increment": 5.0}},
        ],
        "subcontract": None,
    },
    "notes": "test",
}

_WIDGET_DEF = {
    "schema": "ASSEMBLY_DEF_V2",
    "assembly_id": "widget_count",
    "description": "synthetic assembly exercising the subcontract "
                   "stream",
    "provenance": "TEST",
    "applies_to": {"item_class": "MISC", "mode": "COUNT",
                   "designation_includes": ["WIDGET"],
                   "designation_excludes": []},
    "openings": None,
    "streams": {
        "materials": [],
        "labor_hours": [],
        "equipment_time": [],
        "subcontract": [
            {"component_id": "galv",
             "description": "hot-dip galvanizing, scope context",
             "designation": "GALVANIZING",
             "designation_from_driver": False,
             "unit": "EA", "qty_per_unit": 1.0, "waste_factor": 0.0,
             "round": {"direction": "up", "increment": 1.0},
             "cross_check_class": None, "length_lf": None,
             "aisc": False},
        ],
    },
    "notes": "test",
}


@pytest.fixture
def asm_dir(tmp_path):
    d = tmp_path / "asm"
    d.mkdir()
    for definition in (_DECK_DEF, _COL_DEF, _BEAM_DEF, _WIDGET_DEF):
        (d / f"{definition['assembly_id']}.json").write_text(
            json.dumps(definition), encoding="utf-8")
    return d


def _apply(tmp_path, asm_dir, rows):
    path = tmp_path / "takeoff.xlsx"
    _workbook(rows).save(str(path))
    report = apply_assemblies.apply_to_workbook(path, asm_dir=asm_dir)
    return (path, report,
            _read_sheet(path, "BOQ", BOQ_HEADERS),
            _read_sheet(path, "BOM", BOM_HEADERS))


# -- rounding rules ----------------------------------------------------------

def test_round_up_rounds_up_to_increment():
    assert apply_assemblies.round_up(9.828, 1) == 10
    assert apply_assemblies.round_up(491.4, 50) == 500
    assert apply_assemblies.round_up(24.1, 0.25) == 24.25


def test_round_up_exact_multiple_stays_put():
    assert apply_assemblies.round_up(40, 5) == 40
    assert apply_assemblies.round_up(210.0, 5) == 210
    assert apply_assemblies.round_up(24.0, 0.25) == 24


def test_round_up_never_rounds_down():
    assert apply_assemblies.round_up(40.01, 5) == 45
    assert apply_assemblies.round_up(0.001, 1) == 1


def test_round_up_float_artifact_buys_no_increment():
    # 0.1 + 0.2 is 0.30000000000000004; three 0.1 steps, not four.
    assert apply_assemblies.round_up(0.1 + 0.2, 0.1) == 0.3
    assert apply_assemblies.round_up(2.0000000000000004, 1) == 2


def test_round_up_rejects_bad_inputs():
    with pytest.raises(ValueError):
        apply_assemblies.round_up(1.0, 0)
    with pytest.raises(ValueError):
        apply_assemblies.round_up(-1.0, 1)


def test_waste_applies_before_rounding(tmp_path, asm_dir):
    # 240 LF x 1.05 = 252, stick increment 5 rounds UP to 255.
    # 200 LF x 1.05 = 210, already a multiple of 5, stays 210.
    _, _, _, bom = _apply(tmp_path, asm_dir, [
        _row("BEAM-001", "W12X26", "LINEAR", 240, "LF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("BEAM-002", "W16X31", "LINEAR", 200, "LF",
             "S3.00 ROOF FRAMING PLAN"),
    ])
    sticks = {r["source_item_id"]: r["qty"]
              for r in _by_ref(bom, ":stick_material:")}
    assert sticks == {"BEAM-001": 255, "BEAM-002": 210}


def test_labor_and_equipment_rounding(tmp_path, asm_dir):
    # 241 LF x 0.1 HR = 24.1 rounds UP to 24.25 HR.
    # 241 LF x 0.4 MIN = 96.4 rounds UP to 100 MIN.
    _, _, boq, _ = _apply(tmp_path, asm_dir, [
        _row("BEAM-001", "W12X26", "LINEAR", 241, "LF",
             "S3.00 ROOF FRAMING PLAN"),
    ])
    labor = _by_ref(boq, ":fab_labor:")
    crane = _by_ref(boq, ":crane_erect:")
    assert labor[0]["qty"] == 24.25 and labor[0]["unit"] == "HR"
    assert crane[0]["qty"] == 100 and crane[0]["unit"] == "MIN"


# -- openings deduct-and-add (F4) --------------------------------------------

def _deck_rows():
    return [
        _row("DECK-001", "TEST DECK B", "AREA", 1000, "SF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("DECK-002", "OPENING", "AREA", 64, "SF",
             "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001 RTU"),
        _row("DECK-003", "OPENING", "AREA", 9, "SF",
             "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001 small"),
    ]


def test_openings_deduct_from_deck_area(tmp_path, asm_dir):
    # Net area: 1000 - 64 (over threshold) = 936. The 9 SF opening is
    # at or under 16 SF and never deducts.
    # Sheets: 936 x 1.05 / 100 = 9.828, rounds UP to 10 whole sheets.
    # Fasteners: 936 x 0.5 x 1.05 = 491.4, rounds UP to 500.
    # Labor: 936 x 0.01 = 9.36, rounds UP to 9.5 HR.
    # Crane: 936 x 0.005 = 4.68, rounds UP to 5 MIN.
    _, report, boq, bom = _apply(tmp_path, asm_dir, _deck_rows())
    assert _by_ref(bom, ":deck_sheets:")[0]["qty"] == 10
    assert _by_ref(bom, ":deck_fasteners:")[0]["qty"] == 500
    assert _by_ref(boq, ":deck_install:")[0]["qty"] == 9.5
    assert _by_ref(boq, ":crane_bundles:")[0]["qty"] == 5
    assert report["openings"]["qualifying"] == 1
    assert report["openings"]["ignored_under_threshold"] == 1


def test_openings_generate_framed_steel_adds(tmp_path, asm_dir):
    # 64 SF square model: side 8 LF. Perimeter angle 32 x 1.05 = 33.6
    # rounds UP to 34 LF. Headers: 8 x 2 x 1.05 = 16.8 rounds UP to
    # 17 LF. Both cite the OPENING row, closing the F4 double miss.
    _, _, boq, bom = _apply(tmp_path, asm_dir, _deck_rows())
    frame = _by_ref(bom, ":frame_angle:")
    header = _by_ref(bom, ":header_angle:")
    assert frame[0]["qty"] == 34 and frame[0]["unit"] == "LF"
    assert frame[0]["source_item_id"] == "DECK-002"
    assert frame[0]["designation"] == "L3X3X1/4"
    assert header[0]["qty"] == 17
    assert header[0]["source_item_id"] == "DECK-002"
    install = _by_ref(boq, ":opening_frame_install:")
    assert install[0]["qty"] == 2 and install[0]["unit"] == "HR"
    assert install[0]["source_item_id"] == "DECK-002"


def test_under_threshold_opening_adds_nothing(tmp_path, asm_dir):
    path, report, boq, bom = _apply(tmp_path, asm_dir, _deck_rows())
    assert not [r for r in boq + bom
                if r["source_item_id"] == "DECK-003"]
    assert any("DECK-003" in line
               for line in report["openings"]["ignored_detail"])


def test_openings_exceeding_parent_clamp_to_zero(tmp_path, asm_dir):
    _, report, boq, bom = _apply(tmp_path, asm_dir, [
        _row("DECK-001", "TEST DECK B", "AREA", 50, "SF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("DECK-002", "OPENING", "AREA", 64, "SF",
             "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001 RTU"),
    ])
    assert not _by_ref(bom, ":deck_sheets:")
    assert any("clamped" in w for w in report["warnings"])
    # The framed adds still ride: the opening exists and needs steel.
    assert _by_ref(bom, ":frame_angle:")


@pytest.mark.parametrize("bad_first", [True, False])
def test_opening_citing_an_opening_warns_in_either_order(
        tmp_path, asm_dir, bad_first):
    # An OPENING citing another OPENING as its parent is malformed
    # takeoff data. The guard must fire in BOTH sheet orders; the
    # later-parent case used to slip through with no warning, against
    # the 3.4/F4 never-silent rule.
    bad = _row("DECK-002", "OPENING", "AREA", 25, "SF",
               "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-003")
    good = _row("DECK-003", "OPENING", "AREA", 64, "SF",
                "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001")
    rows = [_row("DECK-001", "TEST DECK B", "AREA", 1000, "SF",
                 "S3.00 ROOF FRAMING PLAN")]
    rows += [bad, good] if bad_first else [good, bad]
    _, report, _, _ = _apply(tmp_path, asm_dir, rows)
    assert any("DECK-002" in w and "not a deck AREA row" in w
               for w in report["warnings"])
    assert report["openings"]["qualifying"] == 1  # only DECK-003


def test_every_output_row_cites_a_takeoff_item(tmp_path, asm_dir):
    _, _, boq, bom = _apply(tmp_path, asm_dir, _deck_rows() + [
        _row("COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
             "S4.00 COLUMN SCHEDULE"),
        _row("BEAM-001", "W12X26", "LINEAR", 240, "LF",
             "S3.00 ROOF FRAMING PLAN"),
    ])
    ids = {"DECK-001", "DECK-002", "DECK-003", "COL-001", "BEAM-001"}
    for r in boq + bom:
        assert r["source_item_id"] in ids


# -- measured-class cross-check (section 9) ----------------------------------

def _col_rows(anch_qty=64):
    return [
        _row("COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
             "S4.00 COLUMN SCHEDULE"),
        _row("ANCH-001", "ANCHOR ROD 3/4", "COUNT", anch_qty, "EA",
             "S4.00 COLUMN SCHEDULE + TYPICAL DETAILS"),
    ]


def test_measured_class_emits_cross_check_not_order_line(
        tmp_path, asm_dir):
    _, report, boq, bom = _apply(tmp_path, asm_dir, _col_rows())
    # No anchor order line anywhere; the measured row is the system
    # of record for ordering (P30).
    assert not [r for r in bom if r["designation"] == "ANCHOR ROD"]
    cross = [r for r in boq
             if r["description"].startswith("CROSS-CHECK ONLY")]
    assert len(cross) == 1
    assert cross[0]["source_item_id"] == "ANCH-001"
    assert cross[0]["qty"] == 64  # 16 columns x 4, no waste
    assert "MATCH" in cross[0]["notes"]
    assert any("MATCH" in line for line in report["cross_checks"])


def test_cross_check_mismatch_warns_in_validator(tmp_path, asm_dir):
    path, _, boq, _ = _apply(tmp_path, asm_dir, _col_rows(anch_qty=60))
    cross = [r for r in boq
             if r["description"].startswith("CROSS-CHECK ONLY")]
    assert "MISMATCH" in cross[0]["notes"]
    report = validate_takeoff.validate_file(path)
    assert any("disagrees" in w for w in report["warnings"])
    assert not report["hard_fails"]


def test_unmeasured_class_gets_flagged_order_line(tmp_path, asm_dir):
    # No PLATE rows in the takeoff: the base plate order line IS
    # emitted, flagged for verification against the schedule.
    _, _, _, bom = _apply(tmp_path, asm_dir, _col_rows())
    plates = [r for r in bom if r["designation"] == "BASE PL"]
    assert plates[0]["qty"] == 16
    assert "no measured PLATE rows" in plates[0]["notes"]


# -- dual-mode scoping (3.3) -------------------------------------------------

def test_group_linked_linear_row_emits_materials_only(
        tmp_path, asm_dir):
    _, report, boq, bom = _apply(tmp_path, asm_dir, [
        _row("COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
             "S4.00 COLUMN SCHEDULE"),
        _row("BEAM-001", "HSS8X8X1/4", "LINEAR", 240, "LF",
             "S3.00 ROOF FRAMING PLAN", "GROUP: COL-001"),
    ])
    assert _by_ref(bom, ":stick_material:")  # material rides
    grouped = [r for r in boq if r["source_item_id"] == "BEAM-001"
               and r["stream"] in ("LABOR", "EQUIPMENT")]
    assert not grouped  # the COUNT anchor's assembly drives those
    assert any("BEAM-001" in line
               for line in report["group_suppressed"])


# -- streams and governance --------------------------------------------------

def test_subcontract_stream_is_boq_only(tmp_path, asm_dir):
    _, _, boq, bom = _apply(tmp_path, asm_dir, [
        _row("MISC-001", "WIDGET FRAME", "COUNT", 3, "EA",
             "A5.1 WIDGET DETAILS"),
    ])
    sub = [r for r in boq if r["stream"] == "SUBCONTRACT"]
    assert sub[0]["qty"] == 3 and sub[0]["unit"] == "EA"
    assert not bom  # scope context never becomes an order line


def test_definition_with_pricing_token_fails_to_load(tmp_path):
    d = tmp_path / "asm"
    d.mkdir()
    bad = json.loads(json.dumps(_WIDGET_DEF))
    bad["description"] = "unit rate per SF"
    (d / "widget_count.json").write_text(json.dumps(bad),
                                         encoding="utf-8")
    with pytest.raises(ValueError, match="rate"):
        apply_assemblies.load_definitions(d)


def test_definition_with_dollar_fails_to_load(tmp_path):
    d = tmp_path / "asm"
    d.mkdir()
    bad = json.loads(json.dumps(_WIDGET_DEF))
    bad["notes"] = "about $40 each"
    (d / "widget_count.json").write_text(json.dumps(bad),
                                         encoding="utf-8")
    with pytest.raises(ValueError, match="dollar"):
        apply_assemblies.load_definitions(d)


def test_cross_check_with_per_sheet_sf_rejected(tmp_path):
    # A cross-check component compares counts; per_sheet_sf math has
    # no count meaning and used to crash at accumulation instead of
    # failing loud at load.
    d = tmp_path / "asm"
    d.mkdir()
    bad = json.loads(json.dumps(_DECK_DEF))
    bad["streams"]["materials"][0]["cross_check_class"] = "ANCH"
    (d / "deck_test.json").write_text(json.dumps(bad),
                                      encoding="utf-8")
    with pytest.raises(ValueError, match="qty_per_unit"):
        apply_assemblies.load_definitions(d)


def test_round_down_definition_rejected(tmp_path):
    d = tmp_path / "asm"
    d.mkdir()
    bad = json.loads(json.dumps(_BEAM_DEF))
    bad["streams"]["materials"][0]["round"]["direction"] = "down"
    (d / "beam_test.json").write_text(json.dumps(bad),
                                      encoding="utf-8")
    with pytest.raises(ValueError, match="never .*round down|round down"):
        apply_assemblies.load_definitions(d)


def test_ambiguous_definitions_fail_loud(tmp_path, asm_dir):
    clone = json.loads(json.dumps(_WIDGET_DEF))
    clone["assembly_id"] = "widget_clone"
    (asm_dir / "widget_clone.json").write_text(json.dumps(clone),
                                               encoding="utf-8")
    path = tmp_path / "takeoff.xlsx"
    _workbook([_row("MISC-001", "WIDGET FRAME", "COUNT", 3, "EA",
                    "A5.1 WIDGET DETAILS")]).save(str(path))
    with pytest.raises(ValueError, match="more than one"):
        apply_assemblies.apply_to_workbook(path, asm_dir=asm_dir)


def test_stamped_workbook_refused(tmp_path, asm_dir):
    import openpyxl

    path = tmp_path / "takeoff.xlsx"
    wb = _workbook([_row("COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
                         "S4.00 COLUMN SCHEDULE")])
    wb["TAKEOFF"].cell(row=1, column=1, value="TAKEOFF_SCHEMA_V2")
    wb.save(str(path))
    with pytest.raises(ValueError, match="immutable"):
        apply_assemblies.apply_to_workbook(path, asm_dir=asm_dir)


def test_unassigned_rows_and_orphaned_openings_reported(
        tmp_path, asm_dir):
    _, report, boq, bom = _apply(tmp_path, asm_dir, [
        _row("DECK-001", "GRATING PLANK", "AREA", 500, "SF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("DECK-002", "OPENING", "AREA", 64, "SF",
             "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001"),
    ])
    assert not boq and not bom
    assert any("DECK-001" in line for line in report["unassigned"])
    assert any("DECK-002" in line
               for line in report["openings"]["orphaned"])


def test_zero_qty_removed_rows_derive_nothing(tmp_path, asm_dir):
    _, report, boq, bom = _apply(tmp_path, asm_dir, [
        _row("COL-001", "HSS8X8X1/4", "COUNT", 0, "EA",
             "S4.00 COLUMN SCHEDULE", "REMOVED: R2"),
    ])
    assert not boq and not bom
    assert any("COL-001" in line
               for line in report["skipped_zero_qty"])


# -- shipped definitions, end to end -----------------------------------------

def test_shipped_defs_apply_and_validator_passes(tmp_path):
    """The full loop with the SHIPPED library: apply, then the
    section 14 gate must report zero hard fails. Structure only, no
    seed-factor pinning."""
    rows = [
        _row("COL-001", "HSS8X8X1/4", "COUNT", 16, "EA",
             "S4.00 COLUMN SCHEDULE"),
        _row("ANCH-001", "ANCHOR ROD 3/4", "COUNT", 64, "EA",
             "S4.00 COLUMN SCHEDULE + TYPICAL DETAILS"),
        _row("BEAM-001", "W12X26", "LINEAR", 240, "LF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("JST-001", "28K7", "COUNT", 12, "EA",
             "S3.00 ROOF FRAMING PLAN + JOIST SCHEDULE"),
        _row("DECK-001", "TYPE B ROOF DECK", "AREA", 1000, "SF",
             "S3.00 ROOF FRAMING PLAN"),
        _row("DECK-002", "OPENING", "AREA", 64, "SF",
             "S3.00 ROOF FRAMING PLAN", "OPENING: DECK-001 RTU"),
        _row("MISC-001", "STAIR 1", "COUNT", 2, "EA",
             "A5.1 STAIR DETAILS"),
    ]
    path = tmp_path / "takeoff.xlsx"
    wb = _workbook(rows)
    t = wb["TAKEOFF"]
    t.cell(row=5, column=13, value="=D5*L5")  # BEAM-001 weight
    t.cell(row=5, column=14, value="=M5/2000")
    t.cell(row=5, column=15,
           value="AISC:bridge/aisc_validator.py:W12X26")
    wb.save(str(path))

    report = apply_assemblies.apply_to_workbook(path)
    assert report["status"] == "APPLIED"
    # Five drivers: COL, BEAM, JST, DECK, stair. ANCH-001 is measured
    # (system of record, never a driver) and DECK-002 is an OPENING.
    assert report["drivers_applied"] == 5
    assert not report["unassigned"]
    assert any("ANCH-001" in line
               for line in report["measured_not_drivers"])

    boq = _read_sheet(path, "BOQ", BOQ_HEADERS)
    bom = _read_sheet(path, "BOM", BOM_HEADERS)
    streams = {r["stream"] for r in boq}
    assert {"MATERIAL", "LABOR", "EQUIPMENT"} <= streams
    # Measured ANCH rows: cross-check only, never an order line.
    assert not [r for r in bom if r["designation"] == "ANCHOR ROD"]
    assert [r for r in boq
            if r["description"].startswith("CROSS-CHECK ONLY")]
    # Openings: deck deducted AND framed steel added.
    assert _by_ref(bom, ":frame_angle:")
    assert _by_ref(bom, "roof_deck_sf:deck_sheets:")

    gate = validate_takeoff.validate_file(path)
    assert gate["hard_fails"] == []
