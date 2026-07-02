"""Tests for the Tekla reconciliation (Prompt 8, T13/F5).

Synthetic takeoff workbooks are written exactly the way the Prompt 5
exporter writes them (empty metadata row, schema header on row 2,
formula cells =D*L and =M/2000 on LINEAR rows, never opened in Excel,
so no cached formula values exist). Synthetic Tekla BOM exports carry
known deltas, so every variance number asserted here is hand-derived.
"""

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest

from takeoff_pipeline import tekla_reconcile as tr
from takeoff_pipeline.tekla_reconcile import (
    TAKEOFF_HEADERS,
    diff,
    normalize_designation,
    read_bom,
    read_takeoff,
    run,
    verdict_line,
)

MODE_UNIT = {"COUNT": "EA", "LINEAR": "LF", "AREA": "SF"}


def make_takeoff(path, rows, header=None):
    """Write a takeoff xlsx the way export_xlsx.py does: metadata row
    empty (unstamped), header row 2, exporter formula pair on LINEAR
    rows, lb_per_ft as the operator input cell."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TAKEOFF"
    ws.append([])
    ws.append(list(header or TAKEOFF_HEADERS))
    r = 3
    for row in rows:
        ws.cell(row=r, column=1, value=f"BEAM-{r:03d}")
        ws.cell(row=r, column=2, value=row["designation"])
        ws.cell(row=r, column=3, value=row["mode"])
        ws.cell(row=r, column=4, value=row["qty"])
        ws.cell(row=r, column=5, value=MODE_UNIT[row["mode"]])
        ws.cell(row=r, column=6, value="S2.1 ROOF FRAMING PLAN")
        ws.cell(row=r, column=8, value="medium")
        ws.cell(row=r, column=9, value="S2.1")
        ws.cell(row=r, column=10, value="MANUAL")
        if row.get("notes"):
            ws.cell(row=r, column=11, value=row["notes"])
        if row["mode"] == "LINEAR" and "tons_cell" not in row:
            if row.get("lb_per_ft") is not None:
                ws.cell(row=r, column=12, value=row["lb_per_ft"])
            ws.cell(row=r, column=13, value=f"=D{r}*L{r}")
            ws.cell(row=r, column=14, value=f"=M{r}/2000")
            ws.cell(row=r, column=15,
                    value="AISC:bridge/aisc_validator.py:TEST")
        if "tons_cell" in row:
            ws.cell(row=r, column=14, value=row["tons_cell"])
        r += 1
    wb.save(path)
    return path


def make_bom_csv(path, header, rows, title_rows=()):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for t in title_rows:
            w.writerow(t)
        w.writerow(header)
        w.writerows(rows)
    return path


# -- normalization -----------------------------------------------------------

@pytest.mark.parametrize("a,b", [
    ('HSS 6x6x1/4', 'HSS6X6X1/4'),
    ('HSS6X6X0.25', 'HSS6X6X1/4'),
    ('HSS6X6X.250', 'HSS6X6X1/4'),
    ('TS6X6X.25', 'HSS6X6X1/4'),
    ('HSS6×6×1/4', 'HSS6X6X1/4'),
    ('PL 1/2"x12"', 'PL1/2X12'),
    ('PL0.5X12', 'PL1/2X12'),
    ('FL1/2X12', 'PL1/2X12'),
    ('L3x3x0.25', 'L3X3X1/4'),
    ('2L3X3X0.25', '2L3X3X1/4'),
    ('L8X8X1.375', 'L8X8X1-3/8'),
    ('32LH07', '32LH7'),
    ('28K07', '28K7'),
    ('w12x26', 'W12X26'),
    ('C8X11.50', 'C8X11.5'),
    ('HSS6.625X0.280', 'HSS6.625X.28'),
    ('HSS2.5X2.5X.125', 'HSS2-1/2X2-1/2X1/8'),
    ('L2.5X2.5X0.25', 'L2-1/2X2-1/2X1/4'),
    ('L3.5X3.5X0.25', 'L3-1/2X3-1/2X1/4'),
    ('HSS3.5X3.5X.25', 'HSS3-1/2X3-1/2X1/4'),
    ('L2 1/2x2 1/2x1/4', 'L2.5X2.5X0.25'),
])
def test_normalize_pairs_share_a_key(a, b):
    assert normalize_designation(a) == normalize_designation(b)


def test_normalize_no_false_merge_on_mixed_fractions():
    # 13/16 thickness and 1-3/16 (1.1875 decimal) are different
    # plates; a glued mixed-number form would collide them.
    assert (normalize_designation("PL13/16X8")
            != normalize_designation("PL1.1875X8"))
    assert normalize_designation("PL1.1875X8") == "PL1-3/16X8"
    # WT trailing numbers are weights; never converted.
    assert normalize_designation("WT2X6.5") == "WT2X6.5"


def test_normalize_never_invents_fractions_on_weights():
    # Channel and tee labels trail decimal WEIGHTS, not thickness.
    assert normalize_designation("C8X11.5") == "C8X11.5"
    assert normalize_designation("MC12X10.6") == "MC12X10.6"
    assert normalize_designation("WT12X27.5") == "WT12X27.5"
    # Round HSS diameter is mid-string and stays decimal.
    assert normalize_designation("HSS6.625X.280").startswith("HSS6.625X")


def test_normalize_joists_and_girders():
    assert normalize_designation("28K7") == "28K7"
    assert normalize_designation("24KCS3") == "24KCS3"
    assert normalize_designation("54G8N12.1K") == "54G8N12.1K"
    assert normalize_designation("36G8N11.5K") == "36G8N11.5K"
    assert normalize_designation("") == ""


# -- BOM reading ---------------------------------------------------------------

def test_bom_csv_title_rows_repeated_headers_totals(tmp_path):
    path = make_bom_csv(
        tmp_path / "bom.csv",
        ["Profile", "Qty", "Length", "Total weight"],
        [
            ["W12X26", "10", "300.0", "7800"],
            ["Profile", "Qty", "Length", "Total weight"],
            ["W12X26", "2", "60.0", "1560"],
            ["HSS6X6X0.25", "4", "80.0", "1530.4"],
            ["TOTAL", "", "", "10890.4"],
            ["", "", "", ""],
        ],
        title_rows=[["YOUR COMPANY TEST EXPORT"], [""]],
    )
    out = read_bom(path)
    agg = out["agg"]
    assert agg["W12X26"]["qty"] == 12
    assert agg["W12X26"]["weight_lb"] == 9360
    assert agg["HSS6X6X1/4"]["qty"] == 4
    assert out["columns"]["names"]["profile"] == "PROFILE"
    assert out["columns"]["names"]["weight"] == "TOTAL WEIGHT"


def test_bom_prefers_total_weight_over_unit_weight(tmp_path):
    path = make_bom_csv(
        tmp_path / "bom.csv",
        ["Profile", "No.", "Weight each", "Total weight (lb)"],
        [["W12X26", "10", "780", "7800"]])
    out = read_bom(path)
    assert out["columns"]["names"]["weight"] == "TOTAL WEIGHT (LB)"
    assert out["agg"]["W12X26"]["weight_lb"] == 7800


def test_bom_per_piece_weight_multiplied(tmp_path):
    path = make_bom_csv(
        tmp_path / "bom.csv",
        ["Designation", "Count", "Unit weight"],
        [["W12X26", "10", "780"]])
    out = read_bom(path)
    assert out["agg"]["W12X26"]["weight_lb"] == 7800


def test_bom_kg_converted(tmp_path):
    path = make_bom_csv(
        tmp_path / "bom.csv",
        ["Profile", "Qty", "Weight (kg)"],
        [["W12X26", "1", "1000"]])
    out = read_bom(path)
    assert out["agg"]["W12X26"]["weight_lb"] == pytest.approx(2204.62,
                                                              abs=0.01)


def test_bom_xlsx_with_title_rows(tmp_path):
    import openpyxl

    path = tmp_path / "bom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["TEKLA STRUCTURES BOM"])
    ws.append([])
    ws.append(["Profile", "Qty", "Total weight"])
    ws.append(["W12X26", 10, 7800])
    ws.append(["28K7", 63, None])
    wb.save(path)
    out = read_bom(path)
    assert out["agg"]["W12X26"]["qty"] == 10
    assert out["agg"]["28K7"]["qty"] == 63
    assert out["agg"]["28K7"]["has_weight"] is False


def test_bom_without_detectable_header_fails(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["colA", "colB"], [["W12X26", "10"]])
    with pytest.raises(ValueError, match="no header row"):
        read_bom(path)


def test_bom_unsupported_extension_fails(tmp_path):
    p = tmp_path / "bom.pdf"
    p.write_text("not a bom")
    with pytest.raises(ValueError, match="csv or xlsx"):
        list(tr._bom_rows(p))


# -- takeoff reading -----------------------------------------------------------

def test_takeoff_formula_evaluation_and_warnings(tmp_path):
    path = make_takeoff(tmp_path / "t.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 300.0,
         "lb_per_ft": 26.0},
        {"designation": "W12X26", "mode": "COUNT", "qty": 10},
        {"designation": "HSS6X6X1/4", "mode": "LINEAR", "qty": 80.0,
         "lb_per_ft": None},
        {"designation": "OPENING", "mode": "AREA", "qty": 64},
        {"designation": "RF DECK 1.5B", "mode": "AREA", "qty": 12000},
    ])
    out = read_takeoff(path)
    agg = out["agg"]
    # 300 LF x 26 lb/ft / 2000 = 3.9 tons, from the file's own cells.
    assert agg["W12X26"]["tons"] == pytest.approx(3.9)
    assert agg["W12X26"]["qty_ea"] == 10
    assert agg["W12X26"]["lf"] == 300
    # LINEAR row without lb_per_ft carries no tonnage and warns.
    assert agg["HSS6X6X1/4"]["has_tons"] is False
    assert any("no numeric lb_per_ft" in w for w in out["warnings"])
    assert out["openings"] == 1
    assert "OPENING" not in agg
    assert agg["RFDECK1.5B"]["sf"] == 12000


def test_takeoff_nonstandard_formula_skipped_not_guessed(tmp_path):
    path = make_takeoff(tmp_path / "t.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 100.0,
         "tons_cell": "=Q9*2"},
    ])
    out = read_takeoff(path)
    assert out["agg"]["W12X26"]["has_tons"] is False
    assert any("not the exporter formula" in w for w in out["warnings"])


def test_takeoff_numeric_tons_cell_used(tmp_path):
    path = make_takeoff(tmp_path / "t.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 100.0,
         "tons_cell": 1.3},
    ])
    out = read_takeoff(path)
    assert out["agg"]["W12X26"]["tons"] == pytest.approx(1.3)


def test_takeoff_wrong_header_fails(tmp_path):
    bad = list(TAKEOFF_HEADERS)
    bad[1] = "shape"
    path = make_takeoff(tmp_path / "t.xlsx",
                        [{"designation": "W12X26", "mode": "COUNT",
                          "qty": 1}], header=bad)
    with pytest.raises(ValueError, match="TAKEOFF_SCHEMA_V2 header"):
        read_takeoff(path)


# -- diff and verdict ------------------------------------------------------------

def _scored(v):
    return {"variance_pct": v, "bid_has_tons": True,
            "model_has_tons": True}


def test_verdict_bands():
    assert "PASS" in verdict_line(_scored(1.5))
    assert "tight 2 percent band" in verdict_line(_scored(-1.99))
    inner_missed = verdict_line(_scored(2.5))
    assert "PASS" in inner_missed and "missed" in inner_missed
    assert "FAIL" in verdict_line(_scored(3.01))
    assert "FAIL" in verdict_line(_scored(-12.0))
    for line in (verdict_line(_scored(1.5)), verdict_line(_scored(5))):
        assert "2 to 3 percent" in line


def test_verdict_not_scored_reasons():
    line = verdict_line({"variance_pct": None, "bid_has_tons": False,
                         "model_has_tons": True})
    assert "NOT SCORED" in line and "no derived tonnage" in line
    line = verdict_line({"variance_pct": None, "bid_has_tons": True,
                         "model_has_tons": False})
    assert "NOT SCORED" in line and "no weight column" in line


def test_diff_ranks_by_qty_when_either_side_lacks_tonnage():
    bid = {"W12X26": {"raw": {"W12X26"}, "qty_ea": 10, "lf": 0,
                      "sf": 0, "tons": 0.0, "has_qty_ea": True,
                      "has_tons": False}}
    model = {"W12X26": {"raw": {"W12X26"}, "qty": 14,
                        "weight_lb": 0.0, "has_qty": True,
                        "has_weight": False}}
    d = diff(bid, model)
    assert d["variance_pct"] is None
    assert d["rank_basis"] == "qty"
    assert d["misses"][0]["qty_delta"] == -4


# -- end to end -------------------------------------------------------------------

def test_run_end_to_end(tmp_path):
    takeoff = make_takeoff(tmp_path / "JOB1_TAKEOFF_R1.xlsx", [
        # 300 LF x 26 = 3.9 T; 200 LF x 28.43 = 2.843 T
        {"designation": "W12X26", "mode": "LINEAR", "qty": 300.0,
         "lb_per_ft": 26.0},
        {"designation": "W12X26", "mode": "COUNT", "qty": 10},
        {"designation": "HSS 6x6x1/4", "mode": "LINEAR", "qty": 200.0,
         "lb_per_ft": 28.43},
        {"designation": "28K7", "mode": "COUNT", "qty": 63},
    ])
    # Model: W12X26 heavier (8000 lb = 4.0 T), HSS as decimal profile
    # (5600 lb = 2.8 T), one model-only angle (200 lb = 0.1 T).
    bom = make_bom_csv(tmp_path / "bom.csv",
                       ["Profile", "Qty", "Total weight"],
                       [["W12X26", "11", "8000"],
                        ["HSS6X6X0.25", "4", "5600"],
                        ["L3X3X1/4", "2", "200"]])
    out_dir = tmp_path / "ledger"
    d = run(takeoff, bom, ledger_dir=out_dir)

    bid_tons = 3.9 + 2.843
    model_tons = (8000 + 5600 + 200) / 2000.0
    expected = (bid_tons - model_tons) / model_tons * 100.0
    assert d["variance_pct"] == pytest.approx(expected)
    assert d["job"] == "JOB1"

    report = d["report_path"].read_text(encoding="utf-8")
    assert d["report_path"].name == "reconcile_JOB1.md"
    assert "VERDICT:" in report
    assert "-GP CONFIDENTIAL" in report
    assert "2 to 3 percent" in report
    assert "\u2014" not in report  # no em-dashes, Hard Rule 7
    # The HSS pair matched across notations.
    hss_row = next(r for r in d["rows"] if r["key"] == "HSS6X6X1/4")
    assert hss_row["status"] == "MATCHED"
    # Joists are bid-only when the model excludes them.
    jst_row = next(r for r in d["rows"] if r["key"] == "28K7")
    assert jst_row["status"] == "BID_ONLY"

    ledger = out_dir / "accuracy_ledger.csv"
    rows = list(csv.reader(ledger.open(newline="", encoding="utf-8")))
    assert rows[0] == tr.LEDGER_HEADER
    metrics = {r[3] for r in rows[1:]}
    assert {"bid_tons", "model_tons", "tonnage_variance_pct",
            "designations_matched"} <= metrics
    assert all(r[2] == "tekla_reconcile_T13" for r in rows[1:])

    # Second run appends without duplicating the header.
    run(takeoff, bom, ledger_dir=out_dir)
    rows2 = list(csv.reader(ledger.open(newline="", encoding="utf-8")))
    assert sum(1 for r in rows2 if r == tr.LEDGER_HEADER) == 1
    assert len(rows2) > len(rows)


def test_run_not_scored_when_bid_has_no_tonnage(tmp_path):
    takeoff = make_takeoff(tmp_path / "JOB2_TAKEOFF.xlsx", [
        {"designation": "28K7", "mode": "COUNT", "qty": 63},
        {"designation": "30K7", "mode": "COUNT", "qty": 5},
    ])
    bom = make_bom_csv(tmp_path / "bom.csv",
                       ["Profile", "Qty", "Total weight"],
                       [["28K7", "60", "60000"]])
    d = run(takeoff, bom, ledger_dir=tmp_path / "ledger")
    assert d["variance_pct"] is None
    assert "NOT SCORED" in d["verdict"]
    assert d["rank_basis"] == "qty"
    report = d["report_path"].read_text(encoding="utf-8")
    assert "not computable" in report
    # No variance row lands in the ledger when nothing was scored.
    ledger = tmp_path / "ledger" / "accuracy_ledger.csv"
    rows = list(csv.reader(ledger.open(newline="", encoding="utf-8")))
    assert not any(r[3] == "tonnage_variance_pct" for r in rows)


def test_main_cli(tmp_path, capsys):
    takeoff = make_takeoff(tmp_path / "JOB3_TAKEOFF.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 100.0,
         "lb_per_ft": 26.0},
    ])
    bom = make_bom_csv(tmp_path / "bom.csv",
                       ["Profile", "Qty", "Total weight"],
                       [["W12X26", "4", "2600"]])
    rc = tr.main([str(takeoff), str(bom), "--job", "JOB3",
                  "--ledger-dir", str(tmp_path / "out")])
    captured = capsys.readouterr().out
    assert rc == 0
    assert "VERDICT:" in captured
    assert (tmp_path / "out" / "reconcile_JOB3.md").exists()


def test_main_cli_bad_input(tmp_path, capsys):
    bom = make_bom_csv(tmp_path / "bom.csv", ["colA"], [["x"]])
    takeoff = make_takeoff(tmp_path / "T_TAKEOFF.xlsx", [
        {"designation": "W12X26", "mode": "COUNT", "qty": 1}])
    rc = tr.main([str(takeoff), str(bom)])
    assert rc == 2
    assert "reconcile failed" in capsys.readouterr().out


# -- review fixes (adversarial multi-agent pass) ---------------------------------

def test_matched_count_designation_excluded_from_variance(tmp_path):
    """The blocker repro: a perfect takeoff must not FAIL because the
    model carries joist weight the takeoff schema cannot weigh."""
    takeoff = make_takeoff(tmp_path / "JOB4_TAKEOFF.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 300.0,
         "lb_per_ft": 26.0},
        {"designation": "28K7", "mode": "COUNT", "qty": 63},
    ])
    bom = make_bom_csv(tmp_path / "bom.csv",
                       ["Profile", "Qty", "Total weight"],
                       [["W12X26", "10", "7800"],
                        ["28K7", "63", "60000"]])
    d = run(takeoff, bom, ledger_dir=tmp_path / "ledger")
    assert d["variance_pct"] == pytest.approx(0.0)
    assert "PASS" in d["verdict"]
    assert d["excl_model_tons"] == pytest.approx(30.0)
    jst = next(r for r in d["rows"] if r["key"] == "28K7")
    assert jst["status"] == "MATCHED"
    assert jst["tons_delta"] is None
    assert jst["qty_delta"] == 0
    report = d["report_path"].read_text(encoding="utf-8")
    assert "excluded from the variance" in report
    ledger = tmp_path / "ledger" / "accuracy_ledger.csv"
    rows = list(csv.reader(ledger.open(newline="", encoding="utf-8")))
    excl = [r for r in rows if r[3] == "model_tons_excluded"]
    assert excl and excl[0][4] == "30.0"


def test_bom_tons_header_dropped_then_override(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["Profile", "Qty", "Weight (t)"],
                        [["W12X26", "10", "3.9"]])
    out = read_bom(path)
    assert out["agg"]["W12X26"]["has_weight"] is False
    assert any("tons unit" in w for w in out["warnings"])
    out2 = read_bom(path, weight_unit="ton")
    assert out2["agg"]["W12X26"]["weight_lb"] == pytest.approx(7800)


def test_bom_kg_suffix_in_cell_converted(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["Profile", "Qty", "Weight"],
                        [["W12X26", "1", "1000 kg"]])
    out = read_bom(path)
    assert out["agg"]["W12X26"]["weight_lb"] == pytest.approx(
        2204.62, abs=0.01)


def test_bom_unreadable_qty_warns_not_silent(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["Profile", "Qty", "Total weight"],
                        [["W12X26", "10 PCS", "7800"],
                         ["W12X26", "5", "3900"]])
    out = read_bom(path)
    assert out["agg"]["W12X26"]["qty"] == 5
    assert any("unreadable qty" in w for w in out["warnings"])


def test_bom_nan_weight_warns_and_run_survives(tmp_path):
    takeoff = make_takeoff(tmp_path / "JOB5_TAKEOFF.xlsx", [
        {"designation": "W12X26", "mode": "LINEAR", "qty": 100.0,
         "lb_per_ft": 26.0}])
    bom = make_bom_csv(tmp_path / "bom.csv",
                       ["Profile", "Qty", "Total weight"],
                       [["W12X26", "4", "2600"],
                        ["L3X3X1/4", "2", "nan"],
                        ["L4X4X1/4", "2", "inf"]])
    d = run(takeoff, bom, ledger_dir=tmp_path / "ledger")
    assert d["model_tons"] == pytest.approx(1.3)
    assert d["report_path"].exists()


def test_bom_totals_variants_filtered(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["Profile", "Qty", "Total weight"],
                        [["W12X26", "10", "7800"],
                         ["Total:", "", "7800"],
                         ["Grand Total:", "", "7800"],
                         ["Sub-total", "", "7800"]])
    out = read_bom(path)
    assert list(out["agg"]) == ["W12X26"]
    assert out["agg"]["W12X26"]["weight_lb"] == 7800


def test_bom_semicolon_csv_with_title_row(tmp_path):
    path = tmp_path / "bom.csv"
    path.write_text("TEKLA STRUCTURES EXPORT\n"
                    "Profile;Qty;Total weight\n"
                    "W12X26;10;7800\n", encoding="utf-8")
    out = read_bom(path)
    assert out["agg"]["W12X26"]["qty"] == 10


def test_bom_header_without_data_raises(tmp_path):
    path = make_bom_csv(tmp_path / "bom.csv",
                        ["Profile", "Qty", "Total weight"], [])
    with pytest.raises(ValueError, match="no usable data rows"):
        read_bom(path)


def test_header_map_rejects_single_cell_pseudo_header():
    assert tr._header_map(["Profile Qty Total weight"]) is None


def test_takeoff_numeric_mode_cell_does_not_crash(tmp_path):
    import openpyxl

    path = make_takeoff(tmp_path / "t.xlsx", [
        {"designation": "W12X26", "mode": "COUNT", "qty": 1}])
    wb = openpyxl.load_workbook(str(path))
    wb["TAKEOFF"].cell(row=3, column=3, value=123)
    wb.save(path)
    out = read_takeoff(path)
    assert "W12X26" in out["agg"]


def test_verdict_classifies_on_displayed_rounding():
    assert "PASS" in verdict_line(_scored(3.004))
    assert "FAIL" in verdict_line(_scored(3.006))
    assert "tight 2 percent band" in verdict_line(_scored(2.004))
