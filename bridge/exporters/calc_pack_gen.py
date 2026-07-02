"""Auditable calculation pack (Phase 16, build slot 16, v4.7.0).

Generates a PE-friendly Excel workbook showing exactly how every number
in the bid was computed. When a GC or owner's engineer asks "how did
you get $485,000?", Joseph sends this file. Every weight comes from
AISC v16.0 Table 1-1, every rate from the locked calibration, every
connection cost from the Phase 10 assembly table.

Output: Excel .xlsx with four tabs:
    1. Summary - project totals, tonnage, hours, cost
    2. Members - per-member weight from AISC lb/ft x length x qty
    3. Connections - per-connection hardware cost from assembly table
    4. Rates - current shop rate, overhead, margin, material $/lb

The workbook is self-contained. No macros, no external links, no
circular references. A PE can audit every cell.

Voice rules: zero em-dashes. Hyphens or periods only.
"""


import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

log = logging.getLogger(__name__)


# Lazy import. openpyxl is in the existing stack.
try:
    import openpyxl  # noqa: F401
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_calc_pack(
    takeoff_result: dict,
    bid_number: str = "",
    project_name: str = "",
    output_path: str | Path | None = None,
) -> dict:
    """Generate an auditable calculation pack from takeoff data.

    Args:
        takeoff_result: v1 or v2 takeoff result dict.
        bid_number: Your Company bid number.
        project_name: Project name for the header.
        output_path: If provided, write .xlsx here.

    Returns:
        {
            "success": bool,
            "output_path": str,
            "tabs": list[str],
            "member_count": int,
            "connection_count": int,
            "warnings": list[str],
        }
    """
    warnings: list[str] = []

    if not HAS_OPENPYXL:
        return {
            "success": False,
            "output_path": "",
            "tabs": [],
            "member_count": 0,
            "connection_count": 0,
            "warnings": ["openpyxl_not_installed"],
        }

    wb = openpyxl.Workbook()

    members = takeoff_result.get("valid_members",
                takeoff_result.get("members", []))
    details = takeoff_result.get("details", [])
    assembly = takeoff_result.get("assembly_costs", {})
    breakdown = takeoff_result.get("cost_breakdown", {})

    # Styles
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    hdr_font_white = Font(bold=True, size=11, color="FFFFFF")
    num_fmt = "#,##0.00"
    usd_fmt = "$#,##0.00"
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Tab 1: Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    _write_header(ws, 1, ["Your Company - Calculation Pack"], hdr_font)
    ws.append(["Bid Number:", bid_number or ""])
    ws.append(["Project:", project_name or ""])
    ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M")])  # vj: local-display-ok
    ws.append(["AISC Edition:", "v16.0 (2,299 shapes)"])
    ws.append([])

    ws.append(["PROJECT TOTALS", ""])
    ws["A7"].font = hdr_font
    ws.append(["Total Tons (structural + misc):",
               float(takeoff_result.get("total_tons", 0))])
    ws["B8"].number_format = num_fmt
    ws.append(["Structural Tons:",
               float(takeoff_result.get("structural_tons", 0))])
    ws["B9"].number_format = num_fmt
    ws.append(["Misc Steel Tons:",
               float(takeoff_result.get("misc_tons", 0))])
    ws["B10"].number_format = num_fmt
    ws.append(["Member Count:", len(members)])
    ws.append(["Connection Count:", len(details)])
    ws.append([])

    ws.append(["COST SUMMARY", ""])
    ws["A14"].font = hdr_font
    bd = breakdown.get("breakdown", breakdown)
    ws.append(["Total Bid:",
               float(takeoff_result.get("total_cost", 0))])
    ws["B15"].number_format = usd_fmt
    ws.append(["Material Subtotal:",
               float(bd.get("material_subtotal", 0))])
    ws["B16"].number_format = usd_fmt
    ws.append(["Labor:",
               float(bd.get("labor", 0))])
    ws["B17"].number_format = usd_fmt
    ws.append(["Coatings:",
               float(bd.get("coatings", 0))])
    ws["B18"].number_format = usd_fmt
    ws.append(["Freight:",
               float(bd.get("freight", 0))])
    ws["B19"].number_format = usd_fmt
    ws.append(["Connection Hardware:",
               float(assembly.get("total_connection_cost_usd", 0))])
    ws["B20"].number_format = usd_fmt
    ws.append(["Cost Per Ton:",
               float(takeoff_result.get("cost_per_ton", 0))])
    ws["B21"].number_format = usd_fmt

    # ── Tab 2: Members ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Members")
    headers = ["Mark", "Shape", "Size", "Grade", "Length (ft)", "Qty",
               "lb/ft (AISC)", "Weight (lbs)", "AISC Reference"]
    for ci, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hdr_font_white
        cell.fill = hdr_fill
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["I"].width = 22

    for ri, m in enumerate(members, 2):
        shape = m.get("shape") or m.get("normalized") or ""
        size = m.get("size", "")
        full = f"{shape}{size}" if size else shape
        plf = float(m.get("plf", 0) or 0)
        length = float(m.get("length_ft", 0) or 0)
        qty = int(m.get("qty", 1) or 1)
        weight = plf * length * qty if plf > 0 else \
                 float(m.get("weight_lbs", 0) or 0)

        ws2.cell(row=ri, column=1, value=m.get("mark", "")).border = thin_border
        ws2.cell(row=ri, column=2, value=shape).border = thin_border
        ws2.cell(row=ri, column=3, value=size).border = thin_border
        ws2.cell(row=ri, column=4, value=m.get("grade", "A992")).border = thin_border
        c5 = ws2.cell(row=ri, column=5, value=length)
        c5.number_format = num_fmt
        c5.border = thin_border
        ws2.cell(row=ri, column=6, value=qty).border = thin_border
        c7 = ws2.cell(row=ri, column=7, value=plf)
        c7.number_format = num_fmt
        c7.border = thin_border
        c8 = ws2.cell(row=ri, column=8, value=round(weight, 2))
        c8.number_format = num_fmt
        c8.border = thin_border
        ws2.cell(row=ri, column=9,
                 value=f"AISC v16.0 Table 1-1: {full}").border = thin_border

    # ── Tab 3: Connections ────────────────────────────────────────────
    ws3 = wb.create_sheet("Connections")
    conn_headers = ["Type", "Moment", "Bolt Count", "Welding Hrs",
                     "Hardware Cost", "Assembly Ref"]
    for ci, h in enumerate(conn_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = hdr_font_white
        cell.fill = hdr_fill
        cell.border = thin_border

    per_conn = assembly.get("per_connection", [])
    for ri, c in enumerate(per_conn, 2):
        ws3.cell(row=ri, column=1,
                 value=c.get("connection_type", "")).border = thin_border
        ws3.cell(row=ri, column=2,
                 value="Yes" if "MOMENT" in str(
                     c.get("assembly_key", "")) else "No").border = thin_border
        ws3.cell(row=ri, column=3,
                 value=c.get("bolt_count", 0)).border = thin_border
        c4 = ws3.cell(row=ri, column=4,
                      value=c.get("welding_hrs", 0))
        c4.number_format = num_fmt
        c4.border = thin_border
        c5 = ws3.cell(row=ri, column=5,
                      value=c.get("cost_usd", 0))
        c5.number_format = usd_fmt
        c5.border = thin_border
        ws3.cell(row=ri, column=6,
                 value=c.get("label", "")).border = thin_border

    # ── Tab 4: Rates ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Rates")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 16

    rate_header = ws4.cell(row=1, column=1, value="Rate Schedule")
    rate_header.font = hdr_font
    ws4.append(["Shop Rate ($/hr):", 145.00])
    ws4["B2"].number_format = usd_fmt
    ws4.append(["Engineering Rate ($/hr):", 175.00])
    ws4["B3"].number_format = usd_fmt
    ws4.append(["Overhead Multiplier:", 1.15])
    ws4.append(["Fab Baseline (hrs/ton):", 11.0])
    ws4.append(["Fab Hours (this project):",
               float(takeoff_result.get("fab_hours", 0))])
    ws4["B6"].number_format = num_fmt
    ws4.append(["Erection Hours (this project):",
               float(takeoff_result.get("erect_hours", 0))])
    ws4["B7"].number_format = num_fmt
    ws4.append([])
    ws4.append(["CALIBRATION SOURCE", ""])
    ws4["A9"].font = hdr_font
    ws4.append(["Rates as of:", "Q2 2026 Houston market"])
    ws4.append(["AISC Shape Database:", "v16.0 US Customary (2,299 shapes)"])
    ws4.append(["Margin applied:", f"{float(breakdown.get('margin_pct', 0)) * 100:.1f}%"])

    # Footer
    ws4.append([])
    ws4.append(["This calculation pack is generated by Your Company "
                "Virtual Office. All weights reference AISC Steel "
                "Construction Manual v16.0 Table 1-1. Rates are from "
                "the Q2 2026 Houston-market calibration. Connection "
                "costs reference the Phase 10 assembly costing table."])

    tabs = [ws.title, ws2.title, ws3.title, ws4.title]

    # Write file
    out_path = ""
    if output_path:
        p = Path(output_path)
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(p))
        out_path = str(p)

    return {
        "success": True,
        "output_path": out_path,
        "tabs": tabs,
        "member_count": len(members),
        "connection_count": len(per_conn),
        "warnings": warnings,
    }


def _write_header(ws, row: int, values: list, font: Font) -> None:
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font = font
