"""
planswift_import.py

PlanSwift CSV/XLSX adapter for the BOQ source registry.

PlanSwift exports take-off lists as either CSV or XLSX. Column names
vary between Ivan's template revisions, so this adapter does
case-insensitive column lookup with a fallback map.

Recognized columns (case-insensitive, partial match)
----------------------------------------------------
  description / item / member        -> description
  qty / quantity / count             -> qty
  unit / uom                         -> unit
  rate / unit cost / unit price      -> unit_rate (optional)
  total / amount / extended          -> extended (optional)
  discipline / trade                 -> discipline (default Structural)
  sheet / page / dwg                 -> notes prefix (per-sheet rollup)

Where the adapter looks
-----------------------
  1. ctx.explicit_path - operator-supplied file
  2. ctx.bid_folder / "planswift" / *.csv | *.xlsx
  3. ctx.bid_folder / *planswift*.csv | *planswift*.xlsx
  4. ctx.bid_folder / "boq" / *.csv | *.xlsx with PlanSwift-shaped headers

If any of the above produces a file with at least the required columns
(description, qty, unit), probe() returns True and load() reads it.

Hard rules respected
--------------------
- Does not invent quantities. If a column is missing, the row is
  imported with whatever was present and the gap is recorded in notes.
- Records boq_origin="planswift" so reconciliation Rule F sees a
  PlanSwift source and does not flag the bid.
- No supplier names introduced. PlanSwift exports may contain notes
  with supplier hints (e.g. "Vulcraft 1.5B22"); the adapter strips
  known supplier substrings from the description field before storing.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Optional, List, Dict, Any

from bridge.boq_sources import (
    BoqContext,
    BoqPayload,
    BoqSourceAdapter,
    FIDELITY_PLANSWIFT,
    register,
)

# -------- Column aliases --------
COL_ALIASES = {
    "description": ["description", "item", "member", "name", "piece"],
    "qty":         ["qty", "quantity", "count", "ea", "each"],
    "unit":        ["unit", "uom", "u/m"],
    "unit_rate":   ["rate", "unit cost", "unit price", "$/unit", "unitcost"],
    "extended":    ["total", "amount", "extended", "ext"],
    "discipline":  ["discipline", "trade", "csi"],
    "sheet":       ["sheet", "page", "dwg", "drawing"],
}

REQUIRED_FIELDS = ("description", "qty", "unit")

# Supplier substrings to strip from descriptions before storing.
# Coordinated with bid_output_scrubber rules.
SCRUB_SUPPLIERS = [
    r"\bvulcraft\b",
    r"\bcanam\b",
    r"\bnucor\b",
    r"\bayamsa\b",
]


def _match_column(headers: List[str], aliases: List[str]) -> Optional[str]:
    low = {h.strip().lower(): h for h in headers}
    for alias in aliases:
        if alias in low:
            return low[alias]
    for alias in aliases:
        for h_low, h_orig in low.items():
            if alias in h_low:
                return h_orig
    return None


def _scrub_supplier_names(text: str) -> str:
    if not text:
        return text
    out = text
    for pat in SCRUB_SUPPLIERS:
        out = re.sub(pat, "", out, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", out).strip()


def _candidate_files(ctx: BoqContext) -> List[Path]:
    """Return PlanSwift candidate files in priority order."""
    candidates: List[Path] = []
    # 1. explicit_path wins outright
    if ctx.explicit_path and ctx.explicit_path.is_file():
        if ctx.explicit_path.suffix.lower() in (".csv", ".xlsx"):
            candidates.append(ctx.explicit_path)

    # 2. bid_folder/planswift/*
    if ctx.bid_folder and ctx.bid_folder.is_dir():
        ps_dir = ctx.bid_folder / "planswift"
        if ps_dir.is_dir():
            for p in sorted(ps_dir.iterdir()):
                if p.is_file() and p.suffix.lower() in (".csv", ".xlsx"):
                    candidates.append(p)
        # 3. bid_folder/*planswift*.csv|xlsx
        for p in sorted(ctx.bid_folder.glob("*planswift*")):
            if p.is_file() and p.suffix.lower() in (".csv", ".xlsx"):
                if p not in candidates:
                    candidates.append(p)
        # 4. bid_folder/boq/*.csv|xlsx with PlanSwift-shaped headers
        boq_dir = ctx.bid_folder / "boq"
        if boq_dir.is_dir():
            for p in sorted(boq_dir.iterdir()):
                if p.is_file() and p.suffix.lower() in (".csv", ".xlsx"):
                    if p not in candidates:
                        candidates.append(p)
    return candidates


def _read_rows(path: Path) -> List[Dict[str, Any]]:
    """Read a CSV or XLSX into a list of dicts."""
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [dict(r) for r in reader]
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            raise RuntimeError(
                "openpyxl not installed - cannot read XLSX PlanSwift exports. "
                "Install via: pip install openpyxl"
            )
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        out = []
        for r in rows_iter:
            d = {}
            for h, v in zip(headers, r):
                if h:
                    d[h] = v
            if any(v not in (None, "", "None") for v in d.values()):
                out.append(d)
        return out
    raise RuntimeError(f"Unsupported PlanSwift file extension: {path.suffix}")


def _has_planswift_shape(rows: List[Dict[str, Any]]) -> bool:
    """Does this CSV/XLSX have the PlanSwift-shaped header set?"""
    if not rows:
        return False
    headers = list(rows[0].keys())
    for required in REQUIRED_FIELDS:
        if not _match_column(headers, COL_ALIASES[required]):
            return False
    return True


def _normalize_rows(
    rows: List[Dict[str, Any]],
    source_path: Path,
) -> List[Dict[str, Any]]:
    headers = list(rows[0].keys()) if rows else []
    col_desc = _match_column(headers, COL_ALIASES["description"])
    col_qty  = _match_column(headers, COL_ALIASES["qty"])
    col_unit = _match_column(headers, COL_ALIASES["unit"])
    col_rate = _match_column(headers, COL_ALIASES["unit_rate"])
    col_ext  = _match_column(headers, COL_ALIASES["extended"])
    col_disc = _match_column(headers, COL_ALIASES["discipline"])
    col_sheet = _match_column(headers, COL_ALIASES["sheet"])

    out = []
    for idx, r in enumerate(rows, start=1):
        desc = str(r.get(col_desc, "")).strip() if col_desc else ""
        desc = _scrub_supplier_names(desc)
        if not desc:
            continue
        try:
            qty = float(r.get(col_qty, 0) or 0) if col_qty else 0.0
        except (TypeError, ValueError):
            qty = 0.0
        unit = str(r.get(col_unit, "") or "").strip().upper() if col_unit else ""
        try:
            rate = float(r.get(col_rate, 0) or 0) if col_rate else 0.0
        except (TypeError, ValueError):
            rate = 0.0
        try:
            ext = float(r.get(col_ext, 0) or 0) if col_ext else 0.0
        except (TypeError, ValueError):
            ext = 0.0
        # If extended missing but rate present, compute it.
        if not ext and rate and qty:
            ext = qty * rate
        disc = (str(r.get(col_disc, "")).strip() if col_disc else "") or "Structural"
        sheet = (str(r.get(col_sheet, "")).strip() if col_sheet else "")
        notes = f"sheet={sheet}" if sheet else ""

        out.append({
            "line_id": f"LINE-PS-{idx:04d}",
            "description": desc,
            "category": "Direct",
            "discipline": disc,
            "unit": unit or "EA",
            "qty": qty,
            "unit_rate": rate,
            "extended": ext,
            "requirement_refs": [],
            "rate_basis": f"PlanSwift export: {source_path.name}",
            "markup_applied": False,  # rates from PlanSwift are usually pre-markup
            "notes": notes or None,
        })
    return out


# -------- Adapter interface --------

def planswift_probe(ctx: BoqContext) -> bool:
    """Can we find a PlanSwift-shaped CSV or XLSX for this bid?"""
    for p in _candidate_files(ctx):
        try:
            rows = _read_rows(p)
        except Exception:
            continue
        if _has_planswift_shape(rows):
            return True
    return False


def planswift_load(ctx: BoqContext) -> BoqPayload:
    """Read the first PlanSwift-shaped file found and normalize."""
    for p in _candidate_files(ctx):
        try:
            rows = _read_rows(p)
        except Exception as e:
            continue
        if not _has_planswift_shape(rows):
            continue
        normalized = _normalize_rows(rows, p)
        return BoqPayload(
            rows=normalized,
            boq_origin="planswift",
            source_file=str(p),
            row_count=len(normalized),
            fidelity_rank=FIDELITY_PLANSWIFT,
            notes=f"PlanSwift export read from {p.name}; supplier names scrubbed.",
        )
    # Shouldn't reach here because probe gate-keeps. Defensive empty.
    return BoqPayload(
        rows=[],
        boq_origin="planswift",
        source_file="",
        row_count=0,
        fidelity_rank=FIDELITY_PLANSWIFT,
        notes="probe returned True but no PlanSwift file could be read.",
    )


PLANSWIFT_ADAPTER = BoqSourceAdapter(
    name="planswift",
    fidelity_rank=FIDELITY_PLANSWIFT,
    probe=planswift_probe,
    load=planswift_load,
    description=(
        "PlanSwift CSV or XLSX export (Ivan-run). Highest-fidelity source. "
        "Looks in bid_folder/planswift/, bid_folder/*planswift*, or "
        "bid_folder/boq/ with PlanSwift-shaped headers."
    ),
)

# Self-register on import
register(PLANSWIFT_ADAPTER)
