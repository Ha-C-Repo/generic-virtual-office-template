"""
Your Company Virtual Office - Project Controls (PC4 + PC5)

SPI, CPI, forecast-to-complete, and variance by cost code per WBS line
for awarded projects. Implements P14 (baselines or nothing), P15
(integrated WBS), PC4 (SPI/CPI calc), PC5 (internal dashboard feed).

Data sources (verify, do not generate - every number traces to a file):

  PC3 progress: table progress_log in the shop floor database. The path
      is taken from bridge/shop_floor.py so reader and writer always hit
      the same file (LOCALAPPDATA-aware, resource_path fallback).
      Expected columns per the PC3 capture spec: date, person, wbs_line,
      hours, pieces_done, tons_done, issues_text. Extra columns are
      ignored. A project column, when present, is used as a filter.

  PC1 baseline: the frozen budget xlsx under the project folder
      "Awarded Projects/<project_id> .../09 Financials -GP CONFIDENTIAL".
      The workbook must carry a BASELINE flag cell (P14: no baseline,
      no variance, no claim). The WBS sheet holds one row per WBS line
      with scope, cost code, planned units, unit, planned hours, budget
      cost, start, end, and progress type (production or milestone).

Earned value math (deterministic Python only - constitution Section 11,
the AI never does arithmetic):

  pct_complete  production line: units_done / planned_units, capped 1.0
                milestone line:  rule of credit, issued 20 / approved 75
                / released 100, read as the highest credit recorded
  planned_pct   linear time proration between start and end dates
  PV = planned_pct * budget_cost
  EV = pct_complete * budget_cost
  AC = actual_hours * budget_cost / planned_hours

  AC is labor-hours based: invoice actuals are not integrated yet, so
  cost performance is grounded in recorded shop hours priced at the
  line's budgeted cost rate. SPI = EV / PV. CPI = EV / AC. Lines below
  0.95 on either index are flagged (PC4).

Forecast control limits (Section 07, COWORK-HANDOFF-MASTER-2026-06-11):
at project level, investigate when forecast variance against baseline
is outside minus 1.7 percent / plus 7.3 percent.

Corrective action hierarchy (PC6): diagnose, find root cause, then act:
optimize first, spend second, accept third. Client-caused lines surface
as notice candidates via skills/contract-notice. Never send a notice
without the Owner's approval.

Client-cause convention: a progress row whose issues_text carries the
uppercase tag CLIENT (for example "CLIENT: GC held erection 3 days") or
the phrase "client-caused" marks its WBS line client-caused. Explicit
tag only, no fuzzy matching.

Confidence tagging: every line returns high, medium, or low. Low
confidence lines are flagged for human check, never passed silently.

CONFIDENTIAL - INTERNAL. This module feeds the internal dashboard only.
Never client-facing (PC5). Not subject to the bid-output gate, which
covers client PDFs.
"""

import logging
import os
import re
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from vo_app._resources import resource_path

log = logging.getLogger("bridge.project_controls")

FLAG_THRESHOLD = 0.95
FORECAST_LIMIT_LOW_PCT = -1.7
FORECAST_LIMIT_HIGH_PCT = 7.3
FORECAST_LIMIT_SOURCE = "Section 07 forecast control limits"
MILESTONE_CREDITS = (20.0, 75.0, 100.0)

PC6_HIERARCHY = ("Corrective hierarchy (PC6): diagnose, find root cause, "
                 "then act: optimize first, spend second, accept third.")
CLIENT_CAUSE_NOTE = ("Client-caused variance: consider notice per the "
                     "contract-admin workflow (skills/contract-notice). "
                     "Never send a notice without the Owner's approval.")

# The CLIENT tag stays uppercase-only (explicit tag, no fuzzy matching);
# the documented phrase matches in any case so "Client-caused delay" at
# the start of a sentence is not silently missed.
_CLIENT_TAG = re.compile(r"\bCLIENT\b|(?i:client[- ]caused)", re.UNICODE)

_AWARDED_ROOT = "Awarded Projects"
_FINANCIALS_PREFIX = "09"
_MAX_SCURVE_WEEKS = 520
_MAX_WBS_ROWS = 5000
_MAX_LINE_DAYS = 3650
_DATE_IN_TEXT = re.compile(r"\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4}")

# Tolerant header aliases for the PC1 WBS sheet. Keys are canonical names.
_WBS_ALIASES = {
    "wbs_line": ("wbs_line", "wbs line", "wbs", "wbs id", "line id", "line"),
    "scope": ("scope", "description", "desc"),
    "cost_code": ("cost_code", "cost code", "code"),
    "progress_type": ("progress_type", "progress type", "measure type",
                      "measure", "type"),
    "planned_units": ("planned_units", "planned units", "planned qty",
                      "plan qty", "quantity", "qty"),
    "unit": ("unit", "uom"),
    "planned_hours": ("planned_hours", "planned hours", "budget hours",
                      "plan hours", "hours"),
    "budget_cost": ("budget_cost", "budget cost", "budget", "bac",
                    "cost", "budget usd"),
    "start_date": ("start_date", "start", "planned start",
                   "schedule start"),
    "end_date": ("end_date", "end", "finish", "planned finish",
                 "schedule finish"),
    "schedule_activity": ("schedule_activity", "schedule activity",
                          "activity"),
    "quality_check": ("quality_check", "quality check", "quality"),
    "risk_note": ("risk_note", "risk note", "risk"),
}

# Tolerant column aliases for the PC3 progress_log table.
_PROGRESS_ALIASES = {
    "date": ("date", "production_date", "log_date", "entry_date"),
    "person": ("person", "worker", "worker_id", "name"),
    "wbs_line": ("wbs_line", "wbs", "wbs_id"),
    "hours": ("hours", "hours_worked"),
    "pieces_done": ("pieces_done", "pieces", "pieces_completed",
                    "units_done", "credit"),
    "tons_done": ("tons_done", "tons", "tons_fabricated"),
    "issues_text": ("issues_text", "issues", "notes"),
    "project": ("project", "project_id", "job"),
}


# -- path resolution ------------------------------------------------------

def _progress_db_path() -> Path:
    """Shop floor db path, resolved with the same logic as the PC3 writer
    (takeoff_pipeline/shop_log.py _db_path) so reads and writes always
    land on the same file. Dev: the repo data/shop_floor.db via
    resource_path (Hard Rule 2). Frozen: the LOCALAPPDATA data root, the
    same frozen-safe location _awarded_root already uses; the data/ copy
    inside sys._MEIPASS is a build-time snapshot bundled by
    VirtualOffice.spec, and reading it would silently serve frozen-in-
    time SPI/CPI as current. The bridge.shop_floor _DB import is
    shadowed by the bridge/shop_floor package today and always fails;
    the try is kept so PC3 and PC4 move together if that ever changes."""
    try:
        from bridge.shop_floor import _DB
        return Path(_DB)
    except Exception:
        pass
    try:
        from vo_app._resources import is_frozen
        frozen = is_frozen()
    except Exception:
        frozen = False
    if frozen:
        local = os.environ.get("LOCALAPPDATA")
        if local:
            return Path(local) / "YourCompany" / "VirtualOffice" / "data" \
                / "shop_floor.db"
    return Path(resource_path("data/shop_floor.db"))


def _conn(db_path: Path) -> sqlite3.Connection:
    """SQLite connection per Hard Rule 11: WAL plus busy timeout."""
    c = sqlite3.connect(str(db_path), check_same_thread=False, timeout=10)
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA busy_timeout=10000")
    c.row_factory = sqlite3.Row
    return c


def _awarded_root() -> Path:
    """Resolve the Awarded Projects root. The folder is working data that
    lives next to the installed app, never inside the PyInstaller bundle,
    so the frozen EXE must not look in sys._MEIPASS (resource_path would).
    Frozen: the EXE's directory first, then the LOCALAPPDATA data root
    bridge/shop_floor.py already uses. Dev: repo root via resource_path."""
    try:
        from vo_app._resources import is_frozen
        frozen = is_frozen()
    except Exception:
        frozen = False
    if frozen:
        cand = Path(sys.executable).resolve().parent / _AWARDED_ROOT
        if cand.is_dir():
            return cand
        local = os.environ.get("LOCALAPPDATA")
        if local:
            return Path(local) / "YourCompany" / "VirtualOffice" \
                / _AWARDED_ROOT
        return cand
    return Path(resource_path(_AWARDED_ROOT))


def _find_project_folder(project_id: str, root: Path) -> Path | None:
    """Locate <root>/<project_id> ... by name prefix, then substring."""
    if not root.is_dir():
        return None
    pid = (project_id or "").strip().lower()
    if not pid:
        return None
    # Boundary-aware: the id must not continue with another alphanumeric,
    # or PRJ-2026-ACP-001 would claim the folder of PRJ-2026-ACP-0012 when
    # its own folder is missing. The progress filter is exact-match for
    # the same reason.
    bounded = re.compile(re.escape(pid) + r"(?![0-9a-z])")
    dirs = sorted(d for d in root.iterdir() if d.is_dir())
    for d in dirs:
        if bounded.match(d.name.lower()):
            return d
    for d in dirs:
        if bounded.search(d.name.lower()):
            return d
    return None


def _find_baseline_xlsx(project_folder: Path) -> Path | None:
    """Find the PC1 baseline xlsx in the 09 Financials -GP CONFIDENTIAL
    folder. Prefer filenames containing budget or baseline."""
    fin = None
    for d in sorted(project_folder.iterdir()):
        if d.is_dir() and d.name.startswith(_FINANCIALS_PREFIX):
            fin = d
            break
    if fin is None:
        return None
    candidates = [p for p in sorted(fin.glob("*.xlsx"))
                  if not p.name.startswith("~$")]
    if not candidates:
        return None
    for p in candidates:
        low = p.name.lower()
        if "budget" in low or "baseline" in low:
            return p
    return candidates[0]


# -- value coercion -------------------------------------------------------

def _num(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _match_alias(header: str, aliases: dict) -> str | None:
    h = str(header or "").strip().lower().replace("_", " ")
    for canonical, names in aliases.items():
        for n in names:
            if h == n.replace("_", " "):
                return canonical
    return None


# -- PC1 baseline loader --------------------------------------------------

def _load_baseline(xlsx_path: Path) -> dict:
    """Parse the PC1 baseline xlsx. Returns lines plus the BASELINE flag.
    On failure returns {"error": ..., "fix": ...}."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl is not installed; cannot read the PC1 "
                         "baseline xlsx",
                "fix": "py -3.13 -m pip install openpyxl"}
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True,
                                    read_only=True)
    except Exception as e:
        return {"error": f"could not open baseline xlsx {xlsx_path.name}: {e}",
                "fix": "confirm the file is a valid xlsx and not open in Excel"}
    try:
        baseline_flag = False
        frozen_date = ""
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=40, max_col=15):
                for cell in row:
                    val = cell.value
                    # P14 flag cell is explicit: exactly BASELINE, or
                    # BASELINE plus a freeze marker ("BASELINE frozen
                    # 2026-05-01"). A label that merely contains or starts
                    # with the word (a "Baseline Hours" column header, a
                    # "cost baseline" title) does not freeze a draft.
                    up = val.strip().upper() if isinstance(val, str) else ""
                    if up == "BASELINE" or (up.startswith("BASELINE") and
                            ("FROZEN" in up or _DATE_IN_TEXT.search(up))):
                        baseline_flag = True
                        for sib in row:
                            d = _to_date(sib.value)
                            if d:
                                frozen_date = d.isoformat()
                                break
                if baseline_flag:
                    break
            if baseline_flag:
                break

        wbs_ws = None
        for ws in wb.worksheets:
            if "wbs" in ws.title.lower():
                wbs_ws = ws
                break
        if wbs_ws is None and len(wb.worksheets) > 1:
            wbs_ws = wb.worksheets[1]
        if wbs_ws is None:
            return {"error": f"no WBS sheet found in {xlsx_path.name}",
                    "fix": "add the P15 WBS sheet to the baseline xlsx"}

        header_map = {}
        header_row_idx = None
        for r_idx, row in enumerate(
                wbs_ws.iter_rows(min_row=1, max_row=10), start=1):
            candidate = {}
            for c_idx, cell in enumerate(row):
                canonical = _match_alias(cell.value, _WBS_ALIASES)
                if canonical and canonical not in candidate:
                    candidate[canonical] = c_idx
            if len(candidate) >= 3 and "wbs_line" in candidate:
                header_map = candidate
                header_row_idx = r_idx
                break
        if not header_map:
            return {"error": f"no recognizable WBS header row in "
                             f"{xlsx_path.name} sheet {wbs_ws.title}",
                    "fix": "WBS sheet needs columns such as WBS Line, Cost "
                           "Code, Planned Units, Planned Hours, Budget Cost"}

        lines = []
        warnings = []
        empty_streak = 0
        # Bounded iteration: openpyxl ghost dimensions can report millions
        # of phantom rows; these are synchronous Bridge calls and must not
        # hang the UI. Hard cap plus an empty-row streak break.
        # Columns bounded too: ghost dimensions can report thousands of
        # phantom columns per row; only the mapped header columns matter.
        last_col = max(header_map.values()) + 1
        for row in wbs_ws.iter_rows(min_row=header_row_idx + 1,
                                    max_row=header_row_idx + _MAX_WBS_ROWS,
                                    max_col=last_col):
            cells = [c.value for c in row]

            def pick(name):
                idx = header_map.get(name)
                return cells[idx] if idx is not None and idx < len(cells) \
                    else None

            wbs_line = str(pick("wbs_line") or "").strip()
            if not wbs_line:
                empty_streak += 1
                if empty_streak >= 100:
                    break
                continue
            empty_streak = 0
            ptype = str(pick("progress_type") or "").strip().lower()
            if ptype not in ("production", "milestone"):
                if ptype:
                    warnings.append(f"{wbs_line}: unknown progress type "
                                    f"'{ptype}', treated as production")
                ptype = ptype or ""
            start = _to_date(pick("start_date"))
            end = _to_date(pick("end_date"))
            if start and end and (end - start).days > _MAX_LINE_DAYS:
                # A 12/31/9999-style placeholder would otherwise drive a
                # multi-million-day planned spread in the S-curve.
                warnings.append(f"{wbs_line}: schedule span over 10 years "
                                "looks like a placeholder; dates ignored")
                start = end = None
            lines.append({
                "wbs_line": wbs_line,
                "scope": str(pick("scope") or "").strip(),
                "cost_code": str(pick("cost_code") or "").strip()
                or "uncoded",
                "progress_type": ptype,
                "planned_units": _num(pick("planned_units")),
                "unit": str(pick("unit") or "").strip(),
                "planned_hours": _num(pick("planned_hours")),
                "budget_cost": _num(pick("budget_cost")),
                "start_date": start,
                "end_date": end,
                "schedule_activity": str(pick("schedule_activity") or "")
                .strip(),
                "quality_check": str(pick("quality_check") or "").strip(),
                "risk_note": str(pick("risk_note") or "").strip(),
            })
        if not lines:
            return {"error": f"WBS sheet in {xlsx_path.name} has no data "
                             "rows",
                    "fix": "populate the P15 WBS sheet, one row per line"}
        return {"lines": lines, "baseline_flag": baseline_flag,
                "frozen_date": frozen_date, "path": str(xlsx_path),
                "warnings": warnings}
    finally:
        wb.close()


# -- PC3 progress loader --------------------------------------------------

def _load_progress(db_path: Path, project_id: str) -> dict:
    """Read the PC3 progress_log table. Returns normalized rows. On a
    missing table or db returns {"error": ..., "fix": ...}."""
    if not Path(db_path).exists():
        return {"error": f"shop floor database not found at {db_path}",
                "fix": "run PC3 shop progress capture first; the shop log "
                       "creates the database"}
    c = _conn(Path(db_path))
    try:
        found = c.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name='progress_log'").fetchone()
        if not found:
            return {"error": "PC3 progress table progress_log not found in "
                             f"{Path(db_path).name}",
                    "fix": "run PC3 shop progress capture (shop daily log) "
                           "so the progress_log table exists with data"}
        cols = [r["name"] for r in
                c.execute("PRAGMA table_info(progress_log)")]
        colmap = {}
        for col in cols:
            canonical = _match_alias(col, _PROGRESS_ALIASES)
            if canonical and canonical not in colmap:
                colmap[canonical] = col
        if "wbs_line" not in colmap:
            return {"error": "progress_log has no wbs_line column; per-WBS "
                             "SPI/CPI needs it",
                    "fix": "add wbs_line to the PC3 capture per the shop "
                           "log spec"}
        raw = c.execute("SELECT * FROM progress_log").fetchall()
    finally:
        c.close()

    pid = (project_id or "").strip().lower()
    has_project_col = "project" in colmap
    rows = []
    unattributed = 0
    for r in raw:
        rec = dict(r)
        if has_project_col and pid:
            proj = str(rec.get(colmap["project"]) or "").strip().lower()
            if not proj:
                # No attribution: excluding beats double-counting the row
                # into every project's AC, EV, and S-curve.
                unattributed += 1
                continue
            # Exact match only. Substring matching would pull
            # PRJ-2026-ACP-0012 rows into PRJ-2026-ACP-001.
            if proj != pid:
                continue
        rows.append({
            "date": _to_date(rec.get(colmap.get("date", ""), None)),
            "person": str(rec.get(colmap.get("person", ""), "") or ""),
            "wbs_line": str(rec.get(colmap["wbs_line"]) or "").strip(),
            "hours": _num(rec.get(colmap.get("hours", ""), None)) or 0.0,
            "pieces_done": _num(rec.get(colmap.get("pieces_done", ""),
                                        None)) or 0.0,
            "tons_done": _num(rec.get(colmap.get("tons_done", ""),
                                      None)) or 0.0,
            "issues_text": str(rec.get(colmap.get("issues_text", ""), "")
                               or ""),
        })
    warnings = []
    if unattributed:
        warnings.append(f"{unattributed} progress rows have no project "
                        "attribution; excluded")
    if has_project_col and pid and raw and not rows:
        warnings.append("progress_log has rows but none match project "
                        f"'{project_id}' exactly; check the project value "
                        "used in the shop log")
    if not has_project_col and raw:
        warnings.append("progress_log has no project column; rows are "
                        "scoped only by baseline WBS line membership")
    return {"rows": rows, "path": str(db_path), "warnings": warnings}


# -- earned value math ----------------------------------------------------

def _units_for_row(row: dict, unit: str) -> float:
    """Map a progress row to the line's unit of measure. Tons lines read
    tons_done; everything else (pieces, EA, SF) reads pieces_done. The SF
    convention (SF logged in pieces_done) downgrades confidence."""
    if "ton" in unit:
        return row["tons_done"]
    return row["pieces_done"]


def _planned_pct(start: date | None, end: date | None,
                 as_of: date) -> float | None:
    """Linear time proration of the plan between start and end."""
    if not start or not end or end < start:
        return None
    if as_of < start:
        return 0.0
    if as_of >= end:
        return 1.0
    total = (end - start).days
    if total <= 0:
        return 1.0
    return (as_of - start).days / total


def _safe_ratio(num: float | None, den: float | None) -> float | None:
    if num is None or den is None or den <= 0:
        return None
    return round(num / den, 3)


def _line_metrics(line: dict, rows: list, as_of: date) -> dict:
    """Compute earned value metrics for one WBS line. Pure arithmetic on
    baseline and progress data; nothing here is generated."""
    unit = (line.get("unit") or "").lower()
    ptype = line.get("progress_type") or "production"
    planned_units = line.get("planned_units")
    planned_hours = line.get("planned_hours")
    budget_cost = line.get("budget_cost")
    # Rows without a parseable date cannot honor the as_of cutoff, so they
    # are excluded from every metric (not silently counted as current).
    undated = [r for r in rows if r["date"] is None]
    dated = [r for r in rows
             if r["date"] is not None and r["date"] <= as_of]
    actual_hours = sum(r["hours"] for r in dated)
    client_caused = any(_CLIENT_TAG.search(r["issues_text"]) for r in dated)
    client_evidence = [r["issues_text"] for r in dated
                       if _CLIENT_TAG.search(r["issues_text"])]

    issues = []
    confidence = "high"
    if undated:
        issues.append(f"{len(undated)} progress rows have no parseable "
                      "date; excluded from all metrics")
        confidence = "medium"

    if ptype == "milestone":
        credit = max((r["pieces_done"] for r in dated), default=0.0)
        if credit and credit not in MILESTONE_CREDITS:
            issues.append(f"milestone credit {credit:g} is not one of the "
                          "rule-of-credit steps 20/75/100")
            confidence = "medium"
        pct_complete = min(credit / 100.0, 1.0) if credit else 0.0
        units_done = credit
    else:
        units_done = sum(_units_for_row(r, unit) for r in dated)
        if planned_units and planned_units > 0:
            pct_complete = min(units_done / planned_units, 1.0)
        else:
            pct_complete = None
            issues.append("planned units missing; percent complete not "
                          "computable")
            confidence = "low"
        if unit in ("sf", "sq ft", "sqft"):
            issues.append("SF progress read from pieces_done per the PC3 "
                          "capture convention")
            if confidence == "high":
                confidence = "medium"

    if not line.get("progress_type"):
        issues.append("progress type missing in baseline; assumed "
                      "production")
        if confidence == "high":
            confidence = "medium"

    p_pct = _planned_pct(line.get("start_date"), line.get("end_date"), as_of)
    if p_pct is None:
        issues.append("start or end date missing; SPI not computable")
        if confidence == "high":
            confidence = "medium"

    if budget_cost is None or budget_cost <= 0:
        issues.append("budget cost missing for this line")
        confidence = "low"
    if not dated:
        issues.append("no progress rows recorded yet for this line")
        if confidence == "high":
            confidence = "medium"

    pv = ev = ac = None
    if budget_cost and budget_cost > 0:
        if p_pct is not None:
            pv = round(p_pct * budget_cost, 2)
        if pct_complete is not None:
            ev = round(pct_complete * budget_cost, 2)
        if planned_hours and planned_hours > 0:
            ac = round(actual_hours * budget_cost / planned_hours, 2)
        elif actual_hours:
            issues.append("planned hours missing; actual cost not "
                          "computable from hours")
            confidence = "low"

    if ev is not None and ev > 0 and ac == 0.0:
        # Units cannot be produced with zero hours; the hours are missing,
        # not free. The rollup excludes this line from cost performance.
        issues.append("units logged but no hours recorded; line excluded "
                      "from cost performance")
        confidence = "low"

    spi = _safe_ratio(ev, pv)
    cpi = _safe_ratio(ev, ac)

    flagged = ((spi is not None and spi < FLAG_THRESHOLD)
               or (cpi is not None and cpi < FLAG_THRESHOLD))

    return {
        "wbs_line": line["wbs_line"],
        "scope": line.get("scope", ""),
        "cost_code": line.get("cost_code", "uncoded"),
        "progress_type": ptype,
        "unit": line.get("unit", ""),
        "planned_units": planned_units,
        "units_done": round(units_done, 3),
        "planned_hours": planned_hours,
        "actual_hours": round(actual_hours, 2),
        "budget_cost": budget_cost,
        "pct_complete": round(pct_complete, 4)
        if pct_complete is not None else None,
        "planned_pct": round(p_pct, 4) if p_pct is not None else None,
        "pv": pv, "ev": ev, "ac": ac,
        "spi": spi, "cpi": cpi,
        "flagged": flagged,
        "client_caused": client_caused,
        "client_evidence": client_evidence[:3],
        "confidence": confidence,
        "issues": issues,
        "risk_note": line.get("risk_note", ""),
    }


def _scurve(lines: list, rows_by_line: dict, as_of: date) -> list:
    """Weekly cumulative planned vs earned vs actual cost series. Planned
    spreads each line's budget cost linearly between its dates; earned and
    actual accumulate from progress rows priced at baseline rates."""
    daily_planned = {}
    dates = []
    for ln in lines:
        start, end = ln.get("start_date"), ln.get("end_date")
        cost = ln.get("budget_cost")
        if not start or not end or end < start or not cost:
            continue
        days = (end - start).days
        if days <= 0:
            daily_planned[end] = daily_planned.get(end, 0.0) + cost
        else:
            # Accrue on start+1 .. end so the cumulative planned value at
            # any date equals _planned_pct(date) * cost exactly; the
            # curve and the PV column then use one day-count convention.
            per_day = cost / days
            d = start + timedelta(days=1)
            while d <= end:
                daily_planned[d] = daily_planned.get(d, 0.0) + per_day
                d += timedelta(days=1)
        dates.extend((start, end))

    daily_earned = {}
    daily_actual = {}
    by_key = {ln["wbs_line"]: ln for ln in lines}
    for key, rows in rows_by_line.items():
        ln = by_key.get(key)
        if ln is None:
            continue
        cost = ln.get("budget_cost")
        p_units = ln.get("planned_units")
        p_hours = ln.get("planned_hours")
        unit = (ln.get("unit") or "").lower()
        ptype = ln.get("progress_type") or "production"
        prev_credit = 0.0
        cum_units = 0.0
        for r in sorted(rows, key=lambda x: x["date"] or as_of):
            d = r["date"]
            if d is None or d > as_of:
                continue
            dates.append(d)
            if cost and p_hours and p_hours > 0 and r["hours"]:
                daily_actual[d] = daily_actual.get(d, 0.0) \
                    + r["hours"] * cost / p_hours
            if cost:
                if ptype == "milestone":
                    credit = min(r["pieces_done"], 100.0)
                    if credit > prev_credit:
                        daily_earned[d] = daily_earned.get(d, 0.0) \
                            + (credit - prev_credit) / 100.0 * cost
                        prev_credit = credit
                elif p_units and p_units > 0:
                    # Cap cumulative earned at the line budget so the
                    # curve agrees with the capped line EV (an over-run
                    # or double scan must not push EARNED past PLANNED
                    # at completion).
                    prev_eff = min(cum_units, p_units)
                    cum_units += max(0.0, _units_for_row(r, unit))
                    eff = min(cum_units, p_units)
                    if eff > prev_eff:
                        daily_earned[d] = daily_earned.get(d, 0.0) \
                            + (eff - prev_eff) / p_units * cost

    if not dates:
        return []
    lo, hi = min(dates), max(max(dates), as_of)
    week_end = lo + timedelta(days=(6 - lo.weekday()))
    series = []
    cum_p = cum_e = cum_a = 0.0
    cursor = lo
    weeks = 0
    while week_end - timedelta(days=6) <= hi and weeks < _MAX_SCURVE_WEEKS:
        while cursor <= week_end:
            cum_p += daily_planned.get(cursor, 0.0)
            cum_e += daily_earned.get(cursor, 0.0)
            cum_a += daily_actual.get(cursor, 0.0)
            cursor += timedelta(days=1)
        past = week_end <= as_of
        series.append({
            "week_ending": week_end.isoformat(),
            "planned": round(cum_p, 2),
            "earned": round(cum_e, 2) if past else None,
            "actual": round(cum_a, 2) if past else None,
        })
        week_end += timedelta(days=7)
        weeks += 1
    return series


# -- shared assembly ------------------------------------------------------

def _assemble(project_id: str, db_path=None, baseline_path=None,
              as_of=None) -> dict:
    """Resolve sources, load baseline plus progress, compute line metrics.
    Returns a context dict, or {"error": ..., "fix": ...}."""
    pid = (project_id or "").strip()
    if not pid:
        return {"error": "project_id is required",
                "fix": "pass the awarded project code, for example "
                       "PRJ-2026-ACP-001"}
    as_of = _to_date(as_of) or date.today()

    if baseline_path:
        xlsx = Path(baseline_path)
        if not xlsx.exists():
            return {"error": f"baseline xlsx not found at {xlsx}",
                    "fix": "check the path"}
    else:
        root = _awarded_root()
        if not root.is_dir():
            # Distinct from a wrong project code: the root itself is
            # missing (fresh install, frozen EXE on a new machine).
            return {"error": f"Awarded Projects root not found at {root}",
                    "fix": "create the Awarded Projects folder next to "
                           "the app (repo root in dev) and add the "
                           "project folder inside it"}
        folder = _find_project_folder(pid, root)
        if folder is None:
            return {"error": f"no awarded project folder matches "
                             f"'{pid}' under {root}",
                    "fix": "use the folder's project code, for example "
                           "PRJ-2026-ACP-001"}
        xlsx = _find_baseline_xlsx(folder)
        if xlsx is None:
            return {"error": "no PC1 baseline budget xlsx found in "
                             f"{folder.name}/09 Financials -GP CONFIDENTIAL",
                    "fix": "run the award-to-budget conversion (PC1) to "
                           "produce the frozen baseline first; P14: no "
                           "baseline, no variance"}

    baseline = _load_baseline(xlsx)
    if "error" in baseline:
        return baseline
    if not baseline["baseline_flag"]:
        return {"error": f"{Path(xlsx).name} has no BASELINE flag cell; "
                         "the budget is not frozen",
                "fix": "freeze the budget per P14 (no baseline, no "
                       "variance); edits after freeze need a new version"}

    progress = _load_progress(db_path or _progress_db_path(), pid)
    if "error" in progress:
        return progress

    rows_by_line = {}
    for r in progress["rows"]:
        rows_by_line.setdefault(r["wbs_line"], []).append(r)

    metrics = [_line_metrics(ln, rows_by_line.get(ln["wbs_line"], []),
                             as_of) for ln in baseline["lines"]]

    known = set(rows_by_line) - {ln["wbs_line"] for ln in baseline["lines"]}
    warnings = list(baseline["warnings"]) + list(progress["warnings"])
    if known:
        warnings.append("progress rows reference WBS lines not in the "
                        "baseline: " + ", ".join(sorted(known)[:10]))

    return {
        "project_id": pid,
        "as_of": as_of.isoformat(),
        "as_of_date": as_of,
        "lines": metrics,
        "baseline_lines": baseline["lines"],
        "rows_by_line": rows_by_line,
        "frozen_date": baseline["frozen_date"],
        "warnings": warnings,
        "data_sources": {
            "pc1_baseline_xlsx": baseline["path"],
            "pc3_progress_db": progress["path"],
            "cost_basis": "labor hours priced at baseline line rates; "
                          "invoice actuals not integrated",
        },
    }


def _flag_entries(metrics: list) -> list:
    """Build the dashboard flag list: performance flags (below 0.95) plus
    data flags (low confidence, for human check)."""
    flags = []
    for m in metrics:
        notes = [PC6_HIERARCHY]
        if m["client_caused"]:
            notes.append(CLIENT_CAUSE_NOTE)
        if m["flagged"]:
            reasons = []
            if m["spi"] is not None and m["spi"] < FLAG_THRESHOLD:
                reasons.append(f"SPI {m['spi']} below {FLAG_THRESHOLD}")
            if m["cpi"] is not None and m["cpi"] < FLAG_THRESHOLD:
                reasons.append(f"CPI {m['cpi']} below {FLAG_THRESHOLD}")
            flags.append({"wbs_line": m["wbs_line"], "scope": m["scope"],
                          "cost_code": m["cost_code"], "type": "performance",
                          "spi": m["spi"], "cpi": m["cpi"],
                          "reason": "; ".join(reasons),
                          "client_caused": m["client_caused"],
                          "client_evidence": m["client_evidence"],
                          "confidence": m["confidence"], "notes": notes})
        elif m["confidence"] == "low":
            flags.append({"wbs_line": m["wbs_line"], "scope": m["scope"],
                          "cost_code": m["cost_code"], "type": "data",
                          "spi": m["spi"], "cpi": m["cpi"],
                          "reason": "low confidence: " +
                                    "; ".join(m["issues"]),
                          "client_caused": m["client_caused"],
                          "client_evidence": m["client_evidence"],
                          "confidence": "low", "notes": notes})
    return flags


def _rollup(metrics: list) -> dict:
    """Project / cost-code rollup. SPI and CPI use matched-pair sums: a
    line enters a ratio only when both of its terms are computable.
    Summing asymmetrically would let a line with EV but no AC inflate the
    project CPI (its earned value enters the numerator with nothing in
    the denominator). Excluded lines are named, never dropped silently."""
    pv = sum(m["pv"] for m in metrics if m["pv"] is not None)
    ev = sum(m["ev"] for m in metrics if m["ev"] is not None)
    ac = sum(m["ac"] for m in metrics if m["ac"] is not None)
    bac = sum(m["budget_cost"] for m in metrics
              if m["budget_cost"] is not None)
    spi_ev = spi_pv = cpi_ev = cpi_ac = 0.0
    excluded_spi, excluded_cpi = [], []
    for m in metrics:
        if m["ev"] is not None and m["pv"] is not None:
            spi_ev += m["ev"]
            spi_pv += m["pv"]
        else:
            excluded_spi.append(m["wbs_line"])
        if (m["ev"] is not None and m["ac"] is not None
                and not (m["ev"] > 0 and m["ac"] == 0)):
            cpi_ev += m["ev"]
            cpi_ac += m["ac"]
        else:
            excluded_cpi.append(m["wbs_line"])
    return {"bac": round(bac, 2), "pv": round(pv, 2), "ev": round(ev, 2),
            "ac": round(ac, 2), "spi": _safe_ratio(spi_ev, spi_pv),
            "cpi": _safe_ratio(cpi_ev, cpi_ac),
            "excluded_from_spi": excluded_spi,
            "excluded_from_cpi": excluded_cpi}


def _control_status(variance_pct: float) -> str:
    """Section 07: investigate outside minus 1.7 / plus 7.3 percent."""
    if variance_pct < FORECAST_LIMIT_LOW_PCT \
            or variance_pct > FORECAST_LIMIT_HIGH_PCT:
        return "INVESTIGATE"
    return "WITHIN LIMITS"


# -- public API (wrapped by Bridge methods in bridge/api.py) ---------------

def spi_cpi(project_id: str, db_path=None, baseline_path=None,
            as_of=None) -> dict:
    """SPI and CPI per WBS line plus project rollup, flags, and the weekly
    S-curve series for the internal dashboard (PC4 + PC5)."""
    ctx = _assemble(project_id, db_path, baseline_path, as_of)
    if "error" in ctx:
        return ctx
    metrics = ctx["lines"]
    confidence = {"high": 0, "medium": 0, "low": 0}
    for m in metrics:
        confidence[m["confidence"]] += 1
    roll = _rollup(metrics)
    warnings = list(ctx["warnings"])
    for key, label in (("excluded_from_spi", "SPI"),
                       ("excluded_from_cpi", "CPI")):
        if roll[key]:
            warnings.append(f"project {label} excludes lines with "
                            "incomplete data: " + ", ".join(roll[key]))
    return {
        "project_id": ctx["project_id"],
        "as_of": ctx["as_of"],
        "baseline_frozen": ctx["frozen_date"],
        "project": roll,
        "flag_threshold": FLAG_THRESHOLD,
        "lines": metrics,
        "flags": _flag_entries(metrics),
        "scurve": _scurve(ctx["baseline_lines"], ctx["rows_by_line"],
                          ctx["as_of_date"]),
        "confidence_counts": confidence,
        "pc6_hierarchy": PC6_HIERARCHY,
        "warnings": warnings,
        "data_sources": ctx["data_sources"],
        "classification": "CONFIDENTIAL - INTERNAL",
    }


def forecast_to_complete(project_id: str, db_path=None, baseline_path=None,
                         as_of=None) -> dict:
    """Forecast at completion per line (EAC = BAC / CPI) rolled to project
    level, checked against the Section 07 control limits."""
    ctx = _assemble(project_id, db_path, baseline_path, as_of)
    if "error" in ctx:
        return ctx
    lines = []
    project_bac = project_eac = project_ac = 0.0
    excluded = []
    for m in ctx["lines"]:
        bac = m["budget_cost"]
        ev, ac = m["ev"], m["ac"]
        if bac is None or bac <= 0:
            lines.append({"wbs_line": m["wbs_line"],
                          "cost_code": m["cost_code"], "bac": bac,
                          "ac": ac, "cpi": m["cpi"],
                          "eac": None, "etc": None,
                          "confidence": "low",
                          "note": "budget cost missing; excluded from the "
                                  "project forecast"})
            excluded.append(m["wbs_line"])
            continue
        if ev is not None and ev > 0 and ac is not None and ac > 0:
            # EAC from the raw ratio, not the display-rounded CPI:
            # EAC = BAC / (EV/AC) = BAC * AC / EV.
            eac = round(bac * ac / ev, 2)
            conf = m["confidence"]
            note = ""
        elif ac is not None and ac > 0:
            # Hours burned with zero earned value. BAC / CPI is undefined
            # and holding EAC at BAC would hide the worst case; surface
            # it instead of forecasting it.
            lines.append({"wbs_line": m["wbs_line"],
                          "cost_code": m["cost_code"], "bac": bac,
                          "ac": ac, "cpi": m["cpi"],
                          "eac": None, "etc": None, "confidence": "low",
                          "note": "hours recorded with zero earned value; "
                                  "line forecast not computable, "
                                  "investigate"})
            excluded.append(m["wbs_line"])
            continue
        else:
            eac = bac
            conf = "low"
            if ev is not None and ev > 0:
                # Units logged with zero hours: data exists but the cost
                # side is missing, which is not the same as no data yet.
                note = ("units logged but no hours recorded; cost "
                        "performance unknown, EAC held at BAC")
            else:
                note = "no cost performance data yet; EAC held at BAC"
        etc = round(eac - (ac or 0.0), 2)
        lines.append({"wbs_line": m["wbs_line"], "cost_code": m["cost_code"],
                      "bac": bac, "ac": ac, "cpi": m["cpi"],
                      "eac": eac, "etc": etc, "confidence": conf,
                      "note": note})
        project_bac += bac
        project_eac += eac
        project_ac += ac or 0.0
    if project_bac <= 0:
        return {"error": "no WBS line has a computable forecast (budget "
                         "cost missing or zero earned value on every "
                         "line)",
                "fix": "populate budget cost in the PC1 baseline xlsx and "
                       "check the flagged lines"}
    warnings = list(ctx["warnings"])
    if excluded:
        warnings.append("project forecast excludes lines: "
                        + ", ".join(excluded))
    raw_variance_pct = (project_eac - project_bac) / project_bac * 100
    variance_pct = round(raw_variance_pct, 2)
    # Status from the raw value: rounding first would pull a variance
    # within half a basis point of a control limit onto the boundary.
    status = _control_status(raw_variance_pct)
    out = {
        "project_id": ctx["project_id"],
        "as_of": ctx["as_of"],
        "lines": lines,
        "project": {
            "bac": round(project_bac, 2),
            "eac": round(project_eac, 2),
            "etc": round(project_eac - project_ac, 2),
            "forecast_variance_pct": variance_pct,
            "control_limits": {"low_pct": FORECAST_LIMIT_LOW_PCT,
                               "high_pct": FORECAST_LIMIT_HIGH_PCT,
                               "source": FORECAST_LIMIT_SOURCE},
            "status": status,
        },
        "warnings": warnings,
        "data_sources": ctx["data_sources"],
        "classification": "CONFIDENTIAL - INTERNAL",
    }
    if status == "INVESTIGATE":
        out["project"]["action"] = PC6_HIERARCHY
    return out


def variance_by_cost_code(project_id: str, db_path=None, baseline_path=None,
                          as_of=None) -> dict:
    """Cost and schedule variance grouped by cost code, with client-caused
    lines carrying the contract-admin notice note (PC6)."""
    ctx = _assemble(project_id, db_path, baseline_path, as_of)
    if "error" in ctx:
        return ctx
    groups = {}
    for m in ctx["lines"]:
        groups.setdefault(m["cost_code"], []).append(m)
    codes = []
    for code in sorted(groups):
        ms = groups[code]
        roll = _rollup(ms)
        sv = round(roll["ev"] - roll["pv"], 2)
        cv = round(roll["ev"] - roll["ac"], 2)
        client = any(m["client_caused"] for m in ms)
        flagged = ((roll["spi"] is not None
                    and roll["spi"] < FLAG_THRESHOLD)
                   or (roll["cpi"] is not None
                       and roll["cpi"] < FLAG_THRESHOLD))
        notes = []
        if flagged:
            notes.append(PC6_HIERARCHY)
        if client:
            notes.append(CLIENT_CAUSE_NOTE)
        codes.append({
            "cost_code": code, "bac": roll["bac"], "pv": roll["pv"],
            "ev": roll["ev"], "ac": roll["ac"],
            "schedule_variance": sv, "cost_variance": cv,
            "sv_pct": round(sv / roll["pv"] * 100, 2) if roll["pv"] else
            None,
            "cv_pct": round(cv / roll["ac"] * 100, 2) if roll["ac"] else
            None,
            "spi": roll["spi"], "cpi": roll["cpi"], "flagged": flagged,
            "client_caused": client,
            "wbs_lines": [m["wbs_line"] for m in ms],
            "confidence": min((m["confidence"] for m in ms),
                              key=("low", "medium", "high").index),
            "notes": notes,
        })
    roll = _rollup(ctx["lines"])
    warnings = list(ctx["warnings"])
    for key, label in (("excluded_from_spi", "SPI"),
                       ("excluded_from_cpi", "CPI")):
        if roll[key]:
            warnings.append(f"project {label} excludes lines with "
                            "incomplete data: " + ", ".join(roll[key]))
    return {
        "project_id": ctx["project_id"],
        "as_of": ctx["as_of"],
        "project": roll,
        "codes": codes,
        "flag_threshold": FLAG_THRESHOLD,
        "pc6_hierarchy": PC6_HIERARCHY,
        "warnings": warnings,
        "data_sources": ctx["data_sources"],
        "classification": "CONFIDENTIAL - INTERNAL",
    }
