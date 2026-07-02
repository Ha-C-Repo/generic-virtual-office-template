"""QA overlay: Ivan's verification surface (Prompt 6).

Takes a drawing PDF plus census.db and writes a copy with PyMuPDF
annotation layers, saved alongside the original with the suffix _QA.
Every census hit gets a rectangle annotation at its bbox with a popup
note (designation, count contribution, confidence, primary_source).
Low-confidence and CONFLICT items are additionally listed on a summary
page appended to the PDF so Ivan can work the exception list top to
bottom (P23/P24); when the takeoff xlsx is supplied, manual-entry low
rows (which have no drawable bbox) join that list per schema section 6.
Congestion zones (more hits per square inch than the threshold) get a
zone flag annotation per P24: a dense cluster is where a text census
miscounts, so it is flagged for eyes, never trusted silently.

The output is INTERNAL ONLY. It is Ivan's working surface and is never
sent to a client; every annotated page carries a small INTERNAL QA
marker and the summary page says so in red. Annotations are standard
PDF rect and freetext annots so the file opens in any ordinary reader.

Free tooling only (PyMuPDF, openpyxl). No em-dashes in any generated
text. Standalone: no bridge/ imports; paths are package-relative, and
the census.db default matches the census module.
"""

import json
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from takeoff_pipeline import sheet_router

_PKG = Path(__file__).resolve().parent
DEFAULT_DB = _PKG / "census.db"
# Backup landing zone per the repo operating rule; module-level so
# tests can point it at a sandbox instead of the live changelog.
HANDOFF_ROOT = _PKG.parent / "_handoff"

# One palette, defined once. RGB floats 0..1 for PyMuPDF. Red is
# reserved for conflicts; nothing else may use it.
PALETTE = {
    "COL": (0.00, 0.35, 0.95),      # blue
    "BEAM": (0.00, 0.60, 0.20),     # green
    "JST": (0.95, 0.55, 0.00),      # orange
    "PLATE": (0.55, 0.10, 0.75),    # purple
    "DECK": (0.00, 0.55, 0.55),     # teal
    "ANCH": (0.45, 0.30, 0.10),     # brown
    "MISC": (0.40, 0.40, 0.40),     # gray
    "CONFLICT": (0.85, 0.05, 0.05),  # red, conflicts only
    "CONGESTION": (0.90, 0.70, 0.05),  # amber zone flags
}

# P24 congestion: flag any square inch of sheet carrying more hits
# than this. 72 PDF points = 1 inch.
CONGESTION_THRESHOLD_PER_SQIN = 8
_CELL_PT = 72.0

_INTERNAL_MARK = "INTERNAL QA OVERLAY - NOT FOR ISSUE"
_SUMMARY_TITLE = "QA EXCEPTION LIST - INTERNAL ONLY"
_LINES_PER_SUMMARY_PAGE = 44


def _load_hits(job, db_path=None):
    """census.db is READ-ONLY input: the connection is opened mode=ro
    so the overlay can neither rewrite the journal header nor crash on
    a read-only file attribute."""
    path = Path(db_path) if db_path else DEFAULT_DB
    c = sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True,
                        check_same_thread=False, timeout=10)
    c.execute("PRAGMA busy_timeout=10000")
    c.row_factory = sqlite3.Row
    try:
        hits = [dict(r) for r in c.execute(
            "SELECT * FROM census_hits WHERE job = ? ORDER BY id",
            (job,))]
        conflicts = {r["conflict_group"]: dict(r) for r in c.execute(
            "SELECT * FROM conflicts WHERE job = ?", (job,))}
    finally:
        c.close()
    return hits, conflicts


def _page_map(pdf_path, router_result=None):
    """(sheet number -> page index, ambiguous sheet numbers). A sheet
    number the router reports on MORE than one page is ambiguous: hits
    recorded under it cannot be placed trustworthily, so they route to
    the exception list instead of silently landing on the first page.
    PAGE-<n> fallback names map straight to their index."""
    router_result = router_result or sheet_router.route(pdf_path)
    mapping = {}
    ambiguous = set()
    for entry in router_result["sheets"]:
        num = entry["sheet_number"]
        if num:
            if num in mapping and mapping[num] != entry["page_index"]:
                ambiguous.add(num)
            else:
                mapping.setdefault(num, entry["page_index"])
        mapping.setdefault(f"PAGE-{entry['page_index']}",
                           entry["page_index"])
    for num in ambiguous:
        mapping.pop(num, None)
    return mapping, ambiguous


def _hit_color(hit) -> tuple:
    if hit["conflict_group"]:
        return PALETTE["CONFLICT"]
    return PALETTE.get(hit["item_class"], PALETTE["MISC"])


_SCHEDULE_PRIMARY_CLASSES = ("COL", "JST", "ANCH", "PLATE")


def _norm_designation(text: str) -> str:
    return re.sub(r"[\s\"']", "", (text or "").upper())


def _count_contribution(hit, totals) -> str:
    """Wording follows CLASS PRIMACY (schema section 5 / Appendix A),
    not just source kind: a schedule row of a plan-primary class is
    the cross-check, not the count basis, and the popup must not tell
    Ivan to verify the wrong side."""
    key = (_norm_designation(hit["designation"]), hit["source_kind"])
    total = totals.get(key, 1)
    sched_primary = hit["item_class"] in _SCHEDULE_PRIMARY_CLASSES
    if hit["source_kind"] == "SCHEDULE":
        if not sched_primary:
            return "schedule row, cross-check only, not the count basis"
        if hit["qty"] is not None:
            return f"schedule row, qty {hit['qty']:g}"
        return f"schedule row 1 of {total} (counted, QTY_BASIS min)"
    if sched_primary:
        return (f"plan callout 1 of {total}, cross-check only, "
                "not the count basis")
    return f"plan callout 1 of {total} for this designation"


def _note_text(hit, totals) -> str:
    lines = [
        hit["designation"],
        _count_contribution(hit, totals),
        f"confidence: {hit['confidence']}",
        f"source: {hit['primary_source']}",
    ]
    if hit["conflict_group"]:
        lines.append("CONFLICT row - see exception list")
    return "\n".join(lines)


def _low_reason(hit) -> str:
    raw = (hit["raw_text"] or "").strip().replace("\n", " ")[:140]
    return (f"low confidence {hit['source_kind']} evidence: "
            f"\"{raw}\"")


def _congestion_zones(page_hits, page_rect) -> list:
    """Square-inch regions holding more hits than the threshold.

    A single fixed grid undercounts: a cluster straddling a cell
    corner splits four ways and a 3x-threshold knot can pass unseen.
    Four grids offset by half a cell catch straddlers; overlapping
    flagged cells merge into one zone carrying the highest count."""
    centers = []
    for hit in page_hits:
        try:
            bbox = json.loads(hit["bbox"])
            centers.append(((float(bbox[0]) + float(bbox[2])) / 2.0,
                            (float(bbox[1]) + float(bbox[3])) / 2.0))
        except (ValueError, TypeError, IndexError):
            continue
    flagged = []
    half = _CELL_PT / 2.0
    for ox, oy in ((0.0, 0.0), (half, 0.0), (0.0, half), (half, half)):
        cells = {}
        for cx, cy in centers:
            cell = (int((cx + ox) // _CELL_PT),
                    int((cy + oy) // _CELL_PT))
            cells[cell] = cells.get(cell, 0) + 1
        for (ix, iy), n in cells.items():
            if n > CONGESTION_THRESHOLD_PER_SQIN:
                x0 = ix * _CELL_PT - ox
                y0 = iy * _CELL_PT - oy
                flagged.append([x0, y0, x0 + _CELL_PT, y0 + _CELL_PT,
                                n])
    # Merge intersecting flagged cells into union zones.
    merged = []
    for zone in sorted(flagged):
        for m in merged:
            if not (zone[2] < m[0] or m[2] < zone[0]
                    or zone[3] < m[1] or m[3] < zone[1]):
                m[0] = min(m[0], zone[0])
                m[1] = min(m[1], zone[1])
                m[2] = max(m[2], zone[2])
                m[3] = max(m[3], zone[3])
                m[4] = max(m[4], zone[4])
                break
        else:
            merged.append(list(zone))
    return [((max(z[0], 0.0), max(z[1], 0.0),
              min(z[2], page_rect[2]), min(z[3], page_rect[3])), z[4])
            for z in merged]


def _manual_low_rows(takeoff_xlsx) -> list:
    """Manual-entry low rows from a takeoff xlsx (schema section 6:
    they have no drawable bbox, so they ride the summary page)."""
    import openpyxl

    out = []
    wb = openpyxl.load_workbook(str(takeoff_xlsx))
    try:
        if "TAKEOFF" not in wb.sheetnames:
            return out
        ws = wb["TAKEOFF"]
        for r in range(3, ws.max_row + 1):
            bbox = str(ws.cell(row=r, column=10).value or "")
            conf = str(ws.cell(row=r, column=8).value or "").lower()
            if bbox == "MANUAL" and conf == "low":
                out.append({
                    "sheet": str(ws.cell(row=r, column=9).value or
                                 "MANUAL"),
                    "designation": str(ws.cell(row=r, column=2).value
                                       or ""),
                    "reason": "manual-entry row, low confidence, no "
                              "drawable bbox (takeoff row "
                              f"{ws.cell(row=r, column=1).value})",
                })
    finally:
        wb.close()
    return out


def _draw_summary(doc, exceptions) -> int:
    """Exception list pages appended to the PDF: sheet, designation,
    reason, ordered for top-to-bottom working (conflicts first, then
    low confidence, each in sheet order). Returns pages added."""
    import fitz

    header = [
        _SUMMARY_TITLE,
        "Never send to a client. Ivan works this list top to bottom "
        "(P23/P24).",
        f"Generated {datetime.now(timezone.utc).isoformat()}",
        "",
        f"{'SHEET':<10} {'DESIGNATION':<28} REASON",
        "-" * 96,
    ]
    import textwrap

    lines = []
    for kind, sheet, desig, reason in exceptions:
        tag = "CONFLICT" if kind == "conflict" else "LOW"
        # Wrap, never truncate: a sliced conflict line would drop the
        # second quantity, the exact comparison Ivan adjudicates.
        first = f"{sheet:<10} {desig[:27]:<28} [{tag}] "
        wrapped = textwrap.wrap(reason, width=118 - len(first)) or [""]
        lines.append(first + wrapped[0])
        lines.extend(" " * len(first) + cont for cont in wrapped[1:])
    if not lines:
        lines = ["No conflicts and no low-confidence items. "
                 "Nothing to work."]

    pages = 0
    for start in range(0, len(lines), _LINES_PER_SUMMARY_PAGE):
        page = doc.new_page(width=792, height=612)  # landscape letter
        pages += 1
        chunk = lines[start:start + _LINES_PER_SUMMARY_PAGE]
        text = "\n".join(header + chunk) if start == 0 else \
            "\n".join([_SUMMARY_TITLE + " (continued)", ""] + chunk)
        page.insert_text(
            (36, 40), text, fontname="cour", fontsize=9,
            color=(0, 0, 0))
        page.insert_text(
            (36, 596), _INTERNAL_MARK, fontname="cour", fontsize=8,
            color=PALETTE["CONFLICT"])
    return pages


def build_overlay(pdf_path, job, db_path=None, takeoff_xlsx=None,
                  out_path=None) -> dict:
    """Annotate a copy of the drawing set from census.db. Returns a
    summary dict including the output path and what was drawn."""
    import fitz

    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"drawing set not found: {pdf_path}")
    if pdf_path.stem.endswith("_QA"):
        raise ValueError(
            "input is already a QA overlay; run against the original "
            "drawing set or annotations double and stale exception "
            "pages survive mid-document")
    out_path = Path(out_path) if out_path else \
        pdf_path.with_name(f"{pdf_path.stem}_QA.pdf")
    backed_up = _backup_existing(out_path)

    hits, conflicts = _load_hits(job, db_path)
    if not hits:
        raise ValueError(f"census.db holds no hits for job '{job}'; "
                         "run the census first")
    page_map, ambiguous_sheets = _page_map(str(pdf_path))

    totals = {}
    for h in hits:
        key = (_norm_designation(h["designation"]), h["source_kind"])
        totals[key] = totals.get(key, 0) + 1

    doc = fitz.open(str(pdf_path))
    try:
        drawn = 0
        unmapped = []
        ambiguous_hits = []
        by_page = {}
        exceptions = []

        for hit in hits:
            if hit["sheet"] in ambiguous_sheets:
                ambiguous_hits.append(hit)
                continue
            page_idx = page_map.get(hit["sheet"])
            if page_idx is None or page_idx >= len(doc):
                unmapped.append(hit)
                continue
            by_page.setdefault(page_idx, []).append(hit)

        for page_idx in sorted(by_page):
            page = doc[page_idx]
            for hit in by_page[page_idx]:
                try:
                    bbox = [float(v) for v in json.loads(hit["bbox"])]
                    rect = fitz.Rect(bbox).normalize()
                except (ValueError, TypeError, IndexError):
                    unmapped.append(hit)
                    continue
                if rect.is_empty:
                    rect = fitz.Rect(rect.x0 - 2, rect.y0 - 2,
                                     rect.x0 + 2, rect.y0 + 2)
                annot = page.add_rect_annot(rect)
                annot.set_colors(stroke=_hit_color(hit))
                annot.set_border(width=1.2)
                annot.set_info(
                    title=f"{hit['item_class']} census",
                    content=_note_text(hit, totals))
                annot.update()
                drawn += 1

            zones = _congestion_zones(by_page[page_idx],
                                      tuple(page.rect))
            for zone_rect, n in zones:
                annot = page.add_rect_annot(fitz.Rect(zone_rect))
                annot.set_colors(stroke=PALETTE["CONGESTION"])
                annot.set_border(width=2.0, dashes=[4, 3])
                annot.set_info(
                    title="congestion zone",
                    content=(f"congestion: {n} hits in one square "
                             f"inch (threshold "
                             f"{CONGESTION_THRESHOLD_PER_SQIN}) per "
                             "P24. Dense tags miscount; verify this "
                             "area by eye."))
                annot.update()
            by_page[page_idx] = (by_page[page_idx], len(zones))

            mark = page.add_freetext_annot(
                fitz.Rect(8, 8, 240, 24), _INTERNAL_MARK,
                fontsize=8, text_color=PALETTE["CONFLICT"])
            mark.update()

        # Exception list: conflicts first, then low confidence, each
        # block in natural sheet order so Ivan works top to bottom.
        seen_conflicts = set()
        conflict_entries = []
        for hit in hits:
            ck = hit["conflict_group"]
            if ck and ck not in seen_conflicts:
                seen_conflicts.add(ck)
                note = conflicts.get(ck, {}).get(
                    "note", "conflicting counts; see census.db")
                conflict_entries.append(("conflict", hit["sheet"],
                                         hit["designation"], note))
        conflict_entries.sort(key=lambda e: (sheet_key(e[1]), e[2]))
        exceptions.extend(conflict_entries)
        low = [h for h in hits
               if h["confidence"] == "low" and not h["conflict_group"]]
        for hit in sorted(low, key=lambda h: (
                sheet_key(h["sheet"]), h["designation"])):
            exceptions.append(("low", hit["sheet"], hit["designation"],
                               _low_reason(hit)))
        for row in _manual_low_rows(takeoff_xlsx) if takeoff_xlsx \
                else []:
            exceptions.append(("low", row["sheet"], row["designation"],
                               row["reason"]))
        for hit in unmapped:
            exceptions.append((
                "low", hit["sheet"], hit["designation"],
                "hit could not be drawn (sheet not in this PDF or "
                "bbox unparseable); verify from census.db"))
        for hit in ambiguous_hits:
            exceptions.append((
                "low", hit["sheet"], hit["designation"],
                f"sheet number {hit['sheet']} appears on more than one "
                "page of this PDF; hit drawn nowhere, verify from "
                "census.db"))

        exceptions.sort(key=lambda e: 0 if e[0] == "conflict" else 1)
        summary_pages = _draw_summary(doc, exceptions)

        doc.save(str(out_path), garbage=3, deflate=True)
    finally:
        doc.close()

    return {
        "out_path": str(out_path),
        "job": job,
        "hits_total": len(hits),
        "hits_drawn": drawn,
        "hits_unmapped": len(unmapped),
        "hits_ambiguous_sheet": len(ambiguous_hits),
        "ambiguous_sheets": sorted(ambiguous_sheets),
        "pages_annotated": len(by_page),
        "congestion_zones": sum(z for _, z in by_page.values()),
        "conflicts_listed": len(seen_conflicts),
        "low_listed": len([e for e in exceptions if e[0] == "low"]),
        "summary_pages": summary_pages,
        "previous_backed_up": backed_up,
    }


def sheet_key(sheet: str):
    """Natural sheet order: S2.1 before S10.1."""
    return tuple(int(p) if p.isdigit() else p
                 for p in re.split(r"(\d+)", sheet or "") if p)


def _backup_existing(out_path: Path) -> str:
    """The _QA file is Ivan's working surface and may carry his own
    markup; a regenerate must never destroy it silently. Snapshot to
    _handoff/backups/<UTC-ISO>/ and append a changelog line, per the
    repo operating rule. Returns the backup path or ''."""
    import shutil

    if not Path(out_path).exists():
        return ""
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    backup_dir = HANDOFF_ROOT / "backups" / ts
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / Path(out_path).name
    shutil.copy2(str(out_path), str(dest))
    changelog = HANDOFF_ROOT / "changelog.md"
    with open(changelog, "a", encoding="utf-8") as f:
        f.write(f"\n{ts} - QA overlay regenerate: existing "
                f"{Path(out_path).name} backed up to "
                f"_handoff/backups/{ts}/ before overwrite (it may "
                "carry Ivan's markup).")
    return str(dest)


def main() -> int:
    usage = ("usage: py -m takeoff_pipeline.overlay <pdf> <job> "
             "[--db <census.db>] [--xlsx <takeoff.xlsx>] "
             "[--out <path>]")
    args = sys.argv[1:]
    if len(args) < 2 or args[0].startswith("--"):
        print(usage)
        return 2

    def flag(name):
        if name not in args:
            return None
        i = args.index(name)
        if i + 1 >= len(args):
            raise ValueError(f"{name} needs a value")
        return args[i + 1]

    try:
        info = build_overlay(args[0], args[1], db_path=flag("--db"),
                             takeoff_xlsx=flag("--xlsx"),
                             out_path=flag("--out"))
    except (ValueError, FileNotFoundError) as e:
        print(e)
        print(usage)
        return 1
    for k, v in info.items():
        print(f"{k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
