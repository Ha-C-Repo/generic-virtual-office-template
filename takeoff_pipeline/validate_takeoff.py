"""Validation gate for the four-sheet takeoff workbook.

Implements the section 14 table of TAKEOFF_SCHEMA_V2.md. Any HARD FAIL
stops the pipeline with a non-zero exit and no stamp. CONFLICT rows
and low-confidence rows are findings, not errors: the file PASSES with
a warning block listing each conflict as an RFI candidate and routing
low rows to human review. A validator that resolves, averages, or
drops a conflict is broken (section 8).

The pricing wall (section 12): header cells fail on a case-insensitive
substring of any token ($, rate, price, cost, margin, markup, amount);
data cells fail on any dollar character or a case-insensitive
whole-word match of rate, price, cost, margin, or markup. Words that
merely contain a token pass (accurate, separate, operated, strategy).

The hash re-verification uses the ONE shared digest implementation in
takeoff_pipeline/takeoff_hash.py (13.2). The lb_per_ft cross-check
against bridge/aisc_validator.py runs when that module is importable;
when it is not, the skip is REPORTED as a warning, never silent.

Interpretations this implementation makes of the section 14 table,
stated rather than hidden:
- The section 9 assembly cross-check WARN compares any numeric BOQ/BOM
  row qty whose source_item_id targets a measured ANCH or PLATE row
  against that row's qty. Prompt 9 may refine the emission format.
- Stamp detection fails closed: a populated row 1 or a stamped-pattern
  filename without a valid A1 literal and 64-hex D1 is a hard fail,
  because the alternative lets one cleared cell disable the 13.4 gate.
- Derived weight cells must be CELL-REFERENCING formulas; a constant
  wearing an equals sign (=3.12) fails.
- Content in any column beyond a sheet's spec layout is a hard fail on
  every sheet (the digest covers A-O only; the validator closes the
  rest).

Exit codes: 0 pass (with or without warnings), 1 hard fail, 2 usage.
"""

import re
import sys
from pathlib import Path

from takeoff_pipeline.takeoff_hash import compute_takeoff_hash

SHEET_NAMES = ("TAKEOFF", "BOQ", "BOM", "PRICING_SCHEDULE")

TAKEOFF_HEADERS = (
    "item_id", "designation", "mode", "qty", "unit", "primary_source",
    "secondary_source", "confidence", "sheet", "bbox", "notes",
    "lb_per_ft", "weight_lb", "tons", "formula_ref",
)
BOQ_HEADERS = ("boq_id", "source_item_id", "assembly_id", "stream",
               "description", "qty", "unit", "formula_ref", "notes")
BOM_HEADERS = ("bom_id", "source_item_id", "designation", "description",
               "qty", "unit", "length_lf", "lb_per_ft", "weight_lb",
               "tons", "formula_ref", "notes")
PS_HEADERS = ("line_id", "description", "source_refs", "qty", "unit")

MODES = ("COUNT", "LINEAR", "AREA")
MODE_UNIT = {"COUNT": "EA", "LINEAR": "LF", "AREA": "SF"}
CONFIDENCE = ("high", "medium", "low")
CLASSES = ("COL", "BEAM", "JST", "DECK", "PLATE", "ANCH", "MISC")
_ITEM_ID = re.compile(r"^(COL|BEAM|JST|DECK|PLATE|ANCH|MISC)-\d{3,}$")
_BBOX = re.compile(
    r"^\[\s*-?\d+(?:\.\d+)?\s*(?:,\s*-?\d+(?:\.\d+)?\s*){3}\]$")

# Section 5 conformance: primary_source must contain the keyword.
SOURCE_KEYWORDS = {
    "COL": ("SCHEDULE",),
    "JST": ("SCHEDULE",),
    "ANCH": ("SCHEDULE", "DETAIL"),
    "PLATE": ("SCHEDULE", "DETAIL"),
    "BEAM": ("FRAMING PLAN",),
    "DECK": ("FRAMING PLAN",),
    "MISC": (),
}

PRICING_TOKENS = ("$", "rate", "price", "cost", "margin", "markup",
                  "amount")
_DATA_WORD = re.compile(r"(?i)(?<![a-z])(rate|price|cost|margin|markup)"
                        r"(?![a-z])")
SUPPLIER_TOKENS = ("VULCRAFT", "CANAM", "NUCOR", "AYAMSA")

BOQ_STREAM_UNITS = {
    "MATERIAL": ("EA", "LF", "SF", "LB", "TON"),
    "SUBCONTRACT": ("EA", "LF", "SF", "LB", "TON"),
    "LABOR": ("HR",),
    "EQUIPMENT": ("HR", "MIN"),
}
BOM_UNITS = ("EA", "LF", "SF", "LB", "TON")
PS_UNITS = ("EA", "LF", "SF", "TON")  # LB intentionally excluded


_CELL_REF = re.compile(r"(?i)(?<![A-Z0-9_])\$?[A-Z]{1,3}\$?\d+")


def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _is_derived_formula(cell) -> bool:
    """A derived weight cell must be a formula that actually REFERENCES
    a cell. '=3.12' is a typed constant wearing an equals sign; the
    section 4 rule exists to stop exactly that fabricated-number path
    (Hard Rule 5)."""
    return _is_formula(cell) and bool(_CELL_REF.search(cell.value))


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def validate_file(path) -> dict:
    """Run the full section 14 table. Returns {hard_fails, warnings,
    conflicts, low_rows, summary}."""
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(str(path))  # not data_only: formulas
    fails = []
    warnings = []
    conflicts = []
    low_rows = []

    for name in wb.sheetnames:
        if name not in SHEET_NAMES:
            fails.append(f"worksheet '{name}' is not one of the four "
                         "named sheets")
    for name in SHEET_NAMES:
        if name not in wb.sheetnames:
            fails.append(f"required worksheet '{name}' is missing")
    if fails:
        return _report(path, fails, warnings, conflicts, low_rows)

    ws = wb["TAKEOFF"]
    headers = tuple(_text(ws.cell(row=2, column=c).value)
                    for c in range(1, len(TAKEOFF_HEADERS) + 1))
    if headers != TAKEOFF_HEADERS:
        fails.append("TAKEOFF header row is not exactly the fifteen "
                     f"columns of 3.2 plus 4.2 in order; found {headers}")

    item_ids = {}
    item_rows = {}
    aisc_checks = []
    for r in range(3, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c)
                 for c in range(1, len(TAKEOFF_HEADERS) + 1)]
        if all(c.value is None for c in cells):
            continue
        (item_id, designation, mode, qty, unit, primary, _secondary,
         confidence, _sheet, bbox, notes, lb_per_ft, weight_lb, tons,
         formula_ref) = cells

        iid = _text(item_id.value)
        if not _ITEM_ID.match(iid):
            fails.append(f"TAKEOFF row {r}: item_id '{iid}' does not "
                         "match <CLASS>-<NNN>")
        elif iid in item_ids:
            fails.append(f"TAKEOFF row {r}: item_id '{iid}' duplicates "
                         f"row {item_ids[iid]}")
        else:
            item_ids[iid] = r
            item_rows[iid] = {"mode": _text(mode.value),
                              "class": iid.split("-")[0],
                              "qty": qty.value}

        mode_v = _text(mode.value)
        if not mode_v:
            fails.append(f"TAKEOFF row {r}: missing mode")
        elif mode_v not in MODES:
            fails.append(f"TAKEOFF row {r}: mode '{mode_v}' is not "
                         "COUNT, LINEAR, or AREA")
        unit_v = _text(unit.value)
        if mode_v in MODES and unit_v != MODE_UNIT[mode_v]:
            fails.append(f"TAKEOFF row {r}: unit '{unit_v}' does not "
                         f"match mode {mode_v} "
                         f"(expected {MODE_UNIT[mode_v]})")

        primary_v = _text(primary.value)
        if not primary_v:
            fails.append(f"TAKEOFF row {r}: missing primary_source")
        elif _ITEM_ID.match(iid):
            cls = iid.split("-")[0]
            keywords = SOURCE_KEYWORDS[cls]
            if keywords and not any(k in primary_v.upper()
                                    for k in keywords):
                fails.append(
                    f"TAKEOFF row {r}: primary_source '{primary_v}' "
                    f"fails the section 5 keyword rule for {cls} "
                    f"(needs one of: {', '.join(keywords)})")

        conf_v = _text(confidence.value)
        if conf_v not in CONFIDENCE:
            fails.append(f"TAKEOFF row {r}: confidence '{conf_v}' is "
                         "not high, medium, or low")

        bbox_v = _text(bbox.value)
        if bbox_v != "MANUAL" and not _BBOX.match(bbox_v):
            fails.append(f"TAKEOFF row {r}: bbox '{bbox_v}' is neither "
                         "[x0, y0, x1, y1] nor MANUAL")

        notes_v = _text(notes.value)
        if "CONFLICT:" in notes_v:
            if not notes_v.startswith("CONFLICT:"):
                warnings.append(f"TAKEOFF row {r}: CONFLICT: token is "
                                "not at the start of notes")
            if conf_v != "low":
                fails.append(f"TAKEOFF row {r}: CONFLICT row with "
                             f"confidence '{conf_v}', must be low")
            conflicts.append((r, iid, _text(designation.value), notes_v))
        if conf_v == "low":
            low_rows.append((r, iid, _text(designation.value)))

        if lb_per_ft.value is not None and mode_v != "LINEAR":
            fails.append(f"TAKEOFF row {r}: lb_per_ft populated on a "
                         f"{mode_v or 'blank-mode'} row (LINEAR only)")
        for cell, label in ((weight_lb, "weight_lb"), (tons, "tons")):
            if cell.value is None:
                continue
            if not _is_derived_formula(cell):
                fails.append(f"TAKEOFF row {r}: {label} holds a typed "
                             "constant instead of a cell-referencing "
                             "formula")
            if not _text(formula_ref.value):
                fails.append(f"TAKEOFF row {r}: derived value without "
                             "formula_ref")
        if lb_per_ft.value is not None and not isinstance(
                lb_per_ft.value, (int, float)):
            fails.append(f"TAKEOFF row {r}: lb_per_ft must be a typed "
                         "number the operator filled, not "
                         f"'{lb_per_ft.value}'")
        if isinstance(lb_per_ft.value, (int, float)):
            aisc_checks.append((f"TAKEOFF row {r}",
                                _text(designation.value),
                                float(lb_per_ft.value)))

        for tok in ("GROUP:", "OPENING:"):
            m = re.search(rf"{tok}\s*([A-Z]+-\d+)", notes_v)
            if tok in notes_v and not m:
                fails.append(f"TAKEOFF row {r}: {tok} token with no "
                             "parseable item_id")
            elif m:
                ref = m.group(1)
                if ref not in item_ids and ref not in _all_item_ids(ws):
                    fails.append(f"TAKEOFF row {r}: {tok} cites "
                                 f"'{ref}' which is absent from the "
                                 "sheet")
                else:
                    target = _row_info(ws, ref)
                    if tok == "OPENING:" and target \
                            and target["class"] != "DECK":
                        fails.append(f"TAKEOFF row {r}: OPENING: parent "
                                     f"'{ref}' is not a DECK row")
                    if tok == "GROUP:" and target \
                            and target["mode"] != "COUNT":
                        fails.append(f"TAKEOFF row {r}: GROUP: anchor "
                                     f"'{ref}' is not a COUNT row")

    _check_headers(wb, "BOQ", BOQ_HEADERS, fails)
    _check_headers(wb, "BOM", BOM_HEADERS, fails)
    _check_headers(wb, "PRICING_SCHEDULE", PS_HEADERS, fails)
    _check_no_extra_columns(wb, fails)
    _check_boq_bom(wb, item_rows, item_ids, fails, warnings,
                   aisc_checks)
    _check_pricing_schedule(wb, fails, warnings)
    _check_pricing_tokens(wb, fails)
    _check_stamp(ws, path, fails)
    _check_aisc(aisc_checks, fails, warnings)

    return _report(path, fails, warnings, conflicts, low_rows)


def _all_item_ids(ws) -> set:
    return {_text(ws.cell(row=r, column=1).value)
            for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None}


def _row_info(ws, item_id):
    for r in range(3, ws.max_row + 1):
        if _text(ws.cell(row=r, column=1).value) == item_id:
            return {"mode": _text(ws.cell(row=r, column=3).value),
                    "class": item_id.split("-")[0]}
    return None


def _check_headers(wb, name, expected, fails) -> None:
    ws = wb[name]
    found = tuple(_text(ws.cell(row=1, column=c).value)
                  for c in range(1, len(expected) + 1))
    if found != expected:
        fails.append(f"{name} header row is not exactly the columns of "
                     f"the spec; found {found}")


_SHEET_WIDTH = {"TAKEOFF": len(TAKEOFF_HEADERS),
                "BOQ": len(BOQ_HEADERS), "BOM": len(BOM_HEADERS),
                "PRICING_SCHEDULE": len(PS_HEADERS)}


def _check_no_extra_columns(wb, fails) -> None:
    """The sheet layouts are closed-world (3.1: fifteen columns, A
    through O). The digest covers A-O only, so content parked beyond
    the last spec column would ride a stamped file undetected; any
    non-empty cell out there is a HARD FAIL on every sheet."""
    for name, width in _SHEET_WIDTH.items():
        ws = wb[name]
        for r in range(1, ws.max_row + 1):
            for c in range(width + 1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    fails.append(
                        f"{name} row {r} column {c}: content beyond "
                        f"the sheet's {width}-column layout (outside "
                        "the version hash on TAKEOFF)")
                    break


def _check_boq_bom(wb, item_rows, item_ids, fails, warnings,
                   aisc_checks) -> None:
    for name in ("BOQ", "BOM"):
        ws = wb[name]
        for r in range(2, ws.max_row + 1):
            if all(ws.cell(row=r, column=c).value is None
                   for c in range(1, ws.max_column + 1)):
                continue
            src = _text(ws.cell(row=r, column=2).value)
            if not src:
                fails.append(f"{name} row {r}: missing source_item_id")
            elif src not in item_ids:
                fails.append(f"{name} row {r}: source_item_id '{src}' "
                             "is absent from the TAKEOFF sheet")
            else:
                # Section 9 measured-class rule, section 14 WARN row:
                # an assembly-derived count for a measured ANCH or
                # PLATE row that disagrees with the measured qty is a
                # CONFLICT-style warning for Ivan, never silent.
                target = item_rows.get(src, {})
                qty_col = 6 if name == "BOQ" else 5
                row_qty = ws.cell(row=r, column=qty_col).value
                if (target.get("class") in ("ANCH", "PLATE")
                        and isinstance(row_qty, (int, float))
                        and isinstance(target.get("qty"), (int, float))
                        and float(row_qty) != float(target["qty"])):
                    warnings.append(
                        f"{name} row {r}: assembly-derived count "
                        f"{row_qty:g} disagrees with measured "
                        f"{target['class']} row {src} qty "
                        f"{target['qty']:g} (section 9 cross-check); "
                        "listed for Ivan")
            if name == "BOQ":
                stream = _text(ws.cell(row=r, column=4).value).upper()
                unit = _text(ws.cell(row=r, column=7).value).upper()
                if stream not in BOQ_STREAM_UNITS:
                    fails.append(f"BOQ row {r}: stream '{stream}' is "
                                 "not MATERIAL, LABOR, EQUIPMENT, or "
                                 "SUBCONTRACT")
                elif unit not in BOQ_STREAM_UNITS[stream]:
                    fails.append(f"BOQ row {r}: unit '{unit}' is not "
                                 f"valid for stream {stream}")
            if name == "BOM":
                unit = _text(ws.cell(row=r, column=6).value).upper()
                if unit and unit not in BOM_UNITS:
                    fails.append(f"BOM row {r}: unit '{unit}' outside "
                                 "the BOM enum")
                derived = False
                for col, label in ((9, "weight_lb"), (10, "tons")):
                    cell = ws.cell(row=r, column=col)
                    if cell.value is None:
                        continue
                    derived = True
                    if not _is_derived_formula(cell):
                        fails.append(f"BOM row {r}: {label} holds a "
                                     "typed constant instead of a "
                                     "cell-referencing formula")
                if derived and not _text(
                        ws.cell(row=r, column=11).value):
                    fails.append(f"BOM row {r}: derived value without "
                                 "formula_ref")
                lbf = ws.cell(row=r, column=8).value
                if isinstance(lbf, (int, float)):
                    aisc_checks.append(
                        (f"BOM row {r}",
                         _text(ws.cell(row=r, column=3).value),
                         float(lbf)))


def _check_pricing_schedule(wb, fails, warnings) -> None:
    ws = wb["PRICING_SCHEDULE"]
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value is None
               for c in range(1, 6)):
            continue
        unit = _text(ws.cell(row=r, column=5).value).upper()
        if unit and unit not in PS_UNITS:
            fails.append(f"PRICING_SCHEDULE row {r}: unit '{unit}' "
                         "outside the enum (LB is intentionally "
                         "excluded)")
        for col in (2, 3):
            text = _text(ws.cell(row=r, column=col).value).upper()
            for tok in SUPPLIER_TOKENS:
                if tok in text:
                    warnings.append(
                        f"PRICING_SCHEDULE row {r}: supplier-name "
                        f"token '{tok.title()}' (Hard Rule 4); flag "
                        "before stamping")


def _check_pricing_tokens(wb, fails) -> None:
    """Section 12. Header cells: substring on any token. Data cells:
    the dollar character, or a whole-word rate/price/cost/margin/
    markup. The TAKEOFF metadata row (row 1) is scanned with the
    data-cell rules: the rule says every sheet, and the legitimate
    stamp values carry no tokens."""
    for name in SHEET_NAMES:
        ws = wb[name]
        header_row = 2 if name == "TAKEOFF" else 1
        for c in range(1, ws.max_column + 1):
            text = _text(ws.cell(row=header_row, column=c).value).lower()
            for tok in PRICING_TOKENS:
                if tok in text:
                    fails.append(
                        f"{name} header column {c}: pricing token "
                        f"'{tok}' in header '{text}' (section 12 "
                        "HARD FAIL)")
        data_rows = list(range(header_row + 1, ws.max_row + 1))
        if name == "TAKEOFF":
            data_rows = [1] + data_rows
        for r in data_rows:
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                if "$" in v:
                    # Spec-literal: ANY dollar character, including
                    # inside formula strings (absolute references are
                    # not written by this pipeline).
                    fails.append(f"{name} row {r} column {c}: dollar "
                                 "character in data cell (section 12)")
                m = _DATA_WORD.search(v)
                if m:
                    fails.append(f"{name} row {r} column {c}: "
                                 f"whole-word pricing token "
                                 f"'{m.group(1)}' in data cell "
                                 "(section 12)")


_STAMPED_NAME = re.compile(r".+_TAKEOFF_R\d+_[0-9a-f]{12}\.xlsx$")
_DIGEST = re.compile(r"^[0-9a-f]{64}$")


def _check_stamp(ws, path, fails) -> None:
    """Stamped file: recomputed digest must match D1 (13.4).

    Stamp detection fails CLOSED: the only legitimately unstamped
    artifact in the 13.1 flow has an EMPTY row 1 and an UNSTAMPED
    filename. A file with anything in row 1, or carrying the 13.3
    stamped filename pattern, is a stamped file; if its A1 literal or
    D1 digest is missing or malformed, the stamp was damaged or
    removed, which is itself a hard fail. Otherwise clearing one
    unhashed cell (A1) would disable tamper detection entirely."""
    a1 = _text(ws.cell(row=1, column=1).value)
    row1_populated = any(
        ws.cell(row=1, column=c).value is not None
        for c in range(1, max(ws.max_column, 15) + 1))
    name_stamped = bool(_STAMPED_NAME.match(Path(path).name))
    if a1 != "TAKEOFF_SCHEMA_V2":
        if row1_populated or name_stamped:
            fails.append(
                "file looks stamped (row 1 populated or stamped "
                "filename) but the A1 schema literal is missing; "
                "stamp damaged or removed (13.4)")
        return
    stamped = _text(ws.cell(row=1, column=4).value)
    if not _DIGEST.match(stamped):
        fails.append("stamp present but D1 is not a 64-character "
                     "lowercase hex digest (13.3)")
        return
    recomputed = compute_takeoff_hash(ws)
    if stamped != recomputed:
        fails.append("stamped digest in D1 does not match the "
                     "recomputed TAKEOFF hash; the file was edited "
                     "after stamping (13.4 forbids in-place edits)")


def _check_aisc(aisc_checks, fails, warnings) -> None:
    """Typed lb_per_ft vs bridge/aisc_validator.py (section 14). The
    check needs bridge/ importable; a skip is reported, never silent."""
    if not aisc_checks:
        return
    try:
        from bridge.aisc_validator import validate_shape
    except Exception:
        warnings.append(
            f"{len(aisc_checks)} typed lb_per_ft values could not be "
            "cross-checked: bridge.aisc_validator is not importable "
            "in this environment. Run the validator from the repo "
            "root before stamping.")
        return
    for where, designation, typed in aisc_checks:
        result = validate_shape(designation)
        ref = result.get("weight_per_ft")
        if not result.get("valid") or ref is None:
            continue  # non-AISC designation: no reference weight
        if abs(typed - float(ref)) > 0.051:
            fails.append(
                f"{where}: lb_per_ft {typed:g} for {designation} "
                f"differs from bridge/aisc_validator.py ({ref:g})")


def _report(path, fails, warnings, conflicts, low_rows) -> dict:
    summary = (f"{len(fails)} hard fails, {len(warnings)} warnings, "
               f"{len(conflicts)} conflict rows, "
               f"{len(low_rows)} low-confidence rows")
    return {"path": str(path), "hard_fails": fails, "warnings": warnings,
            "conflicts": conflicts, "low_rows": low_rows,
            "summary": summary}


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: py -m takeoff_pipeline.validate_takeoff "
              "<takeoff.xlsx>")
        return 2
    report = validate_file(sys.argv[1])
    print(f"VALIDATION: {report['path']}")
    for f in report["hard_fails"]:
        print(f"  HARD FAIL: {f}")
    for w in report["warnings"]:
        print(f"  WARN: {w}")
    if report["conflicts"]:
        print("  CONFLICT ROWS, RFI candidates (P26, never resolved "
              "silently):")
        for r, iid, desig, note in report["conflicts"]:
            print(f"    row {r} {iid} {desig}: {note}")
    if report["low_rows"]:
        print("  LOW-CONFIDENCE ROWS, routed to human review (P23/P24):")
        for r, iid, desig in report["low_rows"]:
            print(f"    row {r} {iid} {desig}")
    print(f"  {report['summary']}")
    if report["hard_fails"]:
        print("  RESULT: HARD FAIL, no stamp")
        return 1
    print("  RESULT: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
