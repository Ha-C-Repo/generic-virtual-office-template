"""PC1 award-to-budget conversion plus the PC2 WBS template sheet.

Implements P16 (estimate is not budget): contract value minus target
margin is the cost baseline; the baseline splits into cost codes that
mirror the bid's -GP report lines, with management reserve carried as
an explicit named line, never buried. Implements P14 (baselines or
nothing): the workbook carries a status cell that reads DRAFT until
the freeze step stamps it BASELINE with a frozen date and approver; a
frozen file is never overwritten, edits require a new version.
Implements P15 (integrated WBS): a second sheet where every line
carries scope, cost code, schedule activity, quality check, and risk
note together.

The output format is bound to the PC4 reader,
bridge/project_controls.py :: _load_baseline. Where this module and
that reader disagree, the reader wins:

  - the file lands in "09 Financials -GP CONFIDENTIAL" with "budget"
    in the filename;
  - the frozen flag cell reads exactly "BASELINE frozen <date>", with
    the date repeated in a sibling cell on the same row;
  - the WBS sheet title contains "wbs" and its headers match the
    reader's _WBS_ALIASES;
  - versions are zero-padded (v01) and superseded versions move to a
    superseded/ subfolder at FREEZE time, not at convert time, because
    the reader picks the first alphabetical match in the 09 folder:
    the approved baseline keeps governing while its replacement draft
    is under review (P14), and the padding keeps v10 sorting after
    v02.

BAC partition rule: the PC4 reader sums WBS budget_cost for BAC, so
one WBS line owns each budget dollar. Codes split across several WBS
lines ship with the budget cell blank plus a note; the operator
allocates before freeze and the freeze gate checks the sums. G&A and
management reserve are budget lines without WBS lines: overhead and
reserve are not scheduled work.

No rates live here (P10). Line costs come from the -GP breakdown the
operator feeds in; margin and reserve are explicit inputs, never
defaulted. G&A is read from the source breakdown when present, never
recomputed from a constant. Rates and GP targets stay in
bridge/bid_rates.py.

Paths are package-relative; resource_path() only if this is ever
promoted into bridge/ (see takeoff_pipeline/__init__.py).

CONFIDENTIAL - INTERNAL. The budget xlsx is -GP material: internal
cost basis, margin, and reserve. Never client-facing.

usage: py -m takeoff_pipeline.budget_convert convert
           --source <gp_lines.csv|.xlsx>
           --contract-value <usd> --margin <pct or decimal>
           (--reserve-pct <pct> | --reserve-amount <usd>)
           (--project-folder "<Awarded Projects/...>" | --out <dir>)
           [--project-id <id>] [--force]
       py -m takeoff_pipeline.budget_convert freeze <budget.xlsx>
           --approved-by "<name>" [--date YYYY-MM-DD]
           [--allow-incomplete]
       py -m takeoff_pipeline.budget_convert verify <budget.xlsx>
"""

import csv
import re
import shutil
import sys
from datetime import date, datetime, timezone
from pathlib import Path

_PKG = Path(__file__).resolve().parent
_REPO_ROOT = _PKG.parent

# Cost codes mirror the -GP report lines exactly (PC1 commission).
# Stairs, misc metals, and shop drawings are scope rows on today's -GP
# reports, not cost lines; they stay in the code set and read zero
# when the source breakdown has no such line.
COST_CODES = (
    ("FAB", "Fabrication",
     ("fab", "fabrication", "structural steel fabrication")),
    ("ERE", "Erection",
     ("erection", "erect", "structural steel erection")),
    ("JST", "Joists",
     ("joists", "joist", "steel joists", "steel joists sji k series")),
    ("CDK", "Composite deck",
     ("composite deck", "composite floor deck", "comp deck")),
    ("RDK", "Roof deck", ("roof deck",)),
    ("ANC", "Anchor rods",
     ("anchor rods", "anchor rod", "anchors", "anchor bolts")),
    ("STR", "Stairs", ("stairs", "stair", "metal stairs")),
    ("MSC", "Misc metals",
     ("misc metals", "misc metal", "miscellaneous metals",
      "misc steel")),
    ("SHD", "Shop drawings",
     ("shop drawings", "shop drawings detailing pe stamps",
      "shop drawings and detailing", "detailing")),
    ("GA", "G&A overhead",
     ("g and a", "ga", "g and a overhead", "ga overhead", "overhead")),
)
MR_CODE = "MR"
MR_NAME = "Management reserve"
# Work codes carry WBS lines; GA is overhead, not scheduled work.
WORK_CODES = tuple(c for c, _n, _a in COST_CODES if c != "GA")

_DRAFT_MARK = "DRAFT - not a baseline (P14)"
_CLASSIFICATION = "CONFIDENTIAL - INTERNAL - never client-facing"

# Source breakdown header aliases (tolerant, like the PC4 reader).
_SRC_ALIASES = {
    "line": ("line item", "line", "item", "description", "cost line"),
    "cost": ("cost", "direct cost", "cost usd", "internal cost",
             "amount"),
    "revenue": ("revenue", "sell price", "price", "sell"),
    "gp_pct": ("gp percent", "gp pct", "gp", "margin percent",
               "margin pct"),
}

# Rows on a -GP table that are structure, not cost lines.
_SKIP_EXACT = ("total", "total bid", "base bid", "grand total")
_RESIDUAL_ROW = "deck joists anchors"

# WBS sheet headers; normalized forms match the PC4 reader's
# _WBS_ALIASES canonical names.
WBS_HEADERS = ("WBS Line", "Scope", "Cost Code", "Progress Type",
               "Planned Units", "Unit", "Planned Hours", "Budget Cost",
               "Start Date", "End Date", "Schedule Activity",
               "Quality Check", "Risk Note")

# Read-back aliases for the freeze and verify gates, mirroring the
# tolerance of the PC4 reader's _WBS_ALIASES so a header the reader
# would accept does not silently defeat the gate. Keys the gate needs
# but cannot map make it fail closed (raise), never fail open.
_GATE_ALIASES = {
    "wbs_line": ("wbs line", "wbs", "wbs id", "line id", "line"),
    "cost_code": ("cost code", "code"),
    "progress_type": ("progress type", "measure type", "measure",
                      "type"),
    "planned_units": ("planned units", "planned qty", "plan qty",
                      "quantity", "qty"),
    "planned_hours": ("planned hours", "budget hours", "plan hours",
                      "hours"),
    "budget_cost": ("budget cost", "budget", "bac", "cost",
                    "budget usd"),
    "start_date": ("start date", "start", "planned start",
                   "schedule start"),
    "end_date": ("end date", "end", "finish", "planned finish",
                 "schedule finish"),
}

_DATE_RX = re.compile(r"\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4}")
_PARENS_RX = re.compile(r"\([^)]*\)")

_MAX_SRC_ROWS = 500
_EMPTY_STREAK = 50
_SCAN_ROWS = 40
_SCAN_COLS = 15
_CENTS_TOL = 0.02
_SCALE_WARN = 0.15
_CONTRACT_WARN = 0.05


# -- value parsing --------------------------------------------------------

def _norm(text) -> str:
    """Normalize a label for alias matching: drop parentheticals,
    lower-case, & to and, / and _ and % expanded, spaces collapsed."""
    t = _PARENS_RX.sub(" ", str(text or "")).lower()
    t = (t.replace("&", " and ").replace("/", " ")
         .replace("_", " ").replace("%", " percent"))
    return " ".join(t.split())


def _money(value):
    """Parse a currency cell. Returns float or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    t = str(value).strip().replace("$", "").replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _pct_or_decimal(raw, label) -> float:
    """Margin-style input: 25 and 0.25 both mean 25 percent."""
    v = _money(raw)
    if v is None:
        raise ValueError(f"{label} is not a number: {raw!r}")
    if v >= 1.0:
        v = v / 100.0
    if not 0.0 <= v < 1.0:
        raise ValueError(f"{label} must land in 0..100 percent "
                         f"exclusive; got {raw!r}")
    return v


# -- source breakdown reader ----------------------------------------------

def _src_header_map(cells) -> dict:
    found = {}
    for idx, cell in enumerate(cells):
        h = _norm(cell)
        if not h:
            continue
        for canonical, names in _SRC_ALIASES.items():
            if canonical not in found and h in names:
                found[canonical] = idx
    return found


def _rows_from_csv(path) -> list:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


def _rows_from_xlsx(path) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True,
                                read_only=True)
    try:
        ws = wb.worksheets[0]
        rows = []
        streak = 0
        for row in ws.iter_rows(min_row=1, max_row=_MAX_SRC_ROWS):
            cells = [c.value for c in row]
            if all(c is None or str(c).strip() == "" for c in cells):
                streak += 1
                if streak >= _EMPTY_STREAK:
                    break
                rows.append(cells)
                continue
            streak = 0
            rows.append(cells)
        return rows
    finally:
        wb.close()


def read_source(path) -> dict:
    """Read a -GP style cost breakdown (csv or xlsx). Returns
    {"costs": {code: usd}, "revenue_total": usd or None,
    "skipped": [...], "warnings": [...]}. Raises ValueError on rows
    that cannot be mapped; never re-buckets a line silently."""
    p = Path(path)
    if not p.exists():
        raise ValueError(f"source breakdown not found: {p}")
    if p.suffix.lower() == ".csv":
        rows = _rows_from_csv(p)
    elif p.suffix.lower() == ".xlsx":
        rows = _rows_from_xlsx(p)
    else:
        raise ValueError("source breakdown must be .csv or .xlsx; got "
                         f"{p.name}")

    header_map = {}
    header_idx = None
    for i, cells in enumerate(rows[:10]):
        cand = _src_header_map(cells)
        if "line" in cand and ("cost" in cand or
                               ("revenue" in cand and "gp_pct" in cand)):
            header_map = cand
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            f"no readable header row in {p.name}; need a Line Item "
            "column plus Cost (or Revenue plus GP %), matching the "
            "-GP report table")

    alias_to_code = {}
    for code, _name, aliases in COST_CODES:
        for a in aliases:
            alias_to_code[a] = code

    # Indices that carry a header. Data outside them is almost always
    # an unquoted thousands separator splitting a hand-typed cell
    # ("Fabrication,$1,000" -> cost reads as $1); fail, never smear.
    known_idx = {i for i, c in enumerate(rows[header_idx])
                 if _norm(c)}

    costs = {}
    revenue_total = 0.0
    has_revenue = False
    skipped = []
    warnings = []
    errors = []
    for n, cells in enumerate(rows[header_idx + 1:],
                              start=header_idx + 2):
        def pick(name):
            idx = header_map.get(name)
            return cells[idx] if idx is not None and idx < len(cells) \
                else None

        raw_name = pick("line")
        nrm = _norm(raw_name)
        if not nrm:
            continue
        spill = [i + 1 for i, c in enumerate(cells)
                 if i not in known_idx and c is not None
                 and str(c).strip() != ""]
        if spill:
            errors.append(
                f"row {n} '{raw_name}': data in column(s) "
                f"{spill} with no header; an unquoted thousands "
                "separator splits the cell and truncates the cost")
            continue
        if nrm in _SKIP_EXACT or nrm.startswith("subtotal"):
            skipped.append(str(raw_name).strip())
            continue
        if nrm == _RESIDUAL_ROW:
            errors.append(
                f"row {n} '{raw_name}': blended residual row; itemize "
                "Section B (joists, roof deck, composite deck, anchor "
                "rods) before conversion, a blend cannot be split "
                "into cost codes mechanically")
            continue
        if "management reserve" in nrm:
            errors.append(
                f"row {n} '{raw_name}': management reserve is set by "
                "--reserve-pct or --reserve-amount, not in the source "
                "breakdown (P16: explicit input)")
            continue
        code = alias_to_code.get(nrm)
        if code is None:
            errors.append(
                f"row {n} '{raw_name}': no matching cost code; codes "
                "mirror the -GP lines (fab, erection, joists, "
                "composite deck, roof deck, anchor rods, stairs, misc "
                "metals, shop drawings, G&A)")
            continue
        cost = _money(pick("cost"))
        if cost is None:
            rev = _money(pick("revenue"))
            gp = _money(pick("gp_pct"))
            if rev is not None and gp is not None:
                if gp >= 1.0:
                    gp = gp / 100.0
                cost = round(rev * (1.0 - gp), 2)
        if cost is None:
            errors.append(f"row {n} '{raw_name}': no readable cost "
                          "(need Cost, or Revenue plus GP %)")
            continue
        if cost < 0:
            errors.append(f"row {n} '{raw_name}': negative cost "
                          f"{cost}")
            continue
        rev = _money(pick("revenue"))
        if rev is not None:
            has_revenue = True
            revenue_total += rev
        if code in costs:
            warnings.append(f"'{raw_name}' maps to {code} which "
                            "already has a row; costs summed")
            costs[code] = round(costs[code] + cost, 2)
        else:
            costs[code] = round(cost, 2)
    if errors:
        raise ValueError(f"source breakdown {p.name}:\n  "
                         + "\n  ".join(errors))
    if not costs:
        raise ValueError(f"source breakdown {p.name} has no cost "
                         "lines")
    return {"costs": costs,
            "revenue_total": round(revenue_total, 2)
            if has_revenue else None,
            "skipped": skipped, "warnings": warnings}


# -- margin math (P16) ----------------------------------------------------

def compute_budget(contract_value, margin, reserve_pct,
                   reserve_amount, source_costs) -> dict:
    """Contract value minus margin is the cost baseline; reserve is
    carved as an explicit line; the remainder spreads across the cost
    codes pro-rata by the source -GP line costs. Deterministic Python
    arithmetic only (constitution Section 11)."""
    if contract_value is None or contract_value <= 0:
        raise ValueError("contract value must be a positive amount")
    cost_baseline = round(contract_value * (1.0 - margin), 2)
    if reserve_pct is not None:
        # reserve_pct arrives as a decimal fraction, same convention
        # as margin (_pct_or_decimal: 2 and 0.02 both mean 2 percent).
        reserve = round(cost_baseline * reserve_pct, 2)
    else:
        reserve = round(float(reserve_amount), 2)
    if reserve < 0:
        raise ValueError("management reserve cannot be negative")
    if reserve >= cost_baseline:
        raise ValueError(f"management reserve {reserve} swallows the "
                         f"cost baseline {cost_baseline}")
    allocatable = round(cost_baseline - reserve, 2)

    src_total = round(sum(source_costs.values()), 2)
    if src_total <= 0:
        raise ValueError("source breakdown costs sum to zero; "
                         "nothing to allocate against")
    scale = allocatable / src_total

    warnings = []
    if margin == 0.0:
        warnings.append("target margin is zero: the budget equals "
                        "the contract value")
    if abs(1.0 - scale) > _SCALE_WARN:
        warnings.append(
            f"scale factor {scale:.4f}: the awarded cost baseline "
            f"({allocatable:,.2f}) sits far from the bid cost basis "
            f"({src_total:,.2f}); confirm margin and source with "
            "Owner before freeze")

    lines = []
    for code, name, _aliases in COST_CODES:
        src = source_costs.get(code, 0.0)
        share = src / src_total
        budget = round(src * scale, 2)
        note = ""
        if code not in source_costs:
            note = "no cost line in the source breakdown"
        elif code == "GA":
            note = ("overhead allocation from the source G&A line; "
                    "no WBS line")
        lines.append({"code": code, "name": name, "source": src,
                      "share": share, "budget": budget, "note": note})

    # Cent-rounding residual lands on the largest line so the lines
    # plus reserve reproduce the cost baseline exactly.
    assigned = round(sum(ln["budget"] for ln in lines), 2)
    residual = round(allocatable - assigned, 2)
    if abs(residual) > 0:
        biggest = max(lines, key=lambda ln: ln["budget"])
        biggest["budget"] = round(biggest["budget"] + residual, 2)

    return {"cost_baseline": cost_baseline, "reserve": reserve,
            "allocatable": allocatable, "scale": scale,
            "source_total": src_total, "lines": lines,
            "warnings": warnings}


# -- WBS template (P15) ---------------------------------------------------

def _wbs_rows(budget_by_code) -> list:
    """PC2 template rows. Levels per the commission: shop drawings,
    procurement, fab by sequence/area, delivery, erection by sequence,
    deck, closeout. Planned units, hours, and dates stay blank: the
    operator fills them from the takeoff and the schedule, the tool
    never invents quantities. One WBS line owns each budget dollar."""
    def amt(code):
        v = budget_by_code.get(code, 0.0)
        return v if v > 0 else None

    rows = []
    shd = amt("SHD")
    rows.append((
        "SD-01", "Shop drawing production and approval cycle", "SHD",
        "milestone", None, "credit", None, shd, None, None,
        "Shop drawings",
        "Approval cycle rule of credit: issued 20, approved 75, "
        "released 100; Ivan reviews before issue",
        "Approval cycle delay; log submittals in the correspondence "
        "register" if shd else
        "Approval cycle delay; log submittals in the correspondence "
        "register. No SHD cost line on this bid: shop drawing cost is "
        "folded into the FAB and ERE rates"))
    if amt("JST"):
        rows.append((
            "PR-01", "Joist package buyout and PO", "JST",
            "milestone", None, "credit", None, amt("JST"), None, None,
            "Procurement",
            "PO terms mirror contract flowdowns; supplier names stay "
            "internal (Tier 1)",
            "Joist lead time gates erection start; joist erection "
            "labor is carried in the ERE lines"))
    if amt("ANC"):
        rows.append((
            "PR-02", "Anchor rod package buyout and PO", "ANC",
            "milestone", None, "credit", None, amt("ANC"), None, None,
            "Procurement",
            "Size and grade verified against the column schedule "
            "before PO",
            "Anchor delivery gates foundations; setting labor is "
            "carried in the ERE lines"))
    if amt("FAB"):
        rows.append((
            "FAB-S1",
            "Fabrication, all sequences. Split by sequence or area "
            "before freeze", "FAB", "production", None, "tons", None,
            amt("FAB"), None, None, "Fabrication",
            "Fit-up and weld QC per AWS D1.1; Mario signs off",
            "Template holds the full FAB budget on one line; when "
            "sequences are set, split so the FAB lines sum to the "
            "FAB code total"))
    rows.append((
        "DEL-01", "Delivery to site, by load", "FAB", "production",
        None, "loads", None, None, None, None, "Delivery",
        "Piece-mark load lists checked at shipping and receiving",
        "Freight is carried inside the FAB rate; allocate budget "
        "here only if tracked separately, and keep the FAB lines "
        "summing to the FAB code total"))
    if amt("ERE"):
        rows.append((
            "ERE-S1",
            "Erection, all sequences. Split by sequence before "
            "freeze", "ERE", "production", None, "tons", None,
            amt("ERE"), None, None, "Erection",
            "Plumb, bolt-up, and torque verification per sequence",
            "Site access and GC readiness; tag client-caused delays "
            "CLIENT in the shop log (feeds contract-notice)"))
    if amt("CDK"):
        rows.append((
            "DCK-01", "Composite floor deck supply and install",
            "CDK", "production", None, "sf", None, amt("CDK"), None,
            None, "Deck",
            "Attachment pattern and sidelap fastening QC",
            "Deck supply and install always in Your Company scope; "
            "buyout tracked under 04 Procurement"))
    if amt("RDK"):
        rows.append((
            "DCK-02", "Roof deck supply and install", "RDK",
            "production", None, "sf", None, amt("RDK"), None, None,
            "Deck", "Attachment pattern and sidelap fastening QC",
            "Deck supply and install always in Your Company scope; "
            "buyout tracked under 04 Procurement"))
    if amt("STR"):
        rows.append((
            "MM-01", "Stairs fabrication and erection", "STR",
            "production", None, "pieces", None, amt("STR"), None,
            None, "Fabrication",
            "Handrail, tread, and landing checks per drawings",
            "Stair geometry changes are variation candidates; track "
            "in the obligations register"))
    if amt("MSC"):
        rows.append((
            "MM-02", "Misc metals fabrication and erection", "MSC",
            "production", None, "pieces", None, amt("MSC"), None,
            None, "Fabrication",
            "Scope list reconciled against contract exhibits",
            "Misc metals scope creep is a variation candidate; track "
            "departures"))
    rows.append((
        "CLO-01", "Closeout: punch list, final documents, warranty",
        "CLO", "milestone", None, "credit", None, None, None, None,
        "Closeout",
        "Punch list closed; warranty letter and final documents "
        "delivered per contract",
        "Retainage release rides on closeout; no direct cost line, "
        "effort carried in ERE and G&A"))
    return rows


# -- workbook writer ------------------------------------------------------

def _write_workbook(target, meta, lines, wbs) -> None:
    import openpyxl
    from openpyxl.styles import Font

    bold = Font(bold=True)
    usd = "$#,##0.00"

    wb = openpyxl.Workbook()
    b = wb.active
    b.title = "BUDGET"
    for col, width in (("A", 24), ("B", 44), ("C", 18), ("D", 18),
                       ("E", 10), ("F", 64)):
        b.column_dimensions[col].width = width

    b.append(["PC1 BUDGET", meta["project_id"]])
    b["A1"].font = bold
    b.append(["Status", _DRAFT_MARK])
    b.append(["Classification", _CLASSIFICATION])
    b.append(["Version", f"v{meta['version']}"])
    b.append(["Generated (UTC)", meta["generated"]])
    b.append(["Source breakdown", meta["source_name"]])
    b.append(["Contract value", meta["contract_value"]])
    b.append(["Target margin", meta["margin"],
              "input at award review; company target lives in "
              "bridge/bid_rates.py (net_target_gp_pct), by pointer "
              "per P10"])
    # Note cells must not start with "=": openpyxl would store them
    # as formulas, Excel would show #NAME?, and the freeze gate's
    # uncached-formula check would refuse its own file.
    b.append(["Cost baseline", meta["cost_baseline"],
              "equals contract value * (1 - target margin), per P16"])
    b.append(["Management reserve", meta["reserve"],
              meta["reserve_note"]])
    b.append(["Allocatable to cost codes", meta["allocatable"],
              "equals cost baseline minus management reserve"])
    b.append(["Scale vs source costs", round(meta["scale"], 6),
              "equals allocatable / source direct cost total "
              f"({meta['source_total']:,.2f})"])
    for r in (7, 9, 10, 11):
        b.cell(row=r, column=2).number_format = usd
    b.append([])
    b.append(["Cost Code", "Line Item", "Budget Cost", "Source Cost",
              "Share", "Notes"])
    hdr = b.max_row
    for c in range(1, 7):
        b.cell(row=hdr, column=c).font = bold
    for ln in lines:
        b.append([ln["code"], ln["name"], ln["budget"], ln["source"],
                  ln["share"], ln["note"]])
        r = b.max_row
        b.cell(row=r, column=3).number_format = usd
        b.cell(row=r, column=4).number_format = usd
        b.cell(row=r, column=5).number_format = "0.0%"
    b.append([MR_CODE, MR_NAME, meta["reserve"], None, None,
              "explicit named line (P16); drawdown by approved "
              "change only, never auto-spent"])
    b.cell(row=b.max_row, column=3).number_format = usd
    b.append([None, "Cost baseline total", meta["cost_baseline"]])
    total_row = b.max_row
    b.cell(row=total_row, column=2).font = bold
    b.cell(row=total_row, column=3).font = bold
    b.cell(row=total_row, column=3).number_format = usd
    b.append([])
    b.append(["Note", "G&A is overhead and MR is reserve: neither is "
                      "scheduled work, neither gets a WBS line; PC4 "
                      "BAC covers the work codes only"])
    b.append(["Note", "WBS budget cells must sum to each work cost "
                      "code total; one line owns each dollar"])

    w = wb.create_sheet("WBS")
    for col, width in (("A", 10), ("B", 52), ("C", 11), ("D", 14),
                       ("E", 14), ("F", 8), ("G", 14), ("H", 16),
                       ("I", 12), ("J", 12), ("K", 18), ("L", 56),
                       ("M", 64)):
        w.column_dimensions[col].width = width
    w.append(list(WBS_HEADERS))
    for c in range(1, len(WBS_HEADERS) + 1):
        w.cell(row=1, column=c).font = bold
    for row in wbs:
        w.append(list(row))
        w.cell(row=w.max_row, column=8).number_format = usd

    wb.save(str(target))
    wb.close()


# -- workbook readers (freeze and verify) ---------------------------------

def _flag_state(wb) -> dict:
    """Mirror of the PC4 reader's BASELINE flag scan."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS,
                                max_col=_SCAN_COLS):
            for cell in row:
                val = cell.value
                up = val.strip().upper() if isinstance(val, str) else ""
                if up == "BASELINE" or (up.startswith("BASELINE") and
                                        ("FROZEN" in up or
                                         _DATE_RX.search(up))):
                    frozen_date = ""
                    for sib in row:
                        # Mirror the reader's _to_date tolerance:
                        # Excel may retype the cell as a real date.
                        if isinstance(sib.value, datetime):
                            frozen_date = sib.value.date().isoformat()
                            break
                        if isinstance(sib.value, date):
                            frozen_date = sib.value.isoformat()
                            break
                        s = str(sib.value or "").strip()
                        if _DATE_RX.fullmatch(s):
                            frozen_date = s
                            break
                    return {"frozen": True, "frozen_date": frozen_date}
    return {"frozen": False, "frozen_date": ""}


def _file_frozen(path) -> bool:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True,
                                read_only=True)
    try:
        return _flag_state(wb)["frozen"]
    finally:
        wb.close()


def _read_budget_table(wb) -> dict:
    """Read back our own BUDGET sheet: {code: budget} plus meta."""
    if "BUDGET" not in wb.sheetnames:
        raise ValueError("no BUDGET sheet; not a budget_convert file")
    ws = wb["BUDGET"]
    budgets = {}
    meta = {}
    hdr = None
    for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS + 30):
        a = row[0].value
        if isinstance(a, str) and a.strip() == "Cost Code":
            hdr = row[0].row
            break
        if isinstance(a, str) and len(row) > 1:
            meta[a.strip()] = row[1].value
    if hdr is None:
        raise ValueError("no Cost Code table on the BUDGET sheet")
    for row in ws.iter_rows(min_row=hdr + 1,
                            max_row=hdr + len(COST_CODES) + 3):
        code = str(row[0].value or "").strip().upper()
        if not code:
            break
        budgets[code] = _money(row[2].value) or 0.0
    return {"budgets": budgets, "meta": meta}


def _read_wbs_table(wb) -> list:
    wbs_ws = None
    for ws in wb.worksheets:
        if "wbs" in ws.title.lower():
            wbs_ws = ws
            break
    if wbs_ws is None:
        raise ValueError("no WBS sheet; not a budget_convert file")
    col_of = {}
    for idx, cell in enumerate(next(wbs_ws.iter_rows(min_row=1,
                                                     max_row=1))):
        name = _norm(cell.value)
        if not name:
            continue
        for canonical, names in _GATE_ALIASES.items():
            if canonical not in col_of and name in names:
                col_of[canonical] = idx
    # Fail closed: a header the gate cannot map must refuse, never
    # degrade the partition check into all-blank rows.
    needed = ("wbs_line", "cost_code", "budget_cost")
    gone = [k for k in needed if k not in col_of]
    if gone:
        raise ValueError(
            "WBS sheet is missing the column(s) "
            + ", ".join(gone).replace("_", " ")
            + "; the freeze gate fails closed on renamed headers")
    rows = []
    streak = 0
    for row in wbs_ws.iter_rows(min_row=2, max_row=_MAX_SRC_ROWS):
        cells = [c.value for c in row]

        def pick(k):
            idx = col_of.get(k)
            return cells[idx] if idx is not None and idx < len(cells) \
                else None

        line = str(pick("wbs_line") or "").strip()
        if not line:
            streak += 1
            if streak >= _EMPTY_STREAK:
                break
            continue
        streak = 0
        rows.append({
            "wbs_line": line,
            "cost_code": str(pick("cost_code") or "").strip().upper(),
            "progress_type": str(pick("progress_type") or "").strip(),
            "planned_units": _money(pick("planned_units")),
            "planned_hours": _money(pick("planned_hours")),
            "budget_cost": _money(pick("budget_cost")),
            "start_date": pick("start_date"),
            "end_date": pick("end_date"),
        })
    return rows


def _partition_report(budgets, wbs) -> dict:
    """Check the BAC partition rule: per work code, the WBS budget
    cells either sum to the code total, or blanks remain to allocate.
    Overcounts and unassigned dollars are problems; blanks under an
    under-covered code are incomplete, not broken."""
    problems = []
    incomplete = []
    by_code = {}
    for r in wbs:
        by_code.setdefault(r["cost_code"], []).append(r)
    for code in WORK_CODES:
        total = budgets.get(code, 0.0)
        rows = by_code.get(code, [])
        if total <= 0:
            continue
        if not rows:
            problems.append(f"{code} carries budget {total:,.2f} but "
                            "has no WBS line")
            continue
        filled = round(sum(r["budget_cost"] or 0.0 for r in rows), 2)
        blanks = [r["wbs_line"] for r in rows
                  if r["budget_cost"] is None]
        if filled > total + _CENTS_TOL:
            problems.append(
                f"{code}: WBS lines sum to {filled:,.2f}, over the "
                f"code total {total:,.2f}; one line owns each dollar")
        elif abs(filled - total) > _CENTS_TOL:
            if blanks:
                incomplete.append(
                    f"{code}: {total - filled:,.2f} not yet allocated "
                    "across blank lines " + ", ".join(blanks))
            else:
                problems.append(
                    f"{code}: WBS lines sum to {filled:,.2f}, short "
                    f"of the code total {total:,.2f}, and no blank "
                    "line remains to absorb it")
    # No dollar may hide under a code the loop above never checked:
    # an unknown code (typo), a non-work code (GA, MR, CLO), or a
    # work code whose budget line is zero.
    checked = {c for c in WORK_CODES if budgets.get(c, 0.0) > 0}
    for r in wbs:
        if r["budget_cost"] and r["cost_code"] not in checked:
            problems.append(
                f"{r['wbs_line']} carries budget "
                f"{r['budget_cost']:,.2f} under code "
                f"{r['cost_code'] or '(blank)'} which has no budget "
                "line to own it")
    return {"problems": problems, "incomplete": incomplete}


def _completeness(wbs) -> list:
    """Rows that carry budget need units (production), hours, and
    dates before freeze. Schedule-only rows (blank budget) are
    exempt: their flags belong to PC4, not the freeze gate."""
    out = []
    for r in wbs:
        if not r["budget_cost"]:
            continue
        missing = []
        if (r["progress_type"] == "production"
                and r["planned_units"] is None):
            missing.append("planned units")
        if r["planned_hours"] is None:
            missing.append("planned hours")
        if not r["start_date"] or not r["end_date"]:
            missing.append("start/end dates")
        if missing:
            out.append(f"{r['wbs_line']}: " + ", ".join(missing))
    return out


# -- backup (operating rule: snapshot before overwrite) -------------------

def _log_changelog(text) -> None:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    log = _REPO_ROOT / "_handoff" / "changelog.md"
    with open(log, "a", encoding="utf-8") as f:
        f.write(f"\n- {ts} budget_convert: {text}\n")


def _backup(path, action) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    dest_dir = _REPO_ROOT / "_handoff" / "backups" / ts
    dest_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(path), str(dest_dir / Path(path).name))
    _log_changelog(f"{action}: snapshot of {Path(path).name} to "
                   f"_handoff/backups/{ts}/")
    return str(dest_dir)


# -- convert --------------------------------------------------------------

def _financials_dir(project_folder) -> Path:
    folder = Path(project_folder)
    if not folder.is_dir():
        raise ValueError(f"project folder not found: {folder}")
    for d in sorted(folder.iterdir()):
        if d.is_dir() and d.name.startswith("09"):
            return d
    raise ValueError(f"no '09 Financials -GP CONFIDENTIAL' folder "
                     f"under {folder}")


def _contract_files(project_folder) -> list:
    folder = Path(project_folder)
    for d in sorted(folder.iterdir()):
        if d.is_dir() and d.name.startswith("01"):
            return [f.name for f in sorted(d.iterdir())
                    if f.is_file()
                    and not f.name.startswith((".", "~$"))]
    return []


def _existing_versions(out_dir, project_id) -> list:
    pat = re.compile(re.escape(project_id) + r"_budget_v(\d+)\.xlsx$",
                     re.IGNORECASE)
    found = []
    for p in sorted(Path(out_dir).glob("*.xlsx")):
        m = pat.fullmatch(p.name)
        if m:
            found.append((int(m.group(1)), p))
    return sorted(found)


def convert(source, contract_value, margin, reserve_pct,
            reserve_amount, project_folder=None, out_dir=None,
            project_id=None, force=False) -> dict:
    """Run the PC1 conversion and write the draft budget xlsx."""
    if (project_folder is None) == (out_dir is None):
        raise ValueError("pass exactly one of --project-folder or "
                         "--out")
    contract_note = ""
    if project_folder:
        names = _contract_files(project_folder)
        if not names:
            return {"status": "BLOCKED_NO_CONTRACT",
                    "error": "the 01 Contract folder is empty; no "
                             "signed contract, no budget (P14: "
                             "baselines or nothing)",
                    "fix": "drop the executed contract into "
                           f"{project_folder}/01 Contract and rerun"}
        contract_note = "; ".join(names[:5])
        out = _financials_dir(project_folder)
        if project_id is None:
            project_id = Path(project_folder).name.split()[0]
    else:
        out = Path(out_dir)
        out.mkdir(parents=True, exist_ok=True)
        if project_id is None:
            raise ValueError("--project-id is required with --out")

    src = read_source(source)
    result = compute_budget(contract_value, margin, reserve_pct,
                            reserve_amount, src["costs"])
    warnings = list(src["warnings"]) + list(result["warnings"])
    if src["revenue_total"] is not None:
        diff = abs(src["revenue_total"] - contract_value)
        if diff > contract_value * _CONTRACT_WARN:
            warnings.append(
                f"contract value {contract_value:,.2f} sits more "
                f"than {_CONTRACT_WARN:.0%} from the source revenue "
                f"total {src['revenue_total']:,.2f}; confirm the "
                "award amount against the signed contract")

    # Operator-controlled strings land inside the 40x15 cell window
    # the PC4 reader scans for the BASELINE flag; a name starting
    # with BASELINE plus a date would freeze a draft from the outside.
    for label, value in (("--project-id", project_id),
                         ("--source filename", Path(source).name)):
        if str(value).strip().upper().startswith("BASELINE"):
            raise ValueError(
                f"{label} {value!r} would trip the PC4 BASELINE "
                "flag scan; rename it (names must not start with "
                "'baseline')")

    vers = _existing_versions(out, project_id)
    target = None
    if not vers:
        version = 1
    else:
        last_n, last_p = vers[-1]
        if _file_frozen(last_p):
            # P14: the approved baseline keeps governing. The new
            # draft coexists; freeze() moves prior versions to
            # superseded/ once the replacement is approved.
            version = last_n + 1
            warnings.append(
                f"v{last_n} is a frozen baseline; it stays the "
                f"governing file for PC4 until v{version} freezes, "
                "then moves to superseded/")
        elif force:
            version = last_n
            target = last_p
            _backup(last_p, "overwrite draft")
        else:
            return {"status": "DRAFT_EXISTS",
                    "error": f"draft {last_p.name} already exists",
                    "fix": "review and freeze it, or rerun with "
                           "--force to overwrite the draft (a backup "
                           "snapshot is taken first)"}

    if reserve_pct is not None:
        reserve_note = (f"{reserve_pct * 100:g}% of cost baseline, "
                        "input at award review; drawdown by approved "
                        "change only")
    else:
        reserve_note = ("fixed amount, input at award review; "
                        "drawdown by approved change only")
    meta = {
        "project_id": project_id,
        "version": version,
        "generated": datetime.now(timezone.utc).isoformat(
            timespec="seconds"),
        "source_name": Path(source).name,
        "contract_value": round(float(contract_value), 2),
        "margin": margin,
        "cost_baseline": result["cost_baseline"],
        "reserve": result["reserve"],
        "reserve_note": reserve_note,
        "allocatable": result["allocatable"],
        "scale": result["scale"],
        "source_total": result["source_total"],
    }
    budget_by_code = {ln["code"]: ln["budget"]
                      for ln in result["lines"]}
    if target is None:
        # Zero-padded so v10 sorts after v02 for both the reader's
        # alphabetical pick and a human eyeballing the folder.
        target = Path(out) / f"{project_id}_budget_v{version:02d}.xlsx"
    wbs_rows = _wbs_rows(budget_by_code)
    _write_workbook(target, meta, result["lines"], wbs_rows)
    if _file_frozen(target):
        # Belt and suspenders behind the name guard above: no draft
        # may ever scan as frozen, whatever future meta gets added.
        target.unlink()
        raise ValueError(
            "the freshly written draft scans as frozen; some input "
            "string trips the PC4 BASELINE flag rule. File deleted; "
            "rename the offending input and rerun")

    info = {
        "status": "DRAFT_WRITTEN",
        "path": str(target),
        "version": f"v{version}",
        "project_id": project_id,
        "contract_value": f"{meta['contract_value']:,.2f}",
        "target_margin": f"{margin:.4f}",
        "cost_baseline": f"{result['cost_baseline']:,.2f}",
        "management_reserve": f"{result['reserve']:,.2f}",
        "allocatable": f"{result['allocatable']:,.2f}",
        "scale_factor": f"{result['scale']:.4f}",
        "lines": len(result["lines"]) + 1,
        "wbs_rows": len(wbs_rows),
        "next_step": "review with Owner, fill WBS units, hours, "
                     "and dates, then: py -m "
                     "takeoff_pipeline.budget_convert freeze "
                     f"\"{target}\" --approved-by \"<name>\"",
    }
    if contract_note:
        info["contract_files"] = contract_note
    if src["skipped"]:
        info["skipped_rows"] = ", ".join(src["skipped"])
    if warnings:
        info["warnings"] = " | ".join(warnings)
    return info


# -- freeze (P14) ---------------------------------------------------------

def freeze(path, approved_by, frozen_date=None,
           allow_incomplete=False) -> dict:
    """Stamp the draft as the frozen baseline. After this, the file
    is never edited; changes require a new version (P14)."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        raise ValueError(f"budget xlsx not found: {p}")
    if not approved_by or not str(approved_by).strip():
        raise ValueError("--approved-by is required; freezing is the "
                         "approval act (P14)")
    if frozen_date is None:
        frozen_date = datetime.now(timezone.utc).date().isoformat()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", frozen_date):
        raise ValueError(f"--date must be YYYY-MM-DD; got "
                         f"{frozen_date!r}")

    wb = openpyxl.load_workbook(str(p))
    try:
        if _flag_state(wb)["frozen"]:
            return {"status": "ALREADY_FROZEN", "path": str(p),
                    "error": "this file already carries a BASELINE "
                             "flag; edits require a new version, "
                             "never an overwrite (P14)",
                    "fix": "rerun convert to produce the next "
                           "version"}
        loc = None
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS,
                                    max_col=_SCAN_COLS):
                for cell in row:
                    if (isinstance(cell.value, str)
                            and cell.value.startswith("DRAFT")):
                        loc = (ws, cell.row, cell.column)
                        break
                if loc:
                    break
            if loc:
                break
        if loc is None:
            raise ValueError(f"{p.name} has no DRAFT status cell; "
                             "not a budget_convert draft")

        # Formula cells: the writable view shows formula strings, the
        # gate must check the cached values PC4 will read.
        formulas = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=_MAX_SRC_ROWS,
                                    max_col=20):
                for cell in row:
                    v = cell.value
                    if ((isinstance(v, str) and v.startswith("="))
                            or "Formula" in type(v).__name__):
                        formulas.append((ws.title, cell.coordinate))

        wb2 = openpyxl.load_workbook(str(p), data_only=True)
        try:
            uncached = [f"{t}!{coord}" for t, coord in formulas
                        if wb2[t][coord].value is None]
            if uncached:
                return {"status": "REFUSED",
                        "error": "formula cells with no cached "
                                 "value: " + ", ".join(uncached[:10]),
                        "fix": "open the file in Excel and save so "
                               "the values are cached, or type plain "
                               "numbers; the gate refuses to treat a "
                               "formula as blank"}
            table = _read_budget_table(wb2)
            budgets = table["budgets"]
            wbs = _read_wbs_table(wb2)
        finally:
            wb2.close()

        # The stamped cost baseline must still equal lines plus MR;
        # a hand-bumped line otherwise freezes a self-contradicting
        # baseline (P16 math broken between convert and freeze).
        baseline = _money(table["meta"].get("Cost baseline"))
        line_sum = round(sum(budgets.values()), 2)
        if baseline is None:
            return {"status": "REFUSED",
                    "error": "the Cost baseline meta cell is missing "
                             "or not a number",
                    "fix": "restore the BUDGET sheet meta block or "
                           "reconvert"}
        if abs(line_sum - baseline) > _CENTS_TOL:
            return {"status": "REFUSED",
                    "error": f"cost lines plus reserve sum to "
                             f"{line_sum:,.2f}, not the stamped cost "
                             f"baseline {baseline:,.2f}",
                    "fix": "undo hand edits to the BUDGET sheet or "
                           "reconvert with the new numbers; the "
                           "baseline must equal contract value minus "
                           "margin (P16)"}

        part = _partition_report(budgets, wbs)
        if part["problems"]:
            return {"status": "REFUSED",
                    "error": "BAC partition broken: "
                             + " | ".join(part["problems"]),
                    "fix": "fix the WBS budget cells so each work "
                           "code sums exactly, then freeze"}
        missing = _completeness(wbs) + part["incomplete"]
        if missing and not allow_incomplete:
            return {"status": "REFUSED_INCOMPLETE",
                    "error": "WBS lines are not filled in: "
                             + " | ".join(missing),
                    "fix": "fill planned units, hours, dates, and "
                           "allocations from the takeoff and the "
                           "schedule, or freeze anyway with "
                           "--allow-incomplete (PC4 will flag the "
                           "gaps)"}

        _backup(p, f"freeze approved by {approved_by}")
        ws, r, c = loc
        ws.cell(row=r, column=c,
                value=f"BASELINE frozen {frozen_date}")
        ws.cell(row=r, column=c + 1, value=frozen_date)
        ws.cell(row=r, column=c + 2,
                value=f"approved by {approved_by}")
        wb.save(str(p))
    finally:
        wb.close()

    # The new baseline governs from this moment: prior versions move
    # to superseded/ so the PC4 reader's first-alphabetical pick in
    # the 09 folder cannot land on an old file. Files are moved, not
    # deleted (P14 history stays on disk), and the move is logged.
    moved = []
    name_m = re.fullmatch(r"(.+)_budget_v(\d+)\.xlsx", p.name,
                          re.IGNORECASE)
    if name_m:
        my_n = int(name_m.group(2))
        sup = p.parent / "superseded"
        for n, q in _existing_versions(p.parent, name_m.group(1)):
            if n >= my_n or q.resolve() == p.resolve():
                continue
            sup.mkdir(exist_ok=True)
            dest = sup / q.name
            if dest.exists():
                stamp = datetime.now(timezone.utc).strftime(
                    "%Y%m%dT%H%M%SZ")
                dest = sup / (q.stem + "_" + stamp + q.suffix)
            shutil.move(str(q), str(dest))
            moved.append(q.name)
        if moved:
            _log_changelog(f"freeze of {p.name} superseded "
                           + ", ".join(moved)
                           + "; moved to superseded/")

    info = {"status": "FROZEN", "path": str(p),
            "frozen_date": frozen_date, "approved_by": approved_by,
            "note": "edits from here require a new version, never an "
                    "overwrite (P14)"}
    if moved:
        info["superseded"] = ", ".join(moved)
    if missing:
        info["frozen_with_gaps"] = " | ".join(missing)
    return info


# -- verify ---------------------------------------------------------------

def verify(path) -> dict:
    """Check a budget xlsx against this module's own layout, the BAC
    partition rule, and the real PC4 loader. The cross-check imports
    bridge/project_controls and calls its loader directly so reader
    and writer cannot drift; a skip is reported, never silent."""
    import openpyxl
    p = Path(path)
    if not p.exists():
        raise ValueError(f"budget xlsx not found: {p}")
    problems = []
    notes = []
    wb = openpyxl.load_workbook(str(p), data_only=True)
    try:
        flag = _flag_state(wb)
        table = _read_budget_table(wb)
        budgets = table["budgets"]
        meta = table["meta"]
        wbs = _read_wbs_table(wb)
    finally:
        wb.close()

    baseline = _money(meta.get("Cost baseline"))
    reserve = budgets.get(MR_CODE, 0.0)
    line_sum = round(sum(v for k, v in budgets.items()
                         if k != MR_CODE) + reserve, 2)
    if baseline is None:
        problems.append("Cost baseline meta row missing")
    elif abs(line_sum - baseline) > _CENTS_TOL:
        problems.append(f"cost lines plus reserve sum to "
                        f"{line_sum:,.2f}, not the cost baseline "
                        f"{baseline:,.2f}")
    if MR_CODE not in budgets:
        problems.append("management reserve line missing (P16: "
                        "explicit named line)")
    part = _partition_report(budgets, wbs)
    problems.extend(part["problems"])
    notes.extend(part["incomplete"])

    siblings = [q for q in sorted(p.parent.glob("*.xlsx"))
                if not q.name.startswith("~$")]
    pick = None
    for q in siblings:
        low = q.name.lower()
        if "budget" in low or "baseline" in low:
            pick = q
            break
    if pick is None and siblings:
        pick = siblings[0]
    if pick is not None and pick != p:
        # A frozen lower version outranking this draft is the P14
        # review window working as designed, not a problem.
        ver_rx = re.compile(r"(.+)_budget_v(\d+)\.xlsx", re.IGNORECASE)
        mp = ver_rx.fullmatch(p.name)
        mq = ver_rx.fullmatch(pick.name)
        if (not flag["frozen"] and mp and mq
                and mq.group(1).lower() == mp.group(1).lower()
                and int(mq.group(2)) < int(mp.group(2))
                and _file_frozen(pick)):
            notes.append(f"PC4 currently reads {pick.name}: the old "
                         "baseline governs until this draft freezes "
                         "(P14); freeze moves it to superseded/")
        else:
            problems.append(f"the PC4 reader would pick {pick.name} "
                            "from this folder, not this file; move "
                            "stray xlsx files out of the 09 folder")

    cross = "skipped"
    try:
        from bridge.project_controls import _load_baseline
        res = _load_baseline(p)
        if "error" in res:
            cross = f"PC4 loader error: {res['error']}"
            if not flag["frozen"]:
                cross += " (expected for a draft: no BASELINE flag)"
            else:
                problems.append(cross)
        else:
            cross = (f"PC4 loader parsed {len(res['lines'])} WBS "
                     f"lines, baseline_flag="
                     f"{res['baseline_flag']}, frozen_date="
                     f"{res['frozen_date'] or 'none'}")
            for wmsg in res.get("warnings", []):
                notes.append(f"PC4 loader: {wmsg}")
    except Exception as e:  # cross-check is best-effort, never silent
        cross = (f"cross-check skipped ({e.__class__.__name__}: {e}); "
                 "run from the repo root so bridge/ imports")
        notes.append(cross)

    info = {
        "status": "FAIL" if problems else "PASS",
        "path": str(p),
        "frozen": flag["frozen"],
        "frozen_date": flag["frozen_date"] or "none",
        "cost_lines": len([k for k in budgets if k != MR_CODE]),
        "wbs_rows": len(wbs),
        "pc4_cross_check": cross,
    }
    if problems:
        info["problems"] = " | ".join(problems)
    if notes:
        info["notes"] = " | ".join(notes)
    return info


# -- CLI ------------------------------------------------------------------

def _flag_value(args, flag):
    if flag not in args:
        return None
    i = args.index(flag)
    if i + 1 >= len(args):
        raise ValueError(f"{flag} needs a value")
    return args[i + 1]


def _required(args, flag):
    v = _flag_value(args, flag)
    if v is None:
        raise ValueError(f"{flag} is required")
    return v


def main() -> int:
    usage = (
        "usage: py -m takeoff_pipeline.budget_convert convert "
        "--source <gp_lines.csv|.xlsx> --contract-value <usd> "
        "--margin <pct> (--reserve-pct <pct> | --reserve-amount "
        "<usd>) (--project-folder <dir> | --out <dir>) "
        "[--project-id <id>] [--force]\n"
        "       py -m takeoff_pipeline.budget_convert freeze "
        "<budget.xlsx> --approved-by <name> [--date YYYY-MM-DD] "
        "[--allow-incomplete]\n"
        "       py -m takeoff_pipeline.budget_convert verify "
        "<budget.xlsx>")
    args = sys.argv[1:]
    if not args:
        print(usage)
        return 2
    cmd = args[0]
    try:
        if cmd == "convert":
            contract_value = _money(_required(args,
                                              "--contract-value"))
            if contract_value is None:
                raise ValueError("--contract-value is not a number")
            margin = _pct_or_decimal(_required(args, "--margin"),
                                     "--margin")
            r_pct = _flag_value(args, "--reserve-pct")
            r_amt = _flag_value(args, "--reserve-amount")
            if (r_pct is None) == (r_amt is None):
                raise ValueError(
                    "management reserve is an explicit input (P16): "
                    "pass exactly one of --reserve-pct or "
                    "--reserve-amount; zero is allowed but must be "
                    "stated")
            reserve_pct = None
            reserve_amount = None
            if r_pct is not None:
                # Same convention as --margin: 2 and 0.02 both mean
                # 2 percent. Divergent percent conventions on one
                # command line invite silent 100x reserve errors.
                reserve_pct = _pct_or_decimal(r_pct, "--reserve-pct")
            else:
                reserve_amount = _money(r_amt)
                if reserve_amount is None or reserve_amount < 0:
                    raise ValueError("--reserve-amount must be a "
                                     "non-negative amount")
            info = convert(
                source=_required(args, "--source"),
                contract_value=contract_value,
                margin=margin,
                reserve_pct=reserve_pct,
                reserve_amount=reserve_amount,
                project_folder=_flag_value(args, "--project-folder"),
                out_dir=_flag_value(args, "--out"),
                project_id=_flag_value(args, "--project-id"),
                force="--force" in args)
        elif cmd == "freeze":
            if len(args) < 2 or args[1].startswith("--"):
                raise ValueError("freeze needs the budget xlsx path")
            info = freeze(args[1],
                          approved_by=_required(args, "--approved-by"),
                          frozen_date=_flag_value(args, "--date"),
                          allow_incomplete="--allow-incomplete"
                                           in args)
        elif cmd == "verify":
            if len(args) < 2 or args[1].startswith("--"):
                raise ValueError("verify needs the budget xlsx path")
            info = verify(args[1])
        else:
            print(usage)
            return 2
    except ValueError as e:
        print(e)
        print(usage)
        return 2
    for k, v in info.items():
        print(f"{k}: {v}")
    return 0 if info.get("status") in ("DRAFT_WRITTEN", "FROZEN",
                                       "PASS") else 1


if __name__ == "__main__":
    raise SystemExit(main())
