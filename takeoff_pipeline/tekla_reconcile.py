"""Tekla shop-model BOM reconciliation against the bid takeoff (T13/F5).

Diffs the bid takeoff xlsx (TAKEOFF_SCHEMA_V2, the Prompt 5 exporter
format) against a Tekla shop-model BOM export, per designation: piece
counts and tonnage. Section 07 acceptance: tonnage within 2 to 3
percent of the Tekla shop-model BOM. The verdict states the actual
number, never massaged.

Free accuracy audit (F5): the model gets built regardless. Findings
propose nothing. Owner and Ivan read the report.

Bid-side tonnage comes from the takeoff file's derived columns ONLY.
The exporter writes weight_lb as =D{r}*L{r} and tons as =M{r}/2000 on
LINEAR rows; lb_per_ft (column L) is the operator's input cell, filled
from bridge/aisc_validator.py output. openpyxl carries no cached value
for a formula cell the file has never computed, so this module
evaluates those two exporter formulas from the row's own qty and
lb_per_ft cells. It never recomputes a weight from any other source:
no AISC lookups, no LLM math (Hard Rule 5 stays in the validator).

The variance compares like for like. A COUNT-mode takeoff row never
carries derived tonnage by schema design, so a MATCHED designation
weighed on one side only is excluded from both totals and disclosed in
the report; its piece-count delta still stands. One-sided designations
stay in the totals: steel only one side knows about is a real miss.

Counts only, no pricing (P25). Standalone: stdlib plus openpyxl, no
bridge/ imports. The report and the ledger rows carry -GP CONFIDENTIAL
handling per Section 07: internal only, never client-facing.

Usage:
  py -m takeoff_pipeline.tekla_reconcile <takeoff.xlsx> <bom.csv|xlsx>
     [--job JOB] [--ledger-dir DIR] [--weight-unit lb|kg|ton]
"""

import argparse
import csv
import math
import re
from datetime import date
from fractions import Fraction
from pathlib import Path
from zipfile import BadZipFile

_PKG = Path(__file__).resolve().parent
LEDGER_DIR = _PKG / "ledger"
# Same six columns score_spike.py writes. One ledger, one header.
LEDGER_HEADER = ["date", "job", "test", "metric", "value", "notes"]
TEST_NAME = "tekla_reconcile_T13"

# Section 07: "Tonnage within 2-3% of Tekla shop-model BOM (T13)".
TARGET_INNER_PCT = 2.0
TARGET_OUTER_PCT = 3.0
TOP_MISSES = 10

KG_TO_LB = 2.2046226218

# TAKEOFF sheet header, row 2, columns A-O (TAKEOFF_SCHEMA_V2 3.2/4.2).
TAKEOFF_HEADERS = (
    "item_id", "designation", "mode", "qty", "unit", "primary_source",
    "secondary_source", "confidence", "sheet", "bbox", "notes",
    "lb_per_ft", "weight_lb", "tons", "formula_ref")

# The only formulas this module evaluates: the exporter's own derived
# chain, verbatim (export_xlsx.py writes =D{r}*L{r} and =M{r}/2000).
_TONS_FORMULA = re.compile(r"^=M(\d+)/2000$")
_WEIGHT_FORMULA = re.compile(r"^=D(\d+)\*L(\d+)$")

# BOM header keywords per the Prompt 8 contract: profile/designation,
# qty/count, length, weight. Synonyms cover common Tekla templates.
# Weak qty tokens match only an exact header cell, so a running-number
# or holes column cannot hijack qty when a real Qty column exists.
_PROFILE_KEYS = ("PROFILE", "DESIGNATION", "SECTION", "SHAPE")
_QTY_KEYS = ("QTY", "QUANTITY", "COUNT", "PCS", "PIECES")
_QTY_EXACT = ("NO", "NO.", "N", "#")
_LENGTH_KEY = "LENGTH"
_WEIGHT_KEY = "WEIGHT"

# Totals rows are matched on a letters-only cleanup, so "Total:",
# "Grand total" and "Sub-total" variants are all caught.
_SUMMARY_TOKENS = ("TOTAL", "TOTALS", "GRANDTOTAL", "SUM", "SUBTOTAL")


def _num(x):
    """float for a real finite number, else None. Bools are ints in
    Python and a TRUE cell must never count as 1.0 of anything."""
    if isinstance(x, bool) or not isinstance(x, (int, float)):
        return None
    f = float(x)
    return f if math.isfinite(f) else None


# -- designation normalization ----------------------------------------------

# Matching keys only. Stored designations stay raw on both sides per
# schema 3.2 (never an invented or normalized tag); the report prints
# the raw forms next to each key.

_JOIST = re.compile(r"^(\d+)(KCS|DLH|LH|K|G)0*(\d.*)?$")
_DECIMAL = re.compile(r"^\d*\.\d+$")
_ALPHA_PREFIX = re.compile(r"^(2?[A-Z]+)(.*)$")

# Families whose label dimensions are written as fractions. C and MC
# trail a decimal WEIGHT (C8X11.5) and WT both a decimal depth and a
# decimal weight (WT10.5X31), so those never convert. Round HSS
# diameters survive because they are not exact halves under 9.
_FRACTION_FAMILIES = ("HSS", "2L", "L", "PL")


def _frac(seg: str) -> str:
    """Decimal dimension to the fraction form drawings use: exact
    sixteenths under 3 inches (thickness) plus exact halves from 3 to
    under 9 (legs and sides, L3-1/2 and friends). The output keeps the
    hyphen on mixed numbers, never glued: a glued 13/16 would collide
    with the real 13/16 thickness. Anything else passes through with
    zero cleanup so .250, 0.25, and 1/4 share one key."""
    if not _DECIMAL.match(seg):
        return seg
    f = Fraction(seg)
    if (0 < f < 3 and f.denominator in (2, 4, 8, 16)) or \
            (3 <= f < 9 and f.denominator == 2):
        whole, rem = divmod(f.numerator, f.denominator)
        frac = f"{rem}/{f.denominator}"
        if whole and rem:
            return f"{whole}-{frac}"
        return str(whole) if not rem else frac
    seg = seg.rstrip("0").rstrip(".")
    return re.sub(r"^0(?=\.)", "", seg) or "0"


def _clean_decimal(seg: str) -> str:
    """Trailing-zero and leading-zero cleanup for non-fraction
    families, so 11.50 and 11.5 share one key."""
    if not _DECIMAL.match(seg):
        return seg
    seg = seg.rstrip("0").rstrip(".")
    return re.sub(r"^0(?=\.)", "", seg) or "0"


def normalize_designation(raw: str) -> str:
    """Cross-side matching key. Superset of the census normalizer
    (strip whitespace and inch/foot marks, uppercase) plus the canon a
    Tekla profile string needs to meet a drawing tag: unicode x, the
    legacy TS and Tekla FL prefixes, mixed numbers hyphenated before
    the space strip (2 1/2 and 2-1/2 and 2.5 share one key), decimal
    dimensions to fractions on fraction families, joist chord leading
    zeros (32LH07 = 32LH7)."""
    s = (raw or "").upper()
    s = s.replace("×", "X")
    s = re.sub(r"(\d)\s+(\d+/\d+)", r"\1-\2", s)
    s = re.sub(r"[\s\"']", "", s)
    s = re.sub(r"^TS(?=\d)", "HSS", s)
    s = re.sub(r"^FL(?=\d)", "PL", s)
    if not s:
        return s
    m = _JOIST.match(s)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3) or ''}"
    parts = s.split("X")
    first = _ALPHA_PREFIX.match(parts[0])
    family = first.group(1) if first else ""
    if family in _FRACTION_FAMILIES:
        # A dimension can ride any segment: glued to the prefix on
        # plates (PL.5X12), the legs in the middle, the thickness last.
        if first and first.group(2):
            parts[0] = family + _frac(first.group(2))
        parts[1:] = [_frac(p) for p in parts[1:]]
    else:
        parts = [parts[0]] + [_clean_decimal(p) for p in parts[1:]]
    return "X".join(parts)


# -- bid side: the takeoff xlsx ----------------------------------------------

def _resolve_tons(plain, cached, qty, lb_per_ft, weight_plain, row_idx,
                  warnings) -> float:
    """Tons for one TAKEOFF row, from the file's derived columns only.

    Order: a cached calc result if the file ever passed through a
    calculating app; a typed numeric (the validator forbids it, but a
    number in the file is still the file's number); else the exporter's
    own formula pair evaluated from this row's qty and lb_per_ft cells.
    A formula in any other form is skipped and warned, never guessed.
    """
    if _num(cached) is not None:
        return _num(cached)
    if _num(plain) is not None:
        return _num(plain)
    if plain is None or str(plain).strip() == "":
        return None
    m = _TONS_FORMULA.match(str(plain).strip())
    w = _WEIGHT_FORMULA.match(str(weight_plain or "").strip())
    if m and w and int(m.group(1)) == row_idx and \
            int(w.group(1)) == row_idx and int(w.group(2)) == row_idx:
        if _num(qty) is not None and _num(lb_per_ft) is not None:
            return _num(qty) * _num(lb_per_ft) / 2000.0
        warnings.append(
            f"row {row_idx}: LINEAR row has tons formula but no "
            "numeric lb_per_ft; row carries no tonnage")
        return None
    warnings.append(
        f"row {row_idx}: tons cell is not the exporter formula pair "
        f"({plain!r}); skipped, never recomputed")
    return None


def read_takeoff(path) -> dict:
    """Aggregate the TAKEOFF sheet per normalized designation.

    qty_ea sums COUNT rows (the piece count a model BOM can answer).
    lf and sf are informational. tons sums the derived column per the
    module contract. OPENING rows are deck deductions, not members,
    and are excluded (counted in the result for the method notes).
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(str(path))  # formulas as strings
    wb_data = openpyxl.load_workbook(str(path), data_only=True)
    if "TAKEOFF" not in wb.sheetnames:
        raise ValueError(f"no TAKEOFF sheet in {path.name}")
    ws = wb["TAKEOFF"]
    ws_data = wb_data["TAKEOFF"]

    header = tuple((ws.cell(row=2, column=c).value or "")
                   for c in range(1, 16))
    if tuple(header[:15]) != TAKEOFF_HEADERS:
        raise ValueError(
            f"{path.name} row 2 is not the TAKEOFF_SCHEMA_V2 header; "
            "is this a T8-format takeoff export?")

    job_stamp = ws.cell(row=1, column=2).value
    agg = {}
    warnings = []
    openings = 0
    removed = 0
    for r in range(3, ws.max_row + 1):
        designation = ws.cell(row=r, column=2).value
        mode = str(ws.cell(row=r, column=3).value or "").strip().upper()
        qty = ws.cell(row=r, column=4).value
        notes = str(ws.cell(row=r, column=11).value or "")
        if designation is None and mode == "" and qty is None:
            continue
        raw = str(designation or "").strip()
        if raw.upper() == "OPENING":
            openings += 1
            continue
        if "REMOVED:" in notes:
            removed += 1
        if _num(qty) is None:
            warnings.append(
                f"row {r}: non-numeric qty {qty!r}; row skipped")
            continue
        tons = _resolve_tons(
            ws.cell(row=r, column=14).value,
            ws_data.cell(row=r, column=14).value,
            qty, ws.cell(row=r, column=12).value,
            ws.cell(row=r, column=13).value, r, warnings)
        key = normalize_designation(raw)
        if not key:
            warnings.append(f"row {r}: blank designation; row skipped")
            continue
        e = agg.setdefault(key, {
            "raw": set(), "qty_ea": 0.0, "lf": 0.0, "sf": 0.0,
            "tons": 0.0, "has_qty_ea": False, "has_tons": False})
        e["raw"].add(raw)
        if mode == "COUNT":
            e["qty_ea"] += _num(qty)
            e["has_qty_ea"] = True
        elif mode == "LINEAR":
            e["lf"] += _num(qty)
        elif mode == "AREA":
            e["sf"] += _num(qty)
        if tons is not None:
            e["tons"] += tons
            e["has_tons"] = True
    return {"agg": agg, "warnings": warnings, "openings": openings,
            "removed": removed, "job_stamp": job_stamp}


# -- model side: the Tekla BOM export -----------------------------------------

def _header_map(cells) -> dict:
    """Match one candidate header row against the keyword contract.
    Returns column indices, the matched names, and the weight
    interpretation, or None. A row whose profile cell doubles as the
    qty or weight cell is a mis-split line, not a header."""
    norm = [str(c or "").strip().upper() for c in cells]
    found = {}
    weak_qty = None
    for i, h in enumerate(norm):
        if not h:
            continue
        if "profile" not in found and any(k in h for k in _PROFILE_KEYS):
            found["profile"] = i
        if "qty" not in found and any(k in h for k in _QTY_KEYS):
            found["qty"] = i
        if weak_qty is None and h in _QTY_EXACT:
            weak_qty = i
        if _LENGTH_KEY in h:
            if "length" not in found or "TOTAL" in h:
                found["length"] = i
        if _WEIGHT_KEY in h:
            prev = found.get("weight")
            if prev is None or ("TOTAL" in h
                                and "TOTAL" not in norm[prev]):
                found["weight"] = i
    if "qty" not in found and weak_qty is not None:
        found["qty"] = weak_qty
    if "profile" not in found:
        return None
    if "qty" not in found and "weight" not in found:
        return None
    if found["profile"] in (found.get("qty"), found.get("weight")):
        return None
    out = {"profile": found["profile"], "qty": found.get("qty"),
           "length": found.get("length"), "weight": found.get("weight"),
           "names": {k: norm[v] for k, v in found.items()
                     if v is not None}}
    wh = norm[found["weight"]] if "weight" in found else ""
    out["weight_per_piece"] = ("EACH" in wh or "UNIT" in wh
                               or "PER PIECE" in wh)
    if "KG" in wh:
        out["weight_header_unit"] = "kg"
    elif re.search(r"\(T\)|\bTONNES?\b|\bTONS?\b", wh):
        out["weight_header_unit"] = "ton"
    elif re.search(r"\bLBS?\b|\(LB", wh):
        out["weight_header_unit"] = "lb"
    else:
        out["weight_header_unit"] = None
    return out


_UNIT_SUFFIX = re.compile(r"(?i)\s*(LBS?|KG|TONNES?|TONS?|T)\.?$")
_THOUSANDS = re.compile(r"^\d{1,3}(,\d{3})+(\.\d+)?$")
_EURO_DECIMAL = re.compile(r"^\d+,\d+$")


def _parse_cell(value):
    """(number, unit) from a BOM cell. The unit token comes from a
    trailing suffix: lb, kg, t/ton/tonne. Commas: thousands groups
    drop; a single non-group comma reads as a decimal point."""
    if isinstance(value, bool) or value is None:
        return None, None
    if isinstance(value, (int, float)):
        f = float(value)
        return (f, None) if math.isfinite(f) else (None, None)
    s = str(value).strip()
    unit = None
    m = _UNIT_SUFFIX.search(s)
    if m:
        token = m.group(1).upper().rstrip(".")
        unit = {"LB": "lb", "LBS": "lb", "KG": "kg"}.get(token, "ton")
        s = s[:m.start()].strip()
    if _THOUSANDS.match(s):
        s = s.replace(",", "")
    elif _EURO_DECIMAL.match(s):
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        f = float(s)
    except ValueError:
        return None, None
    return (f, unit) if math.isfinite(f) else (None, None)


def _to_number(value):
    return _parse_cell(value)[0]


def _is_summary(raw: str) -> bool:
    return re.sub(r"[^A-Z]", "", raw.upper()) in _SUMMARY_TOKENS


def _bom_rows(path):
    """Yield raw cell rows from a csv or xlsx BOM export."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            sample = f.read(8192)
            f.seek(0)
            # Sniff on lines that carry a candidate delimiter, so a
            # delimiter-free title line cannot derail the dialect.
            lines = [ln for ln in sample.splitlines()
                     if any(d in ln for d in ",;\t")]
            try:
                dialect = csv.Sniffer().sniff(
                    "\n".join(lines) or sample, ",;\t")
                yield from csv.reader(f, dialect)
            except csv.Error:
                counts = {d: sample.count(d) for d in ",;\t"}
                delim = max(counts, key=counts.get)
                yield from csv.reader(f, delimiter=delim)
    elif path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True,
                                    read_only=True)
        try:
            for ws in wb.worksheets:
                probe = []
                for row in ws.iter_rows(values_only=True):
                    probe.append(list(row))
                if any(_header_map(r) for r in probe[:15]):
                    yield from probe
                    return
            # No sheet carries a detectable header; surface the first
            # sheet so the caller raises the real error message.
            yield from (list(r) for r in
                        wb.worksheets[0].iter_rows(values_only=True))
        finally:
            wb.close()
    else:
        raise ValueError(
            f"unsupported BOM format {path.suffix!r}; csv or xlsx only")


def read_bom(path, weight_unit=None) -> dict:
    """Aggregate the Tekla BOM per normalized designation: piece count
    and pounds. Columns are detected by header keywords; the matched
    names are reported so the data source is explicit. weight_unit
    ('lb', 'kg', 'ton' short) overrides detection; without it a tons
    unit is ambiguous (short vs metric) and dropped with a warning,
    never guessed."""
    cols = None
    agg = {}
    warnings = []
    skipped = 0
    tons_dropped = 0
    for cells in _bom_rows(path):
        if cols is None:
            cols = _header_map(cells)
            continue
        if not any(str(c or "").strip() for c in cells):
            continue
        if _header_map(cells):  # repeated page-break header
            continue
        if cols["profile"] >= len(cells):
            skipped += 1
            continue
        raw = str(cells[cols["profile"]] or "").strip()
        if not raw or _is_summary(raw):
            continue
        qty_cell = (cells[cols["qty"]]
                    if cols["qty"] is not None
                    and cols["qty"] < len(cells) else None)
        qty = _to_number(qty_cell)
        if qty is None and str(qty_cell or "").strip():
            warnings.append(
                f"{raw}: unreadable qty {qty_cell!r}; piece count "
                "dropped for this row")
        wt_cell = (cells[cols["weight"]]
                   if cols["weight"] is not None
                   and cols["weight"] < len(cells) else None)
        weight, cell_unit = _parse_cell(wt_cell)
        if weight is None and str(wt_cell or "").strip():
            warnings.append(
                f"{raw}: unreadable weight {wt_cell!r}; weight "
                "dropped for this row")
        if weight is not None and cols["weight_per_piece"]:
            if qty is None:
                warnings.append(
                    f"{raw}: per-piece weight without a qty; "
                    "weight dropped")
                weight = None
            else:
                weight *= qty
        if weight is not None:
            unit = weight_unit or cell_unit or \
                cols["weight_header_unit"] or "lb"
            if unit == "kg":
                weight *= KG_TO_LB
            elif unit == "ton":
                if weight_unit == "ton":
                    weight *= 2000.0
                else:
                    tons_dropped += 1
                    weight = None
        if qty is None and weight is None:
            skipped += 1
            continue
        key = normalize_designation(raw)
        e = agg.setdefault(key, {"raw": set(), "qty": 0.0,
                                 "weight_lb": 0.0, "has_qty": False,
                                 "has_weight": False})
        e["raw"].add(raw)
        if qty is not None:
            e["qty"] += qty
            e["has_qty"] = True
        if weight is not None:
            e["weight_lb"] += weight
            e["has_weight"] = True
    if cols is None:
        raise ValueError(
            f"no header row found in {Path(path).name}: need a "
            "profile/designation column plus qty/count or weight")
    if tons_dropped:
        warnings.append(
            f"{tons_dropped} weight values carry a tons unit; short "
            "vs metric is ambiguous, so those weights were dropped, "
            "never guessed. Rerun with --weight-unit ton if the "
            "export is short tons.")
    if not agg:
        raise ValueError(
            f"BOM header matched in {Path(path).name} but no usable "
            "data rows followed")
    return {"agg": agg, "columns": cols, "warnings": warnings,
            "skipped": skipped, "weight_unit": weight_unit}


def _weight_note(cols, weight_unit) -> str:
    if cols["weight"] is None:
        return "no weight column detected"
    parts = ["read as per piece and multiplied by qty"
             if cols["weight_per_piece"] else "read as the row total"]
    if weight_unit == "ton":
        parts.append("operator override: short tons x 2000 lb")
    elif weight_unit:
        parts.append(f"operator override: values read as {weight_unit}")
    elif cols["weight_header_unit"] == "kg":
        parts.append("header indicates kg, converted to lb")
    elif cols["weight_header_unit"] == "ton":
        parts.append("header indicates tons, ambiguous without "
                     "--weight-unit; weights dropped, never guessed")
    elif cols["weight_header_unit"] == "lb":
        parts.append("header indicates lb")
    else:
        parts.append("assumed lb")
    return "; ".join(parts)


# -- diff ---------------------------------------------------------------------

def diff(bid, model) -> dict:
    """Per-designation deltas plus the overall tonnage variance.
    Variance sign convention: (bid - model) / model, the model BOM is
    the reference. Like for like: a MATCHED designation weighed on one
    side only (COUNT-mode bid rows carry no derived tonnage by schema
    design) is excluded from both totals and disclosed; its qty delta
    still stands. One-sided designations stay in the totals; steel
    only one side knows about IS the miss."""
    rows = []
    excl_bid = []
    excl_model = []
    bid_total = 0.0
    model_total = 0.0
    bid_counted = False
    bid_has_tons = False
    model_has_tons = False
    for key in sorted(set(bid) | set(model)):
        b = bid.get(key)
        m = model.get(key)
        bid_qty = b["qty_ea"] if b and b["has_qty_ea"] else None
        model_qty = m["qty"] if m and m["has_qty"] else None
        bid_tons = b["tons"] if b and b["has_tons"] else None
        model_tons = (m["weight_lb"] / 2000.0
                      if m and m["has_weight"] else None)
        status = ("MATCHED" if b and m
                  else "BID_ONLY" if b else "MODEL_ONLY")
        bid_has_tons = bid_has_tons or bid_tons is not None
        model_has_tons = model_has_tons or model_tons is not None
        if status == "MATCHED" and \
                (bid_tons is None) != (model_tons is None):
            tons_delta = None
            if bid_tons is not None:
                excl_bid.append((key, bid_tons))
            else:
                excl_model.append((key, model_tons))
        else:
            if bid_tons is not None:
                bid_total += bid_tons
                bid_counted = True
            if model_tons is not None:
                model_total += model_tons
            tons_delta = ((bid_tons or 0.0) - (model_tons or 0.0)
                          if bid_tons is not None
                          or model_tons is not None else None)
        rows.append({
            "key": key,
            "raw_bid": sorted(b["raw"]) if b else [],
            "raw_model": sorted(m["raw"]) if m else [],
            "bid_qty": bid_qty, "model_qty": model_qty,
            "qty_delta": (bid_qty - model_qty
                          if bid_qty is not None
                          and model_qty is not None else None),
            "bid_tons": bid_tons, "model_tons": model_tons,
            "tons_delta": tons_delta,
            "status": status,
        })
    variance_pct = None
    if bid_counted and model_total > 0:
        variance_pct = (bid_total - model_total) / model_total * 100.0
    # Rank misses by tonnage only when a real tonnage comparison
    # happened; otherwise piece counts are the honest ranking.
    misses = []
    rank_basis = "qty"
    if variance_pct is not None:
        misses = sorted(
            (r for r in rows if r["tons_delta"] is not None
             and abs(r["tons_delta"]) > 1e-9),
            key=lambda r: abs(r["tons_delta"]),
            reverse=True)[:TOP_MISSES]
        if misses:
            rank_basis = "tons"
    if not misses:
        misses = sorted(
            (r for r in rows if r["qty_delta"] is not None
             and abs(r["qty_delta"]) > 1e-9),
            key=lambda r: abs(r["qty_delta"]),
            reverse=True)[:TOP_MISSES]
    return {"rows": rows, "bid_tons": bid_total,
            "model_tons": model_total, "variance_pct": variance_pct,
            "bid_has_tons": bid_has_tons,
            "model_has_tons": model_has_tons,
            "excl_bid": excl_bid, "excl_model": excl_model,
            "excl_bid_tons": sum(t for _, t in excl_bid),
            "excl_model_tons": sum(t for _, t in excl_model),
            "misses": misses, "rank_basis": rank_basis}


def verdict_line(d) -> str:
    """One line against the Section 07 target: tonnage within 2 to 3
    percent of the Tekla shop-model BOM. Classifies on the same
    2-decimal value it displays and the ledger stores."""
    v = d["variance_pct"]
    if v is None:
        if not d["bid_has_tons"]:
            reason = ("the bid takeoff carries no derived tonnage "
                      "(no LINEAR rows with lb_per_ft filled)")
        elif not d["model_has_tons"]:
            reason = "the BOM export carries no weight column data"
        elif d.get("excl_bid") or d.get("excl_model"):
            reason = ("after excluding matched designations weighed "
                      "on one side only, no comparable tonnage "
                      "remains; see the method notes")
        else:
            reason = "the model BOM tonnage totals zero"
        return ("VERDICT: NOT SCORED against the Section 07 target. "
                f"Tonnage variance cannot be computed: {reason}.")
    v = round(v, 2)
    a = abs(v)
    if a <= TARGET_INNER_PCT:
        return (f"VERDICT: PASS. Overall tonnage variance {v:+.2f} "
                "percent is inside the tight 2 percent band of the "
                "Section 07 target (within 2 to 3 percent of the "
                "Tekla shop-model BOM).")
    if a <= TARGET_OUTER_PCT:
        return (f"VERDICT: PASS. Overall tonnage variance {v:+.2f} "
                "percent is within the Section 07 target (within 2 "
                "to 3 percent of the Tekla shop-model BOM); the "
                "tighter 2 percent band was missed.")
    return (f"VERDICT: FAIL. Overall tonnage variance {v:+.2f} "
            "percent is outside the Section 07 target (within 2 to "
            "3 percent of the Tekla shop-model BOM).")


# -- report and ledger ---------------------------------------------------------

def _fmt(value, places=2) -> str:
    if value is None:
        return "-"
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return f"{value:.{places}f}"


def _raw_forms(row) -> str:
    forms = sorted(set(row["raw_bid"]) | set(row["raw_model"]))
    return "; ".join(forms[:3]) + (" ..." if len(forms) > 3 else "")


def _key_list(pairs, cap=10) -> str:
    keys = [k for k, _ in pairs[:cap]]
    return ", ".join(keys) + (" ..." if len(pairs) > cap else "")


def build_report(job, d, bid_meta, bom_meta, takeoff_path,
                 bom_path) -> str:
    """The reconcile_<job>.md body. Findings propose nothing."""
    cols = bom_meta["columns"]
    lines = [
        f"# Tekla reconciliation: {job}",
        "",
        "-GP CONFIDENTIAL. Internal accuracy audit (T13/F5). Never "
        "client-facing.",
        "",
        "This is a free accuracy audit; the model gets built "
        "regardless. Findings propose nothing. Owner and Ivan read "
        "this report.",
        "",
        f"- Date: {date.today().isoformat()}",
        f"- Bid takeoff: {Path(takeoff_path).name}",
        f"- Tekla BOM: {Path(bom_path).name}",
        "- BOM columns matched: " + ", ".join(
            f"{k}={v!r}" for k, v in sorted(cols["names"].items())),
        "- Weight interpretation: "
        + _weight_note(cols, bom_meta["weight_unit"]),
        "",
        "## Overall tonnage",
        "",
        "| side | tons compared | tons excluded |",
        "|---|---|---|",
        f"| bid takeoff (derived columns) | {_fmt(d['bid_tons'])} | "
        f"{_fmt(d['excl_bid_tons'])} |",
        f"| Tekla shop-model BOM | {_fmt(d['model_tons'])} | "
        f"{_fmt(d['excl_model_tons'])} |",
        "",
        ("Overall tonnage variance: "
         + (f"{d['variance_pct']:+.2f} percent (bid vs model; the "
            "model BOM is the reference)"
            if d["variance_pct"] is not None else "not computable")),
        "",
        verdict_line(d),
        "",
        f"## Top {TOP_MISSES} absolute misses "
        f"(ranked by {d['rank_basis']} delta)",
        "",
    ]
    if d["misses"]:
        lines += [
            "| designation | raw forms | bid qty EA | model qty | "
            "qty delta | bid tons | model tons | tons delta | status |",
            "|---|---|---|---|---|---|---|---|---|",
        ]
        for r in d["misses"]:
            lines.append(
                f"| {r['key']} | {_raw_forms(r)} | "
                f"{_fmt(r['bid_qty'])} | {_fmt(r['model_qty'])} | "
                f"{_fmt(r['qty_delta'])} | {_fmt(r['bid_tons'])} | "
                f"{_fmt(r['model_tons'])} | {_fmt(r['tons_delta'])} | "
                f"{r['status']} |")
    else:
        lines.append("No comparable per-designation deltas. The "
                     "coverage section below shows what each side "
                     "carries.")
    bid_only = [r for r in d["rows"] if r["status"] == "BID_ONLY"]
    model_only = [r for r in d["rows"] if r["status"] == "MODEL_ONLY"]
    matched = [r for r in d["rows"] if r["status"] == "MATCHED"]
    lines += [
        "",
        "## Coverage",
        "",
        f"- Designations matched on both sides: {len(matched)}",
        f"- Bid only ({len(bid_only)}): "
        + (", ".join(r["key"] for r in bid_only) or "none"),
        f"- Model only ({len(model_only)}): "
        + (", ".join(r["key"] for r in model_only) or "none"),
        "",
        "## Method notes",
        "",
        "- Bid tonnage comes from the takeoff file's derived columns "
        "only: the exporter's =D*L and =M/2000 formula cells evaluated "
        "from each row's own qty and lb_per_ft values. lb_per_ft was "
        "filled by the operator from bridge/aisc_validator.py. Nothing "
        "here recomputes a weight.",
        "- Designations were normalized on both sides for matching "
        "only (spaces and inch marks stripped, TS and FL prefixes, "
        "decimal dimensions to fractions, joist chord leading zeros). "
        "Raw tags are shown beside each key.",
        "- Model tons are short tons, BOM weight divided by 2000.",
    ]
    if d["excl_model"]:
        lines.append(
            f"- {_fmt(d['excl_model_tons'])} tons of model BOM weight "
            "sit on matched designations with no bid-side derived "
            f"tonnage ({_key_list(d['excl_model'])}). The takeoff "
            "carries derived tonnage on LINEAR rows only, so these "
            "are excluded from the variance; their piece-count "
            "deltas still stand.")
    if d["excl_bid"]:
        lines.append(
            f"- {_fmt(d['excl_bid_tons'])} tons of bid takeoff "
            "tonnage sit on matched designations with no model-side "
            f"weight ({_key_list(d['excl_bid'])}); excluded from the "
            "variance, piece-count deltas still stand.")
    if bid_meta["openings"]:
        lines.append(
            f"- {bid_meta['openings']} deck OPENING deduction rows "
            "were excluded; they are not members.")
    if bid_meta["removed"]:
        lines.append(
            f"- {bid_meta['removed']} bid rows carry a REMOVED: "
            "revision token; their qty already reflects removal.")
    if bom_meta["skipped"]:
        lines.append(
            f"- {bom_meta['skipped']} BOM rows had no readable "
            "profile, qty, or weight and were skipped.")
    warnings = bid_meta["warnings"] + bom_meta["warnings"]
    if warnings:
        lines += ["", "## Warnings", ""]
        lines += [f"- {w}" for w in warnings]
    lines.append("")
    return "\n".join(lines)


def _append_ledger(rows, ledger_dir) -> None:
    ledger_dir = Path(ledger_dir)
    ledger_dir.mkdir(parents=True, exist_ok=True)
    ledger = ledger_dir / "accuracy_ledger.csv"
    new_file = not ledger.exists()
    with open(ledger, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(LEDGER_HEADER)
        w.writerows(rows)


def ledger_rows(job, d) -> list:
    """Section 07 ledger: T13 diffs, one metric per row, numbers as
    computed."""
    today = date.today().isoformat()
    compared_note = (" (compared rows only)"
                     if d["excl_bid"] or d["excl_model"] else "")
    rows = [
        [today, job, TEST_NAME, "bid_tons",
         round(d["bid_tons"], 3),
         "takeoff derived columns" + compared_note],
        [today, job, TEST_NAME, "model_tons",
         round(d["model_tons"], 3),
         "Tekla shop-model BOM / 2000" + compared_note],
    ]
    if d["variance_pct"] is not None:
        rows.append(
            [today, job, TEST_NAME, "tonnage_variance_pct",
             round(d["variance_pct"], 2),
             "Section 07 target within 2 to 3 pct of the model BOM"])
    if d["excl_model_tons"]:
        rows.append(
            [today, job, TEST_NAME, "model_tons_excluded",
             round(d["excl_model_tons"], 3),
             "matched designations with no bid-side derived tonnage: "
             + _key_list(d["excl_model"], cap=3)])
    if d["excl_bid_tons"]:
        rows.append(
            [today, job, TEST_NAME, "bid_tons_excluded",
             round(d["excl_bid_tons"], 3),
             "matched designations with no model-side weight: "
             + _key_list(d["excl_bid"], cap=3)])
    matched = sum(1 for r in d["rows"] if r["status"] == "MATCHED")
    bid_only = sum(1 for r in d["rows"] if r["status"] == "BID_ONLY")
    model_only = sum(1 for r in d["rows"]
                     if r["status"] == "MODEL_ONLY")
    rows.append(
        [today, job, TEST_NAME, "designations_matched", matched,
         f"bid_only {bid_only}, model_only {model_only}"])
    if d["misses"]:
        top = d["misses"][0]
        basis = ("tons_delta" if d["rank_basis"] == "tons"
                 else "qty_delta")
        rows.append(
            [today, job, TEST_NAME, f"top_miss_abs_{d['rank_basis']}",
             round(abs(top[basis]), 3 if basis == "tons_delta" else 1),
             f"worst single designation: {top['key']}"])
    return rows


# -- entry points ---------------------------------------------------------------

def _job_from_inputs(job, takeoff_path, stamp) -> str:
    if job:
        return job
    if stamp:
        return str(stamp)
    name = Path(takeoff_path).stem
    return name.split("_TAKEOFF")[0] if "_TAKEOFF" in name else name


def run(takeoff_path, bom_path, job=None, ledger_dir=None,
        weight_unit=None) -> dict:
    """Full reconcile: read both sides, diff, write the report, append
    the ledger rows. Returns the diff dict plus output paths."""
    bid = read_takeoff(takeoff_path)
    bom = read_bom(bom_path, weight_unit=weight_unit)
    job = _job_from_inputs(job, takeoff_path, bid["job_stamp"])
    job_token = re.sub(r"[^A-Za-z0-9_-]+", "_", job)
    d = diff(bid["agg"], bom["agg"])
    report = build_report(job, d, bid, bom, takeoff_path, bom_path)
    out_dir = Path(ledger_dir) if ledger_dir else LEDGER_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / f"reconcile_{job_token}.md"
    report_path.write_text(report, encoding="utf-8")
    rows = ledger_rows(job, d)
    _append_ledger(rows, out_dir)
    d["job"] = job
    d["report_path"] = report_path
    d["ledger_rows"] = rows
    d["verdict"] = verdict_line(d)
    return d


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        description="Bid takeoff vs Tekla shop-model BOM diff "
                    "(T13/F5, Section 07 ledger). Free accuracy "
                    "audit; findings propose nothing.")
    p.add_argument("takeoff", help="bid takeoff xlsx, T8 format")
    p.add_argument("bom", help="Tekla BOM export, csv or xlsx")
    p.add_argument("--job", help="job id; default from the stamp or "
                                 "the takeoff filename")
    p.add_argument("--ledger-dir",
                   help="report and ledger directory; default "
                        "takeoff_pipeline/ledger")
    p.add_argument("--weight-unit", choices=("lb", "kg", "ton"),
                   help="override BOM weight unit detection; ton "
                        "means short tons x 2000 lb")
    args = p.parse_args(argv)
    try:
        d = run(args.takeoff, args.bom, job=args.job,
                ledger_dir=args.ledger_dir,
                weight_unit=args.weight_unit)
    except (ValueError, FileNotFoundError, BadZipFile) as exc:
        print(f"reconcile failed: {exc}")
        return 2
    print(f"report: {d['report_path']}")
    print(d["verdict"])
    print(f"ledger: {len(d['ledger_rows'])} rows appended "
          f"({TEST_NAME})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
