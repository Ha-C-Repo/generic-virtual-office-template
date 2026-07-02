"""Exporter for the four-sheet takeoff workbook (TAKEOFF_SCHEMA_V2).

Builds TAKEOFF, BOQ, BOM, and PRICING_SCHEDULE from census.db plus
manual-entry rows, per takeoff_pipeline/docs/TAKEOFF_SCHEMA_V2.md.
Where this code and that document disagree, the document wins.

Pipeline order (13.1): build unstamped, operator fills lb_per_ft from
bridge/aisc_validator.py output, validate_takeoff.py gates, THEN the
stamp is written and the file takes its final name
<job>_TAKEOFF_R<n>_<hash12>.xlsx. main() runs build, validate, stamp.

Derived weights are FORMULA cells only (=qty*lb_per_ft, =weight/2000).
The exporter never computes a weight itself and never calls an LLM for
one; lb_per_ft is an input column the operator fills from
bridge/aisc_validator.py output (Hard Rule 5). The PRICING_SCHEDULE
sheet is structure only: line organization and item references, no
pricing columns, no typed numbers (its qty cells are formulas summing
TAKEOFF cells).

Census rows are COUNT/EA: a text census counts callouts and schedule
rows, it does not measure lengths or areas. Schedule-backed quantities
without a schedule QTY column carry QTY_BASIS: min (the row count is a
floor); plan-callout counts carry QTY_BASIS: approx. LINEAR and AREA
quantities enter as manual-entry rows (section 7): a csv with columns
item_class, designation, mode, qty, unit, primary_source,
secondary_source, confidence, sheet, bbox, notes. confidence is set by
the operator, never defaulted (a blank is a build error). Attribute
evidence in the census (camber text, printed SF figures) never becomes
a TAKEOFF row; attributes ride in notes per section 5.

Imperial units only. No pricing anywhere (P25). Free tooling
(openpyxl). Paths are package-relative; resource_path() only if this
is ever promoted into bridge/.
"""

import csv
import json
import re
import sqlite3
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from takeoff_pipeline.takeoff_hash import compute_takeoff_hash
from takeoff_pipeline.validate_takeoff import SOURCE_KEYWORDS

_PKG = Path(__file__).resolve().parent
DEFAULT_DB = _PKG / "census.db"
DEFAULT_OUT_DIR = _PKG / "exports"

SCHEMA_LITERAL = "TAKEOFF_SCHEMA_V2"

TAKEOFF_HEADERS = (
    "item_id", "designation", "mode", "qty", "unit", "primary_source",
    "secondary_source", "confidence", "sheet", "bbox", "notes",
    "lb_per_ft", "weight_lb", "tons", "formula_ref",
)
BOQ_HEADERS = ("boq_id", "source_item_id", "assembly_id", "stream",
               "description", "qty", "unit", "formula_ref", "notes")
BOM_HEADERS = ("bom_id", "source_item_id", "designation", "description",
               "qty", "unit", "length_lf", "lb_per_ft", "weight_lb",
               "tons", "formula_ref", "notes")
PS_HEADERS = ("line_id", "description", "source_refs", "qty", "unit")

CLASS_ORDER = ("COL", "BEAM", "JST", "DECK", "PLATE", "ANCH", "MISC")
MODE_UNIT = {"COUNT": "EA", "LINEAR": "LF", "AREA": "SF"}
COUNT_CLASSES = ("COL", "JST", "ANCH", "PLATE")

# Census-derived rows are always COUNT: a text-and-schedule census
# counts callouts and schedule rows, it never measures a length or an
# area. LINEAR rows (member sticks) and AREA rows (deck and footprint)
# enter as manual-entry rows (section 7) or from grid_geometry.py, each
# carrying its own mode. This mapping makes the per-class census mode
# explicit so it is not a hard-coded literal on every row; it is the one
# place to change if the census ever derives a non-count quantity.
CENSUS_MODE_BY_CLASS = {cls: "COUNT" for cls in CLASS_ORDER}

# Census evidence that is an attribute, never a member row (section 5:
# attributes ride in notes; they never form COUNT/LINEAR/AREA rows).
_ATTRIBUTE_EVIDENCE = re.compile(
    r"CAMBER|^\d[\d,]*\s?SF$", re.IGNORECASE)


# -- census aggregation ----------------------------------------------------

TYPE_ONLY_SUPPRESS_CLASSES = ("COL", "PLATE")


def _is_type_only_schedule_group(base_class, sched, plan) -> bool:
    """A COL or PLATE designation whose only quantity evidence is a
    schedule that names TYPES without a quantity column (A1). A base-
    plate or column-size schedule lists types, never a member count, so
    it produces no count row: the column count is read from the
    foundation plan, the base-plate count is derived one per verified
    column downstream (P29), and Ivan verifies both. Counts come only
    from an explicit QTY column; a row-count of types is never inferred.

    For PLATE a plan callout IS a real count and blocks suppression. For
    COL there are no plan-class hits to weigh (a plan HSS is BEAM, not
    COL), so a column type schedule suppresses regardless."""
    if not sched or any(h["qty"] is not None for h in sched):
        return False
    if base_class == "COL":
        return True
    if base_class == "PLATE":
        return not plan
    return False


def _uncounted_types(job, db_path, item_class, require_no_plan) -> list:
    """TYPES of a class present only on a no-quantity schedule. Surfaced
    in the build report so the absence of a count is loud, never silent;
    the overlay draws the underlying census hits for Ivan."""
    path = Path(db_path) if db_path else DEFAULT_DB
    c = sqlite3.connect(str(path), check_same_thread=False, timeout=10)
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA busy_timeout=10000")
    c.row_factory = sqlite3.Row
    try:
        hits = [dict(r) for r in c.execute(
            "SELECT * FROM census_hits WHERE job = ? AND item_class = ?"
            " ORDER BY id", (job, item_class))]
    finally:
        c.close()

    def norm(d):
        return re.sub(r"[\s\"']", "", (d or "").upper())

    groups = {}
    for h in hits:
        groups.setdefault(norm(h["designation"]), []).append(h)
    out = []
    for key in sorted(groups):
        group = groups[key]
        sched = [h for h in group if h["source_kind"] == "SCHEDULE"]
        plan = [h for h in group if h["source_kind"] == "PLAN"]
        if not sched or any(h["qty"] is not None for h in sched):
            continue
        if require_no_plan and plan:
            continue
        out.append(f"{group[0]['designation']} ({sched[0]['primary_source']})")
    return out


def uncounted_col_types(job, db_path=None) -> list:
    """Column SIZES on a schedule with no quantity column. The member
    count is read from the foundation plan, never from these types."""
    return _uncounted_types(job, db_path, "COL", require_no_plan=False)


def uncounted_plate_types(job, db_path=None) -> list:
    """Base-plate SIZES on a schedule with no quantity column. The base-
    plate count is derived one per verified column downstream (P29), so
    a type schedule yields an honest null pending the column count,
    never a wrong per-type total. Plan callout plates are counted
    normally and are not listed here."""
    return _uncounted_types(job, db_path, "PLATE", require_no_plan=True)


def _census_rows(job, db_path=None) -> list:
    """Aggregate census hits into TAKEOFF row dicts per Appendix A:
    one row per (designation, mode, primary_source). Schedule is
    primary for count classes, plan for BEAM; the cross-check side is
    cited in secondary_source. Two schedule sources that disagree on
    one designation become a CONFLICT row (section 8), never a silent
    sum. All ordering is deterministic so anyone rebuilding from the
    same census reproduces the same takeoff (section 1)."""
    path = Path(db_path) if db_path else DEFAULT_DB
    c = sqlite3.connect(str(path), check_same_thread=False, timeout=10)
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA busy_timeout=10000")
    c.row_factory = sqlite3.Row
    try:
        hits = [dict(r) for r in c.execute(
            "SELECT * FROM census_hits WHERE job = ? ORDER BY id",
            (job,))]
        conflicts = {r["conflict_group"]: dict(r) for r in c.execute(
            "SELECT * FROM conflicts WHERE job = ?", (job,))}
    finally:
        c.close()

    def norm(d):
        return re.sub(r"[\s\"']", "", (d or "").upper())

    groups = {}
    for h in hits:
        if _ATTRIBUTE_EVIDENCE.search(h["designation"].strip()):
            continue
        groups.setdefault(norm(h["designation"]), []).append(h)

    rows = []
    for key in sorted(groups):
        group = sorted(groups[key], key=_reading_order)
        sched = [h for h in group if h["source_kind"] == "SCHEDULE"]
        plan = [h for h in group if h["source_kind"] == "PLAN"]
        base_class = sched[0]["item_class"] if sched \
            else group[0]["item_class"]
        # A1: a COL or PLATE SIZE listed on a base-plate or column
        # schedule with no quantity column is a TYPE, not a member
        # count. It produces NO TAKEOFF count row: a row-count of types
        # would read as the building total (4 types, 16 columns). The
        # column count comes from the foundation plan and the base-plate
        # count is derived one per column (P29); both verified by Ivan.
        # uncounted_col_types() / uncounted_plate_types() surface the
        # types and the overlay still draws every census hit for him.
        if _is_type_only_schedule_group(base_class, sched, plan):
            continue
        schedule_primary = bool(sched) and (
            base_class in COUNT_CLASSES or not plan)

        conflict_notes = []
        for ck in sorted({h["conflict_group"] for h in group
                          if h["conflict_group"]}):
            if ck in conflicts:
                conflict_notes.append(conflicts[ck]["note"])

        if schedule_primary:
            rows.extend(_schedule_primary_rows(
                sched, plan, base_class, conflict_notes))
        else:
            primary, cross = (plan, sched) if plan else (sched, [])
            # One row per primary source (Appendix A). A designation
            # called out on two framing plans is two rows, each citing
            # its own sheet.
            for src in _source_order(primary):
                side = [h for h in primary if h["primary_source"] == src]
                rows.append(_make_row(
                    side, base_class, float(len(side)),
                    ["QTY_BASIS: approx"], cross, conflict_notes))
    return rows


def _schedule_primary_rows(sched, plan, base_class, conflict_notes):
    """Rows for a schedule-primary designation. One schedule source:
    its quantity stands (rows within one table are additive; the same
    designation on two marks is two real quantities). Multiple
    schedule sources that AGREE collapse to one row with a note; ones
    that DISAGREE become a CONFLICT row per section 8, never a sum."""
    by_src = {}
    for src in _source_order(sched):
        side = [h for h in sched if h["primary_source"] == src]
        qtys = [h["qty"] for h in side if h["qty"] is not None]
        by_src[src] = {
            "hits": side,
            "qty": sum(qtys) if qtys else float(len(side)),
            "counted": not qtys,
        }
    srcs = list(by_src)
    values = {round(v["qty"], 6) for v in by_src.values()}
    first = by_src[srcs[0]]

    notes = list(conflict_notes)
    if len(srcs) == 1:
        if first["counted"]:
            notes.append("QTY_BASIS: min")
        return [_make_row(first["hits"], base_class, first["qty"],
                          notes, plan, conflict_notes)]
    if len(values) == 1:
        all_hits = [h for v in by_src.values() for h in v["hits"]]
        notes.append(f"schedule appears on {len(srcs)} sources with "
                     "the same quantity")
        if first["counted"]:
            notes.append("QTY_BASIS: min")
        return [_make_row(all_hits, base_class, first["qty"], notes,
                          plan, conflict_notes)]
    # Two-source disagreement (section 5/8): CONFLICT, never a sum.
    detail = " vs ".join(
        f"{src} {by_src[src]['qty']:g} EA" for src in srcs)
    conflict = (f"CONFLICT: {detail}. RFI candidate. Never resolved "
                "silently.")
    all_hits = [h for v in by_src.values() for h in v["hits"]]
    row = _make_row(all_hits, base_class, first["qty"],
                    [conflict] + notes, plan, [conflict])
    row["confidence"] = "low"
    return [row]


def _make_row(primary_hits, item_class, qty, notes, cross,
              conflict_notes) -> dict:
    """One TAKEOFF row from one primary side. Reclassification per
    Appendix A and section 5 happens here, per row, from the row's own
    cited source."""
    primary_hits = sorted(primary_hits, key=_reading_order)
    primary_src = _conforming_source(primary_hits, item_class)

    notes = list(notes)
    # Appendix A maps shape families to BEAM "from framing plans" and
    # deck callouts to DECK on framing plans; the same evidence on a
    # detail sheet or a non-column schedule is unresolved, so MISC.
    # JST stays JST even plan-only: a joist count without its joist
    # schedule is meant to fail the gate until the schedule exists.
    if item_class in ("BEAM", "DECK") \
            and "FRAMING PLAN" not in primary_src.upper():
        item_class = "MISC"
        notes.append("classed MISC per Appendix A: shape evidence "
                     "outside a framing plan")
    elif item_class in ("ANCH", "PLATE") and not any(
            k in primary_src.upper()
            for k in SOURCE_KEYWORDS[item_class]):
        item_class = "MISC"
        notes.append("classed MISC per section 5 loose-plate pattern: "
                     "no schedule or detail source in the census "
                     "evidence")

    rank = {"low": 0, "medium": 1, "high": 2}
    confidence = min((h["confidence"] for h in primary_hits),
                     key=lambda v: rank[v])
    if conflict_notes:
        confidence = "low"
    # Section 8: the CONFLICT token leads the notes.
    notes.sort(key=lambda n: 0 if n.startswith("CONFLICT:") else 1)

    # Deterministic display designation: the most frequent raw variant,
    # ties broken lexicographically.
    variants = {}
    for h in primary_hits:
        variants[h["designation"]] = variants.get(h["designation"],
                                                  0) + 1
    designation = sorted(variants.items(),
                         key=lambda kv: (-kv[1], kv[0]))[0][0]

    anchor = primary_hits[0]
    cross_srcs = _source_order(cross) if cross else []
    return {
        "item_class": item_class,
        "designation": designation,
        "mode": CENSUS_MODE_BY_CLASS.get(item_class, "COUNT"),
        "qty": qty,
        "unit": "EA",
        "primary_source": primary_src,
        "secondary_source": cross_srcs[0] if cross_srcs else "",
        "confidence": confidence,
        "sheet": anchor["sheet"],
        "bbox": anchor["bbox"],
        "notes": " ".join(notes),
    }


def _source_order(hits) -> list:
    """Distinct primary_source strings in reading order of their first
    hit. Deterministic."""
    seen = []
    for h in sorted(hits, key=_reading_order):
        if h["primary_source"] not in seen:
            seen.append(h["primary_source"])
    return seen


def _conforming_source(hits_side, item_class) -> str:
    """First contributing source that satisfies the section 5 keyword
    rule for the class; falls back to the first hit's source."""
    keywords = SOURCE_KEYWORDS.get(item_class, ())
    if keywords:
        for h in hits_side:
            if any(k in h["primary_source"].upper() for k in keywords):
                return h["primary_source"]
    return hits_side[0]["primary_source"]


def _sheet_key(sheet: str):
    """Natural sort: S2.1 sorts before S10.1 (Appendix A reading order
    says lowest sheet NUMBER, not lowest string)."""
    return tuple(int(p) if p.isdigit() else p
                 for p in re.split(r"(\d+)", sheet or "") if p)


def _reading_order(hit):
    """Lowest sheet number first, then top to bottom, left to right."""
    try:
        bbox = json.loads(hit["bbox"])
        x0, y0 = float(bbox[0]), float(bbox[1])
    except (ValueError, TypeError, IndexError):
        x0 = y0 = 0.0
    return (_sheet_key(hit["sheet"]), y0, x0)


# -- manual rows -------------------------------------------------------------

MANUAL_COLUMNS = ("item_class", "designation", "mode", "qty", "unit",
                  "primary_source", "secondary_source", "confidence",
                  "sheet", "bbox", "notes")


def load_manual_rows(csv_path) -> list:
    """Manual-entry rows per section 7. confidence is never defaulted;
    a blank is a build error, not a silent high."""
    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        missing = [c for c in MANUAL_COLUMNS
                   if c not in (reader.fieldnames or [])]
        if missing:
            raise ValueError(
                f"manual csv missing columns: {', '.join(missing)}")
        for i, raw in enumerate(reader, start=2):
            mode = (raw["mode"] or "").strip().upper()
            if mode not in MODE_UNIT:
                raise ValueError(
                    f"manual csv line {i}: mode '{raw['mode']}' is not "
                    "COUNT, LINEAR, or AREA")
            unit = (raw["unit"] or "").strip().upper() or MODE_UNIT[mode]
            if unit != MODE_UNIT[mode]:
                raise ValueError(
                    f"manual csv line {i}: unit {unit} does not match "
                    f"mode {mode} ({MODE_UNIT[mode]})")
            confidence = (raw["confidence"] or "").strip().lower()
            if confidence not in ("high", "medium", "low"):
                raise ValueError(
                    f"manual csv line {i}: confidence must be high, "
                    "medium, or low; it is set by the operator, never "
                    "defaulted (section 7)")
            item_class = (raw["item_class"] or "").strip().upper()
            if item_class not in CLASS_ORDER:
                raise ValueError(
                    f"manual csv line {i}: item_class '{item_class}' "
                    f"is not one of {', '.join(CLASS_ORDER)}")
            try:
                qty = float(raw["qty"])
            except (TypeError, ValueError):
                raise ValueError(
                    f"manual csv line {i}: qty '{raw['qty']}' is not "
                    "a number") from None
            rows.append({
                "item_class": item_class,
                "designation": (raw["designation"] or "").strip(),
                "mode": mode,
                "qty": qty,
                "unit": unit,
                "primary_source": (raw["primary_source"] or "").strip(),
                "secondary_source": (raw["secondary_source"] or "")
                .strip(),
                "confidence": confidence,
                "sheet": (raw["sheet"] or "").strip() or "MANUAL",
                "bbox": (raw["bbox"] or "").strip() or "MANUAL",
                "notes": (raw["notes"] or "").strip(),
            })
    return rows


def _normalize_extra_row(raw, index) -> dict:
    """Validate and canonicalize a pre-built TAKEOFF row dict (such as
    a grid_geometry.py AREA row) into the eleven 3.2 columns. Same
    gates as load_manual_rows: a bad mode, unit, confidence, class, or
    qty fails loud, never a silent invalid row."""
    missing = [c for c in MANUAL_COLUMNS if c not in raw]
    if missing:
        raise ValueError(
            f"extra row {index} missing keys: {', '.join(missing)}")
    mode = (raw["mode"] or "").strip().upper()
    if mode not in MODE_UNIT:
        raise ValueError(
            f"extra row {index}: mode '{raw['mode']}' is not COUNT, "
            "LINEAR, or AREA")
    unit = (raw["unit"] or "").strip().upper() or MODE_UNIT[mode]
    if unit != MODE_UNIT[mode]:
        raise ValueError(
            f"extra row {index}: unit {unit} does not match mode "
            f"{mode} ({MODE_UNIT[mode]})")
    confidence = (raw["confidence"] or "").strip().lower()
    if confidence not in ("high", "medium", "low"):
        raise ValueError(
            f"extra row {index}: confidence must be high, medium, or "
            "low (set at source, never defaulted)")
    item_class = (raw["item_class"] or "").strip().upper()
    if item_class not in CLASS_ORDER:
        raise ValueError(
            f"extra row {index}: item_class '{item_class}' is not one "
            f"of {', '.join(CLASS_ORDER)}")
    try:
        qty = float(raw["qty"])
    except (TypeError, ValueError):
        raise ValueError(
            f"extra row {index}: qty '{raw['qty']}' is not a number"
        ) from None
    return {
        "item_class": item_class,
        "designation": (raw["designation"] or "").strip(),
        "mode": mode,
        "qty": qty,
        "unit": unit,
        "primary_source": (raw["primary_source"] or "").strip(),
        "secondary_source": (raw["secondary_source"] or "").strip(),
        "confidence": confidence,
        "sheet": (raw["sheet"] or "").strip() or "MANUAL",
        "bbox": (raw["bbox"] or "").strip() or "MANUAL",
        "notes": (raw["notes"] or "").strip(),
    }


# -- workbook build ----------------------------------------------------------

def _formula_ref_for(item_class, designation) -> str:
    if item_class == "JST":
        return "SJI:joist load table"
    return f"AISC:bridge/aisc_validator.py:{designation}"


def build_workbook(job, db_path=None, manual_csv=None, extra_rows=None):
    """The unstamped four-sheet workbook. Returns (workbook, info).

    extra_rows are pre-built TAKEOFF row dicts in the manual-row shape
    (the eleven 3.2 columns as keys), merged exactly like manual rows.
    grid_geometry.py feeds its AREA rows here so footprint and deck SF
    are first-class takeoff rows; each carries its own mode (AREA), so
    it never passes through the census COUNT path."""
    import openpyxl

    census_rows = _census_rows(job, db_path)
    manual_rows = load_manual_rows(manual_csv) if manual_csv else []
    grid_rows = [_normalize_extra_row(r, i)
                 for i, r in enumerate(extra_rows or [], start=1)]
    rows = census_rows + manual_rows + grid_rows

    # Sort by class order, then mode, then designation: deterministic,
    # and it makes each PRICING_SCHEDULE partition a CONTIGUOUS block
    # so qty formulas are single ranges, not one SUM argument per row
    # (Excel caps a function at 255 arguments).
    rows.sort(key=lambda r: (CLASS_ORDER.index(r["item_class"]),
                             r["mode"], r["designation"],
                             r["primary_source"]))

    counters = {cls: 0 for cls in CLASS_ORDER}
    for row in rows:
        counters[row["item_class"]] += 1
        row["item_id"] = f"{row['item_class']}-" \
                         f"{counters[row['item_class']]:03d}"

    wb = openpyxl.Workbook()
    takeoff = wb.active
    takeoff.title = "TAKEOFF"

    # Row 1 is the metadata row, written only by the stamp step (13.1:
    # stamping is the last write). Row 2 holds the fifteen headers.
    takeoff.append([])
    takeoff.append(list(TAKEOFF_HEADERS))
    for r_idx, row in enumerate(rows, start=3):
        takeoff.append([
            row["item_id"], row["designation"], row["mode"], row["qty"],
            row["unit"], row["primary_source"], row["secondary_source"],
            row["confidence"], row["sheet"], row["bbox"], row["notes"],
        ])
        if row["mode"] == "LINEAR":
            # lb_per_ft (column L) stays empty for the operator. The
            # weight cells are formulas, never typed constants.
            takeoff.cell(row=r_idx, column=13,
                         value=f"=D{r_idx}*L{r_idx}")
            takeoff.cell(row=r_idx, column=14, value=f"=M{r_idx}/2000")
            takeoff.cell(row=r_idx, column=15,
                         value=_formula_ref_for(row["item_class"],
                                                row["designation"]))

    boq = wb.create_sheet("BOQ")
    boq.append(list(BOQ_HEADERS))
    bom = wb.create_sheet("BOM")
    bom.append(list(BOM_HEADERS))

    ps = wb.create_sheet("PRICING_SCHEDULE")
    ps.append(list(PS_HEADERS))
    deviations = _build_pricing_schedule(ps, rows)

    info = {
        "job": job,
        "rows": len(rows),
        "census_rows": len(census_rows),
        "manual_rows": len(manual_rows),
        "grid_rows": len(grid_rows),
        "conflict_rows": sum(1 for r in rows
                             if r["notes"].startswith("CONFLICT:")),
        "low_rows": sum(1 for r in rows if r["confidence"] == "low"),
        "col_types_uncounted": uncounted_col_types(job, db_path),
        "plate_types_uncounted": uncounted_plate_types(job, db_path),
        "spec_deviations": deviations,
    }
    return wb, info


def _sum_formula(col, rows, excel_row) -> str:
    """=SUM over contiguous ranges of the given rows' cells. Rows are
    pre-sorted so partitions are contiguous; ranges keep the formula
    far inside Excel's 255-argument and 8192-character limits."""
    idxs = sorted(excel_row[r["item_id"]] for r in rows)
    ranges = []
    start = prev = idxs[0]
    for i in idxs[1:]:
        if i == prev + 1:
            prev = i
            continue
        ranges.append((start, prev))
        start = prev = i
    ranges.append((start, prev))
    parts = [f"TAKEOFF!{col}{a}:{col}{b}" for a, b in ranges]
    return f"=SUM({','.join(parts)})"


def _build_pricing_schedule(ps, rows) -> list:
    """Skeleton per section 11: one line per quantity class present, in
    class order; qty cells are formulas over TAKEOFF cells, never typed
    numbers; source_refs lists every contributing item_id. Returns a
    list of explicit spec deviations.

    Known deviation, surfaced not silent: section 11 mandates one TON
    line for COL/BEAM/JST, summing derived tons cells. Tons cells exist
    only on LINEAR rows; a class carrying only COUNT rows has nothing
    to sum, so the exporter emits an EA count line (the PLATE/ANCH
    pattern) and reports the deviation. Spec amendment candidate.

    OPENING rows (3.4) never form their own line: section 11 cites
    them in the parent deck line's source_refs; deductions ride
    downstream in BOQ and BOM."""
    excel_row = {}
    for r_idx, row in enumerate(rows, start=3):
        excel_row[row["item_id"]] = r_idx

    deviations = []
    seq = 0

    def add_line(cls, description, refs, qty_formula, unit):
        nonlocal seq
        seq += 1
        suffix = "" if seq == 1 else f"-{seq}"
        ps.append([f"PS-{cls}{suffix}", description, ", ".join(refs),
                   qty_formula, unit])

    for cls in CLASS_ORDER:
        cls_rows = [r for r in rows if r["item_class"] == cls]
        if not cls_rows:
            continue
        seq = 0
        linear = [r for r in cls_rows if r["mode"] == "LINEAR"]
        count = [r for r in cls_rows if r["mode"] == "COUNT"]
        area = [r for r in cls_rows if r["mode"] == "AREA"]
        if linear:
            add_line(cls, f"{cls} derived tonnage",
                     [r["item_id"] for r in linear],
                     _sum_formula("N", linear, excel_row), "TON")
        if count:
            add_line(cls, f"{cls} measured counts",
                     [r["item_id"] for r in count],
                     _sum_formula("D", count, excel_row), "EA")
            if cls in ("COL", "BEAM", "JST") and not linear:
                deviations.append(
                    f"PS-{cls}: section 11 mandates a TON line but the "
                    "class has no LINEAR rows (no tons cells to sum); "
                    "emitted an EA count line instead. Spec amendment "
                    "candidate.")
        if area:
            if cls == "DECK":
                openings = [r for r in area
                            if r["designation"].upper() == "OPENING"]
                opening_parent = {}
                for r in openings:
                    m = re.search(r"OPENING:\s*([A-Z]+-\d+)",
                                  r["notes"])
                    if m:
                        opening_parent.setdefault(m.group(1),
                                                  []).append(r)
                    else:
                        deviations.append(
                            f"{r['item_id']}: OPENING row with no "
                            "parseable parent; cited on no PS line")
                by_desig = {}
                for r in area:
                    if r["designation"].upper() == "OPENING":
                        continue
                    by_desig.setdefault(r["designation"], []).append(r)
                for desig in sorted(by_desig):
                    drows = by_desig[desig]
                    refs = [r["item_id"] for r in drows]
                    # Section 11: OPENING rows are cited in
                    # source_refs of their parent deck line, never
                    # summed into the gross area.
                    for parent_id in list(refs):
                        for op in opening_parent.get(parent_id, []):
                            refs.append(op["item_id"])
                    add_line(cls, f"DECK {desig} gross area", refs,
                             _sum_formula("D", drows, excel_row), "SF")
            else:
                add_line(cls, f"{cls} measured area",
                         [r["item_id"] for r in area],
                         _sum_formula("D", area, excel_row), "SF")
    return deviations


# -- stamp -------------------------------------------------------------------

def _next_export_number(job, out_dir) -> int:
    """Monotonic per job (13.3). The register file survives moved or
    archived exports, so an R number is never reissued even when the
    file it named has left the directory; the glob is a floor for
    registers lost or hand-edited."""
    register = Path(out_dir) / "export_register.json"
    issued = {}
    if register.exists():
        try:
            issued = json.loads(register.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            issued = {}
    pat = re.compile(rf"^{re.escape(job)}_TAKEOFF_R(\d+)_[0-9a-f]{{12}}"
                     r"\.xlsx$")
    n = int(issued.get(job, 0))
    for p in Path(out_dir).glob(f"{job}_TAKEOFF_R*.xlsx"):
        m = pat.match(p.name)
        if m:
            n = max(n, int(m.group(1)))
    issued[job] = n + 1
    register.write_text(json.dumps(issued, indent=1), encoding="utf-8")
    return n + 1


def _assembly_version() -> str:
    """Git short hash of takeoff_pipeline/assemblies/. Raises when no
    version is determinable: F1 = NONE means exactly one thing per
    13.3 (BOQ and BOM are empty), so a populated workbook with an
    unknown assembly version must refuse the stamp rather than lie."""
    asm = _PKG / "assemblies"
    if asm.is_dir():
        try:
            out = subprocess.run(
                ["git", "log", "-1", "--format=%h", "--", str(asm)],
                capture_output=True, text=True, cwd=str(_PKG.parent),
                timeout=10)
            short = out.stdout.strip()
            if short:
                return short
        except Exception:
            pass
    raise ValueError(
        "cannot stamp: BOQ/BOM hold data but the assembly-library "
        "version is undeterminable (takeoff_pipeline/assemblies/ "
        "missing or not committed). F1=NONE is reserved for empty "
        "BOQ and BOM per 13.3.")


def stamp_workbook(unstamped_path, job, out_dir=None) -> dict:
    """13.3: write the stamp row and save under the final name. Runs
    AFTER validation per 13.1; nothing writes to the file after this."""
    import openpyxl

    out_dir = Path(out_dir) if out_dir else DEFAULT_OUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(str(unstamped_path))
    ws = wb["TAKEOFF"]

    boq_empty = wb["BOQ"].max_row <= 1
    bom_empty = wb["BOM"].max_row <= 1
    digest = compute_takeoff_hash(ws)
    export_n = _next_export_number(job, out_dir)

    ws.cell(row=1, column=1, value=SCHEMA_LITERAL)
    ws.cell(row=1, column=2, value=job)
    ws.cell(row=1, column=3,
            value=datetime.now(timezone.utc).isoformat())
    ws.cell(row=1, column=4, value=digest)
    ws.cell(row=1, column=5, value=f"R{export_n}")
    ws.cell(row=1, column=6,
            value="NONE" if (boq_empty and bom_empty)
            else _assembly_version())

    final = out_dir / f"{job}_TAKEOFF_R{export_n}_{digest[:12]}.xlsx"
    wb.save(str(final))
    return {"path": str(final), "hash": digest,
            "export_number": f"R{export_n}"}


# -- pipeline entry ----------------------------------------------------------

def export(job, db_path=None, manual_csv=None, out_dir=None,
           stamp=False, extra_rows=None) -> dict:
    """13.1 steps 1 and 4: build the unstamped workbook and validate
    it. The DEFAULT stops there (status UNSTAMPED_VALID): step 2 is
    the operator filling lb_per_ft from bridge/aisc_validator.py
    output, and stamping before that step would freeze blank weights
    into an immutable file. stamp=True chains straight to the stamp
    and is only sane for workbooks with no LINEAR rows awaiting
    operator input; stamp_file() is the operator's step 5 for
    everything else. The unstamped file always survives on validation
    failure for inspection; nothing failing validation is stamped."""
    out_dir = Path(out_dir) if out_dir else DEFAULT_OUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    wb, info = build_workbook(job, db_path, manual_csv, extra_rows)
    unstamped = out_dir / f"{job}_TAKEOFF_UNSTAMPED.xlsx"
    wb.save(str(unstamped))

    from takeoff_pipeline import validate_takeoff
    report = validate_takeoff.validate_file(unstamped)
    info["validation"] = report["summary"]
    info["unstamped"] = str(unstamped)
    if report["hard_fails"]:
        info["status"] = "VALIDATION_FAILED"
        return info
    if not stamp:
        info["status"] = "UNSTAMPED_VALID"
        info["next_step"] = ("operator fills lb_per_ft on LINEAR rows "
                             "from bridge/aisc_validator.py output, "
                             "then stamp_file()")
        return info

    stamped = stamp_workbook(unstamped, job, out_dir)
    unstamped.unlink()
    del info["unstamped"]
    info.update(stamped)
    info["status"] = "STAMPED"
    return info


def stamp_file(path, job, out_dir=None) -> dict:
    """13.1 step 5 for an unstamped workbook the operator finished:
    re-validate, then stamp. Validation failure refuses the stamp."""
    from takeoff_pipeline import validate_takeoff
    report = validate_takeoff.validate_file(path)
    if report["hard_fails"]:
        return {"status": "VALIDATION_FAILED",
                "validation": report["summary"],
                "hard_fails": report["hard_fails"]}
    out = stamp_workbook(path, job, out_dir)
    Path(path).unlink()
    out["status"] = "STAMPED"
    return out


def _flag_value(args, flag):
    if flag not in args:
        return None
    i = args.index(flag)
    if i + 1 >= len(args):
        raise ValueError(f"{flag} needs a value")
    return args[i + 1]


def main() -> int:
    usage = ("usage: py -m takeoff_pipeline.export_xlsx <job> "
             "[--db <census.db>] [--manual <rows.csv>] [--out <dir>] "
             "[--stamp] | <job> --stamp-file <unstamped.xlsx> "
             "[--out <dir>]")
    args = sys.argv[1:]
    if not args or args[0].startswith("--"):
        print(usage)
        return 2
    job = args[0]
    try:
        stamp_path = _flag_value(args, "--stamp-file")
        out_dir = _flag_value(args, "--out")
        if stamp_path:
            info = stamp_file(stamp_path, job, out_dir)
        else:
            info = export(job, db_path=_flag_value(args, "--db"),
                          manual_csv=_flag_value(args, "--manual"),
                          out_dir=out_dir,
                          stamp="--stamp" in args)
    except ValueError as e:
        print(e)
        print(usage)
        return 2
    for k, v in info.items():
        print(f"{k}: {v}")
    return 0 if info.get("status") in ("STAMPED", "UNSTAMPED_VALID") \
        else 1


if __name__ == "__main__":
    raise SystemExit(main())
