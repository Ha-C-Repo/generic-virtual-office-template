"""PC3 shop progress capture (Prompt 11, implements P17).

The rule: if it takes more than 60 seconds a day to fill in, the shop
will not fill it in. One table, two entry surfaces (a printable
one-page daily sheet plus a 5-field CLI), one weekly rollup CSV.

Storage. Table progress_log in the shop floor database. The legacy
daily_production table in that db is per-project grain with no worker
and no WBS line, so PC3 adds a table to the SAME database rather than
retrofitting it; nothing here is a parallel store. Path resolution is
identical to the PC4 reader: dev resolves to the repo data/
shop_floor.db, frozen resolves to the LOCALAPPDATA data root (the
data/ copy inside the PyInstaller bundle is a build-time snapshot,
never a live store). Hard Rule 11: every connection sets WAL plus a
busy timeout.

Project attribution. PC4 filters rows by project with an exact match
and excludes unattributed rows, so every entry needs a project. To
keep the daily entry at five fields, the project is a sticky default:
set once with the project command, inherited by every entry that does
not pass --project explicitly. An entry with no project and no default
is rejected with the one-time fix, never silently dropped downstream.

Schema contract. bridge/project_controls.py (PC4, already shipped) is
the consumer and fixes the column meanings:

  date, person, wbs_line, hours, pieces_done, tons_done, issues_text,
  project (optional filter, exact match). Extra columns are ignored.
  - production lines log units: tons in tons_done, pieces in
    pieces_done, SF also in pieces_done (the PC3 capture convention;
    PC4 downgrades confidence on SF lines for exactly this reason).
  - milestone lines (shop drawings) log the rule of credit in
    pieces_done: issued 20, approved 75, released 100. PC4 reads the
    highest credit recorded.
  - client-caused issues carry the uppercase tag CLIENT in
    issues_text, for example "CLIENT: GC held erection 3 days".

P17 separation. Production-driven and milestone-driven progress never
mix. The CLI has separate commands (log vs milestone), the daily sheet
has separate sections, and the writers reject an entry whose type
contradicts the line's recorded history.

Weekly rollup. Per WBS line, planned vs actual units and hours for the
week ending Sunday (same week convention as the PC4 S-curve), exported
as CSV for PC4. Planned values come from the PC1 baseline xlsx WBS
sheet when one is found (parsed by the same loader PC4 uses); without
a baseline the actuals still roll up and the planned columns stay
empty, with a warning. An unfrozen baseline is reported as draft (P14).

Free tooling: stdlib plus openpyxl (already a dependency) for the
sheet. No Bridge methods here; PC4 owns the Bridge surface.
"""

import argparse
import csv
import os
import sqlite3
import sys
import threading
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

_PKG = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = _PKG / "exports"

MILESTONE_CREDITS = (20.0, 75.0, 100.0)
CREDIT_LEGEND = "issued 20, approved 75, released 100"
PRODUCTION = "production"
MILESTONE = "milestone"

_lock = threading.Lock()


# -- storage ---------------------------------------------------------------

def _db_path() -> Path:
    """Shop floor db path, resolved with the same logic as the PC4 reader
    (bridge/project_controls.py _progress_db_path) so writes and reads
    always land on the same file. Dev: the repo data/shop_floor.db via
    resource_path. Frozen: the LOCALAPPDATA data root; resource_path
    would point inside sys._MEIPASS, where the bundled data/ copy is a
    read-only build-time snapshot, and reading or writing it would
    silently split the writer from the reader. The bridge.shop_floor _DB
    import is shadowed by the bridge/shop_floor package today and always
    fails; the try is kept so PC3 and PC4 move together if that ever
    changes."""
    try:
        from bridge.shop_floor import _DB
        return Path(_DB)
    except Exception:
        pass
    from vo_app._resources import is_frozen, resource_path
    if is_frozen():
        local = os.environ.get("LOCALAPPDATA")
        if local:
            return Path(local) / "YourCompany" / "VirtualOffice" / "data" \
                / "shop_floor.db"
    return Path(resource_path("data/shop_floor.db"))


def _conn(db_path: Path) -> sqlite3.Connection:
    """SQLite connection per Hard Rule 11: WAL plus busy timeout."""
    db_path.parent.mkdir(parents=True, exist_ok=True)
    c = sqlite3.connect(str(db_path), check_same_thread=False, timeout=10)
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA busy_timeout=10000")
    c.row_factory = sqlite3.Row
    return c


def ensure_schema(db_path=None) -> Path:
    """Create progress_log if missing. Column names match the PC4 alias
    table in bridge/project_controls.py; progress_type and logged_at are
    PC3-side extras PC4 ignores."""
    path = Path(db_path) if db_path else _db_path()
    with _lock:
        c = _conn(path)
        try:
            c.executescript("""
                CREATE TABLE IF NOT EXISTS progress_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    person TEXT NOT NULL,
                    project TEXT DEFAULT '',
                    wbs_line TEXT NOT NULL,
                    progress_type TEXT NOT NULL DEFAULT 'production',
                    hours REAL DEFAULT 0,
                    pieces_done REAL DEFAULT 0,
                    tons_done REAL DEFAULT 0,
                    issues_text TEXT DEFAULT '',
                    logged_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_progress_date
                    ON progress_log(date);
                CREATE INDEX IF NOT EXISTS idx_progress_wbs
                    ON progress_log(wbs_line);
                CREATE TABLE IF NOT EXISTS shop_log_meta (
                    key TEXT PRIMARY KEY,
                    value TEXT
                );
            """)
            c.commit()
        finally:
            c.close()
    return path


# -- value handling --------------------------------------------------------

def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _line_type_on_record(c: sqlite3.Connection, wbs_line: str,
                         project: str) -> str:
    """The progress type this WBS line has already been logged under for
    THIS project, or empty when the line has no history. Project-scoped:
    the P15 WBS template reuses line ids (SD-01, FAB-S1) across projects,
    and another project's history must not reject a valid entry."""
    row = c.execute(
        "SELECT progress_type FROM progress_log WHERE wbs_line=? "
        "AND project=? ORDER BY id DESC LIMIT 1",
        (wbs_line, project)).fetchone()
    return row["progress_type"] if row else ""


def set_default_project(project: str, db_path=None) -> dict:
    """Sticky project default: set once, every entry without --project
    inherits it. Keeps the daily entry at five fields without writing
    unattributed rows PC4 would exclude from every metric."""
    pid = (project or "").strip()
    if not pid:
        return {"error": "project id is empty",
                "fix": "pass the awarded project code, for example "
                       "PRJ-2026-ACP-001"}
    path = ensure_schema(db_path)
    with _lock:
        c = _conn(path)
        try:
            c.execute(
                "INSERT INTO shop_log_meta (key, value) VALUES "
                "('default_project', ?) ON CONFLICT(key) DO UPDATE SET "
                "value=excluded.value", (pid,))
            c.commit()
        finally:
            c.close()
    return {"default_project": pid, "db": str(path)}


def get_default_project(db_path=None) -> str:
    """The sticky default project, or empty. Never creates the db: a
    read-only question must not materialize an empty store."""
    path = Path(db_path) if db_path else _db_path()
    if not path.exists():
        return ""
    with _lock:
        c = _conn(path)
        try:
            row = c.execute(
                "SELECT value FROM shop_log_meta WHERE "
                "key='default_project'").fetchone()
        except sqlite3.OperationalError:
            return ""
        finally:
            c.close()
    return row["value"] if row else ""


def _resolve_project(project: str, db_path):
    """Explicit project, else the sticky default, else an error dict.
    PC4 excludes unattributed rows from every project-filtered metric,
    so an entry that would write project='' is rejected up front."""
    pid = (project or "").strip()
    if pid:
        return pid
    pid = get_default_project(db_path)
    if pid:
        return pid
    return {"error": "no project given and no default set; an entry "
                     "without project attribution is excluded from every "
                     "PC4 metric for every project",
            "fix": "set it once: py -3.13 -m takeoff_pipeline.shop_log "
                   "project <project-id>  (or pass --project on this "
                   "entry)"}


def _baseline_line(project: str, wbs_line: str, baseline_path=None):
    """Best-effort lookup of one WBS line in the PC1 baseline. Returns
    the line dict or None. Never raises and never blocks an entry on a
    missing or unreadable baseline; capture must keep working before
    PC1 exists."""
    try:
        baseline = _find_baseline(project, baseline_path)
        if not baseline or "error" in baseline:
            return None
        for ln in baseline.get("lines", []):
            if ln["wbs_line"] == wbs_line:
                return ln
    except Exception:
        return None
    return None


def _insert(db_path, entry_date, person, project, wbs_line, progress_type,
            hours, pieces_done, tons_done, issues_text) -> dict:
    path = ensure_schema(db_path)
    now = datetime.now(timezone.utc).isoformat()
    with _lock:
        c = _conn(path)
        try:
            recorded = _line_type_on_record(c, wbs_line, project)
            if recorded and recorded != progress_type:
                return {"error": f"WBS line {wbs_line} is on record as "
                                 f"{recorded}-driven; a {progress_type} "
                                 "entry would mix the two progress types "
                                 "and corrupt percent complete (P17)",
                        "fix": f"use the {recorded} entry command for this "
                               "line, or pick the correct WBS line"}
            cur = c.execute(
                "INSERT INTO progress_log (date, person, project, wbs_line,"
                " progress_type, hours, pieces_done, tons_done, issues_text,"
                " logged_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (entry_date.isoformat(), person, project, wbs_line,
                 progress_type, hours, pieces_done, tons_done, issues_text,
                 now))
            c.commit()
            return {"id": cur.lastrowid, "date": entry_date.isoformat(),
                    "person": person, "wbs_line": wbs_line,
                    "progress_type": progress_type, "db": str(path)}
        finally:
            c.close()


def _common_checks(person, wbs_line, entry_date, hours) -> dict | None:
    if not (person or "").strip():
        return {"error": "person is required",
                "fix": "pass --person, for example --person Mario"}
    if not (wbs_line or "").strip():
        return {"error": "wbs_line is required",
                "fix": "pass --wbs with the WBS line id, for example "
                       "--wbs FAB-S1"}
    if entry_date is None:
        return {"error": "date not understood",
                "fix": "use YYYY-MM-DD, for example 2026-06-12"}
    if hours < 0:
        return {"error": "hours cannot be negative", "fix": "check the "
                "value; hours are crew hours for the line, zero is fine"}
    return None


def log_production(person: str, wbs_line: str, hours: float = 0.0,
                   pieces: float = 0.0, tons: float = 0.0, sf: float = 0.0,
                   issues: str = "", entry_date=None, project: str = "",
                   db_path=None, baseline_path=None) -> dict:
    """One production-driven entry: units done plus crew hours for one WBS
    line on one day. SF is stored in pieces_done per the PC3 convention
    PC4 reads (a line is pieces OR SF, never both)."""
    d = _parse_date(entry_date) if entry_date else date.today()
    bad = _common_checks(person, wbs_line, d, hours)
    if bad:
        return bad
    if pieces < 0 or tons < 0 or sf < 0:
        return {"error": "units cannot be negative",
                "fix": "check pieces, tons, and sf values"}
    if pieces and sf:
        return {"error": "pieces and sf both given; they share the "
                         "pieces_done column and would overwrite each "
                         "other",
                "fix": "a WBS line counts pieces or SF, not both; log "
                       "the one the line's unit uses"}
    if hours == 0 and pieces == 0 and tons == 0 and sf == 0:
        return {"error": "nothing to record: hours and units are all zero",
                "fix": "log the hours worked even when no units finished"}
    pid = _resolve_project(project, db_path)
    if isinstance(pid, dict):
        return pid
    note = ""
    line = _baseline_line(pid, wbs_line.strip(), baseline_path)
    if line:
        if line.get("progress_type") == MILESTONE:
            return {"error": f"WBS line {wbs_line.strip()} is "
                             "milestone-driven in the baseline; production "
                             "units on it would mix the two progress types "
                             "(P17)",
                    "fix": "use the milestone command with the "
                           f"rule-of-credit step: {CREDIT_LEGEND}"}
        unit = (line.get("unit") or "").lower()
        if "ton" in unit and tons == 0 and (pieces or sf):
            note = (f"baseline unit for this line is {line['unit']}; PC4 "
                    "reads tons_done on it, so these units will not count "
                    "toward percent complete")
        elif unit and "ton" not in unit and tons and not (pieces or sf):
            note = (f"baseline unit for this line is {line['unit']}; PC4 "
                    "reads pieces_done on it, so these tons will not count "
                    "toward percent complete")
    result = _insert(db_path, d, person.strip(), pid, wbs_line.strip(),
                     PRODUCTION, float(hours), float(pieces or sf),
                     float(tons), (issues or "").strip())
    if note and "error" not in result:
        result["note"] = note
    return result


def log_milestone(person: str, wbs_line: str, credit: float,
                  hours: float = 0.0, issues: str = "", entry_date=None,
                  project: str = "", db_path=None,
                  baseline_path=None) -> dict:
    """One milestone-driven entry (shop drawings): the rule-of-credit step
    reached, plus drafting or checking hours. Credit lands in pieces_done;
    PC4 reads the highest credit recorded for the line."""
    d = _parse_date(entry_date) if entry_date else date.today()
    bad = _common_checks(person, wbs_line, d, hours)
    if bad:
        return bad
    if float(credit) not in MILESTONE_CREDITS:
        return {"error": f"credit {credit:g} is not a rule-of-credit step",
                "fix": f"use one of {CREDIT_LEGEND}"}
    pid = _resolve_project(project, db_path)
    if isinstance(pid, dict):
        return pid
    line = _baseline_line(pid, wbs_line.strip(), baseline_path)
    if line and line.get("progress_type") == PRODUCTION:
        return {"error": f"WBS line {wbs_line.strip()} is production-"
                         "driven in the baseline; a rule-of-credit entry "
                         "on it would mix the two progress types (P17)",
                "fix": "use the log command with the units done (pieces, "
                       "tons, or SF)"}
    result = _insert(db_path, d, person.strip(), pid, wbs_line.strip(),
                     MILESTONE, float(hours), float(credit), 0.0,
                     (issues or "").strip())
    if "error" in result:
        return result
    path = Path(result["db"])
    with _lock:
        c = _conn(path)
        try:
            prior = c.execute(
                "SELECT MAX(pieces_done) AS m FROM progress_log "
                "WHERE wbs_line=? AND project=? AND progress_type=? "
                "AND id<?",
                (wbs_line.strip(), pid, MILESTONE,
                 result["id"])).fetchone()
        finally:
            c.close()
    if prior and prior["m"] is not None and float(credit) <= prior["m"]:
        result["note"] = (f"line was already at credit {prior['m']:g}; "
                          "entry recorded, PC4 reads the highest credit")
    return result


def recent_entries(days: int = 7, db_path=None) -> list:
    """Latest entries for a quick did-it-land check."""
    path = ensure_schema(db_path)
    cutoff = (date.today() - timedelta(days=days)).isoformat()
    with _lock:
        c = _conn(path)
        try:
            rows = c.execute(
                "SELECT * FROM progress_log WHERE date >= ? "
                "ORDER BY date DESC, id DESC", (cutoff,)).fetchall()
        finally:
            c.close()
    return [dict(r) for r in rows]


# -- baseline (planned side) -----------------------------------------------

def _find_baseline(project: str = "", baseline_path=None) -> dict:
    """Locate and parse the PC1 baseline xlsx with the same loader PC4
    uses, so planned numbers here always agree with the SPI/CPI view.
    Returns {"lines": [...], "baseline_flag": bool, ...} or
    {"error": ...} or {} when there is nothing to look for."""
    from bridge.project_controls import (_awarded_root, _find_baseline_xlsx,
                                         _find_project_folder,
                                         _load_baseline)
    if baseline_path:
        xlsx = Path(baseline_path)
        if not xlsx.exists():
            return {"error": f"baseline xlsx not found at {xlsx}",
                    "fix": "check the path"}
        return _load_baseline(xlsx)
    if not (project or "").strip():
        return {}
    root = _awarded_root()
    folder = _find_project_folder(project, root) if root.is_dir() else None
    if folder is None:
        return {}
    xlsx = _find_baseline_xlsx(folder)
    if xlsx is None:
        return {}
    return _load_baseline(xlsx)


def _units_for_row(row: dict, unit: str) -> float:
    """Same unit mapping as PC4: tons lines read tons_done, everything
    else (pieces, EA, SF) reads pieces_done."""
    if "ton" in unit:
        return row["tons_done"] or 0.0
    return row["pieces_done"] or 0.0


# -- weekly rollup ----------------------------------------------------------

def _week_window(week_ending=None) -> tuple:
    """Inclusive Monday-to-Sunday window. Default: the week containing
    today, ending Sunday (the PC4 S-curve week convention)."""
    end = _parse_date(week_ending) if week_ending else None
    if end is None:
        today = date.today()
        end = today + timedelta(days=6 - today.weekday())
    start = end - timedelta(days=6)
    return start, end


def _load_rows(db_path, project: str) -> dict:
    """Read progress_log rows, optionally filtered to one project (exact
    match, mirroring PC4: unattributed rows are excluded when filtering,
    never double-counted)."""
    path = Path(db_path) if db_path else _db_path()
    if not path.exists():
        return {"rows": [], "path": str(path),
                "warnings": [f"shop floor database not found at {path}; "
                             "no actuals to roll up"]}
    with _lock:
        c = _conn(path)
        try:
            found = c.execute(
                "SELECT name FROM sqlite_master WHERE type='table' "
                "AND name='progress_log'").fetchone()
            if not found:
                return {"rows": [], "path": str(path),
                        "warnings": ["progress_log table not found; log an "
                                     "entry first"]}
            raw = c.execute("SELECT * FROM progress_log").fetchall()
        finally:
            c.close()
    pid = (project or "").strip().lower()
    rows, unattributed, undated = [], 0, 0
    for r in raw:
        rec = dict(r)
        if pid:
            proj = str(rec.get("project") or "").strip().lower()
            if not proj:
                unattributed += 1
                continue
            if proj != pid:
                continue
        d = _parse_date(rec.get("date"))
        if d is None:
            undated += 1
            continue
        rec["date"] = d
        rows.append(rec)
    warnings = []
    if unattributed:
        warnings.append(f"{unattributed} rows have no project attribution; "
                        "excluded from this project rollup")
    if undated:
        warnings.append(f"{undated} rows have no parseable date; excluded")
    return {"rows": rows, "path": str(path), "warnings": warnings}


def _production_line_rollup(rows: list, unit: str, week_start: date,
                            week_end: date) -> dict:
    todate = [r for r in rows if r["date"] <= week_end]
    week = [r for r in todate if r["date"] >= week_start]
    return {
        "units_done_todate": round(sum(_units_for_row(r, unit)
                                       for r in todate), 3),
        "units_done_week": round(sum(_units_for_row(r, unit)
                                     for r in week), 3),
        "hours_todate": round(sum(r["hours"] or 0.0 for r in todate), 2),
        "hours_week": round(sum(r["hours"] or 0.0 for r in week), 2),
    }


def _milestone_line_rollup(rows: list, week_start: date,
                           week_end: date) -> dict:
    """Credit is a level, not a flow: to-date is the highest credit on or
    before week end; the week column is the advance during the week."""
    todate = [r for r in rows if r["date"] <= week_end]
    before = [r for r in rows if r["date"] < week_start]
    credit_now = max((r["pieces_done"] or 0.0 for r in todate), default=0.0)
    credit_before = max((r["pieces_done"] or 0.0 for r in before),
                        default=0.0)
    week = [r for r in todate if r["date"] >= week_start]
    return {
        "units_done_todate": round(credit_now, 3),
        "units_done_week": round(max(credit_now - credit_before, 0.0), 3),
        "hours_todate": round(sum(r["hours"] or 0.0 for r in todate), 2),
        "hours_week": round(sum(r["hours"] or 0.0 for r in week), 2),
    }


def _pct(num, den):
    if num is None or not den or den <= 0:
        return None
    return round(num / den, 4)


def weekly_rollup(week_ending=None, project: str = "", baseline_path=None,
                  db_path=None) -> dict:
    """Per WBS line, planned vs actual units and hours for one week plus
    cumulative to date. Feeds the PC4 CSV export. Pure arithmetic on
    logged rows and the baseline; nothing here is generated."""
    if week_ending and _parse_date(week_ending) is None:
        return {"error": f"week ending '{week_ending}' not understood",
                "fix": "use YYYY-MM-DD, for example 2026-06-14"}
    week_start, week_end = _week_window(week_ending)
    project = (project or "").strip() or get_default_project(db_path)
    loaded = _load_rows(db_path, project)
    warnings = list(loaded["warnings"])

    if not project:
        blended = {str(r.get("project") or "").strip()
                   for r in loaded["rows"]}
        blended.discard("")
        if len(blended) > 1:
            warnings.append(f"rows from {len(blended)} projects are "
                            "blended in this rollup; pass --project for "
                            "a per-project report")

    by_line = {}
    for r in loaded["rows"]:
        by_line.setdefault(str(r["wbs_line"]).strip(), []).append(r)

    baseline = _find_baseline(project, baseline_path)
    if "error" in baseline:
        return baseline
    base_lines = baseline.get("lines", [])
    if not base_lines:
        warnings.append("no PC1 baseline found; planned columns are empty "
                        "(actuals only)")
    elif not baseline.get("baseline_flag"):
        warnings.append("baseline xlsx has no BASELINE flag; planned "
                        "values are draft, not frozen (P14)")
    keys = [ln["wbs_line"] for ln in base_lines]
    dupes = sorted({k for k in keys if keys.count(k) > 1})
    if dupes:
        warnings.append("duplicate WBS lines in the baseline repeat "
                        "their actuals once per row in this report: "
                        + ", ".join(dupes[:5]))

    out_rows = []
    seen = set()
    for ln in base_lines:
        key = ln["wbs_line"]
        seen.add(key)
        rows = by_line.get(key, [])
        ptype = ln.get("progress_type") or PRODUCTION
        unit = (ln.get("unit") or "").lower()
        note = ""
        if ptype == MILESTONE:
            roll = _milestone_line_rollup(rows, week_start, week_end)
            planned_units = 100.0
            unit_label = "credit"
        else:
            roll = _production_line_rollup(rows, unit, week_start, week_end)
            planned_units = ln.get("planned_units")
            unit_label = ln.get("unit") or ""
            if unit in ("sf", "sq ft", "sqft"):
                note = "SF logged in pieces_done per the PC3 convention"
        logged_types = {r["progress_type"] for r in rows
                        if r.get("progress_type")}
        if logged_types and logged_types != {ptype}:
            note = (f"logged as {'/'.join(sorted(logged_types))} but the "
                    f"baseline says {ptype}; check the entries (P17)")
        out_rows.append({
            "wbs_line": key,
            "progress_type": ptype,
            "unit": unit_label,
            "planned_units": planned_units,
            "units_done_todate": roll["units_done_todate"],
            "units_done_week": roll["units_done_week"],
            "pct_units_complete": _pct(roll["units_done_todate"],
                                       planned_units),
            "planned_hours": ln.get("planned_hours"),
            "hours_todate": roll["hours_todate"],
            "hours_week": roll["hours_week"],
            "pct_hours_used": _pct(roll["hours_todate"],
                                   ln.get("planned_hours")),
            "in_baseline": True,
            "note": note,
        })

    for key in sorted(set(by_line) - seen):
        rows = by_line[key]
        types = {r.get("progress_type") or PRODUCTION for r in rows}
        ptype = MILESTONE if types == {MILESTONE} else PRODUCTION
        note = "not in baseline" if base_lines else "no baseline"
        if len(types) > 1:
            note += "; mixed progress types logged, check entries (P17)"
        if ptype == MILESTONE:
            roll = _milestone_line_rollup(rows, week_start, week_end)
            unit_label = "credit"
            planned_units = None
        else:
            tons = sum(r["tons_done"] or 0.0 for r in rows)
            pieces = sum(r["pieces_done"] or 0.0 for r in rows)
            unit_label = "tons" if tons and not pieces else \
                "pieces" if pieces and not tons else ""
            roll = _production_line_rollup(rows, unit_label, week_start,
                                           week_end)
            planned_units = None
            if not unit_label and (tons or pieces):
                note += "; both pieces and tons logged, unit unknown"
        out_rows.append({
            "wbs_line": key,
            "progress_type": ptype,
            "unit": unit_label,
            "planned_units": planned_units,
            "units_done_todate": roll["units_done_todate"],
            "units_done_week": roll["units_done_week"],
            "pct_units_complete": _pct(roll["units_done_todate"],
                                       planned_units),
            "planned_hours": None,
            "hours_todate": roll["hours_todate"],
            "hours_week": roll["hours_week"],
            "pct_hours_used": None,
            "in_baseline": False,
            "note": note,
        })

    return {
        "week_start": week_start.isoformat(),
        "week_ending": week_end.isoformat(),
        "project": (project or "").strip(),
        "rows": out_rows,
        "warnings": warnings,
        "sources": {
            "progress_db": loaded["path"],
            "baseline_xlsx": baseline.get("path", ""),
        },
    }


CSV_COLUMNS = ("wbs_line", "progress_type", "unit", "planned_units",
               "units_done_todate", "units_done_week", "pct_units_complete",
               "planned_hours", "hours_todate", "hours_week",
               "pct_hours_used", "in_baseline", "note")


def write_rollup_csv(rollup: dict, out_path=None) -> Path:
    """Write the weekly rollup as the PC4 hand-off CSV. None values become
    empty cells, never zeros (a blank planned column is missing data, not
    a zero plan)."""
    if out_path:
        out = Path(out_path)
    else:
        DEFAULT_OUT_DIR.mkdir(parents=True, exist_ok=True)
        out = DEFAULT_OUT_DIR / \
            f"shop_rollup_week_{rollup['week_ending']}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["week_start", rollup["week_start"],
                    "week_ending", rollup["week_ending"],
                    "project", rollup.get("project", "")])
        w.writerow(CSV_COLUMNS)
        for r in rollup["rows"]:
            w.writerow(["" if r[k] is None else r[k] for k in CSV_COLUMNS])
    return out


# -- daily sheet -------------------------------------------------------------

def build_daily_sheet(sheet_date=None, out_path=None, project: str = "",
                      baseline_path=None) -> dict:
    """Printable one-page daily log for Mario's day shift. Two sections,
    never mixed (P17): production lines count units; the milestone section
    records the rule-of-credit step only. When a PC1 baseline is found the
    WBS lines are pre-filled so filling the sheet is circling and writing
    numbers, nothing else."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.worksheet.properties import PageSetupProperties
    except ImportError:
        return {"error": "openpyxl is not installed; cannot build the "
                         "daily sheet",
                "fix": "py -3.13 -m pip install openpyxl"}
    d = _parse_date(sheet_date) if sheet_date else date.today()
    if d is None:
        return {"error": "date not understood",
                "fix": "use YYYY-MM-DD, for example 2026-06-12"}

    baseline = _find_baseline(project, baseline_path)
    if "error" in baseline:
        return baseline
    prod_lines = [ln for ln in baseline.get("lines", [])
                  if (ln.get("progress_type") or PRODUCTION) == PRODUCTION]
    mile_lines = [ln for ln in baseline.get("lines", [])
                  if ln.get("progress_type") == MILESTONE]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAILY LOG"

    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    h1 = Font(bold=True, size=14)
    h2 = Font(bold=True, size=11)
    small = Font(size=9)
    wrap = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 16, "B": 24, "C": 12, "D": 8, "E": 8, "F": 8, "G": 8,
              "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws["A1"] = "YOUR COMPANY - SHOP DAILY LOG"
    ws["A1"].font = h1
    ws["A2"] = f"Date: {d.isoformat()}"
    ws["C2"] = f"Project: {project or '____________'}"
    ws["F2"] = "Shift lead: ____________"
    ws["A3"] = ("One row per person per WBS line. Under 60 seconds total. "
                "Hours are crew hours on that line.")
    ws["A3"].font = small

    row = 5
    ws.cell(row=row, column=1,
            value="SECTION A - PRODUCTION (count units)").font = h2
    row += 1
    headers = ("WBS line", "Scope", "Person", "Hours", "Pieces", "Tons",
               "SF", "Issues / delays (write CLIENT: if client-caused)")
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = h2
        cell.border = box
    row += 1
    prod_rows = prod_lines[:14] if prod_lines else []
    blank_prod = 12 if not prod_rows else max(2, 14 - len(prod_rows))
    for ln in prod_rows:
        ws.cell(row=row, column=1, value=ln["wbs_line"]).border = box
        scope = ws.cell(row=row, column=2, value=ln.get("scope", ""))
        scope.border = box
        scope.font = small
        unit = (ln.get("unit") or "").lower()
        unit_col = 6 if "ton" in unit else 7 if unit in (
            "sf", "sq ft", "sqft") else 5
        for col in range(3, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = box
            if col in (5, 6, 7) and col != unit_col and unit:
                cell.value = "x"
                cell.font = small
                cell.alignment = Alignment(horizontal="center")
        row += 1
    for _ in range(blank_prod):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = box
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value="SECTION B - MILESTONES (shop drawings, rule of credit: "
                  f"{CREDIT_LEGEND})").font = h2
    row += 1
    m_headers = ("WBS line", "Scope", "Person", "Hours",
                 "Credit reached (circle ONE: 20 / 75 / 100)", "", "",
                 "Issues / delays")
    for col, text in enumerate(m_headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = h2
        cell.border = box
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    row += 1
    mile_rows = mile_lines[:4] if mile_lines else []
    blank_mile = 3 if not mile_rows else max(1, 4 - len(mile_rows))
    for ln in mile_rows:
        ws.cell(row=row, column=1, value=ln["wbs_line"]).border = box
        scope = ws.cell(row=row, column=2, value=ln.get("scope", ""))
        scope.border = box
        scope.font = small
        credit = ws.cell(row=row, column=5, value="20   /   75   /   100")
        credit.alignment = Alignment(horizontal="center")
        for col in range(3, 9):
            ws.cell(row=row, column=col).border = box
        ws.merge_cells(start_row=row, start_column=5, end_row=row,
                       end_column=7)
        row += 1
    for _ in range(blank_mile):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = box
        credit = ws.cell(row=row, column=5, value="20   /   75   /   100")
        credit.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=5, end_row=row,
                       end_column=7)
        row += 1

    not_shown = max(0, len(prod_lines) - len(prod_rows)) \
        + max(0, len(mile_lines) - len(mile_rows))
    row += 1
    proj_arg = f" --project {project}" if project else ""
    foot_text = (
        "Production and milestone lines never share a row (P17). Hand the "
        "filled sheet to Joseph, or enter directly: py -3.13 -m "
        "takeoff_pipeline.shop_log log --person <name> --wbs <line> "
        f"--hours <h> --pieces/--tons/--sf <n>{proj_arg}")
    if not_shown:
        foot_text += (f"  NOTE: {not_shown} baseline lines did not fit "
                      "this page; use the blank rows or a second sheet.")
    foot = ws.cell(row=row, column=1, value=foot_text)
    foot.font = small
    foot.alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1,
                   end_column=8)

    ws.print_area = f"A1:H{row + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    if out_path:
        out = Path(out_path)
    else:
        DEFAULT_OUT_DIR.mkdir(parents=True, exist_ok=True)
        out = DEFAULT_OUT_DIR / f"shop_daily_sheet_{d.isoformat()}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    prefilled = len(prod_rows) + len(mile_rows)
    result = {"path": str(out), "date": d.isoformat(),
              "prefilled_lines": prefilled,
              "note": "WBS lines pre-filled from the PC1 baseline" if
              prefilled else "no baseline found; blank line column"}
    if not_shown:
        result["lines_not_shown"] = not_shown
    return result


# -- CLI ---------------------------------------------------------------------

def _print_result(result: dict) -> int:
    if "error" in result:
        print(f"ERROR: {result['error']}")
        if result.get("fix"):
            print(f"  fix: {result['fix']}")
        return 1
    for key, val in result.items():
        print(f"  {key}: {val}")
    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        prog="shop_log",
        description="PC3 shop progress capture. Five fields, under 60 "
                    "seconds a day.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_log = sub.add_parser("log", help="production entry: units plus hours")
    p_log.add_argument("--person", required=True)
    p_log.add_argument("--wbs", required=True, help="WBS line id")
    p_log.add_argument("--hours", type=float, default=0.0)
    p_log.add_argument("--pieces", type=float, default=0.0)
    p_log.add_argument("--tons", type=float, default=0.0)
    p_log.add_argument("--sf", type=float, default=0.0,
                       help="square feet decked (stored in pieces_done "
                            "per the PC3 convention)")
    p_log.add_argument("--issues", default="",
                       help="delays or problems; start with CLIENT: when "
                            "client-caused")
    p_log.add_argument("--date", default=None, help="YYYY-MM-DD, "
                       "default today")
    p_log.add_argument("--project", default="")
    p_log.add_argument("--db", default=None)

    p_mile = sub.add_parser("milestone",
                            help="shop drawing entry: rule of credit")
    p_mile.add_argument("--person", required=True)
    p_mile.add_argument("--wbs", required=True)
    p_mile.add_argument("--credit", type=float, required=True,
                        help=CREDIT_LEGEND)
    p_mile.add_argument("--hours", type=float, default=0.0)
    p_mile.add_argument("--issues", default="")
    p_mile.add_argument("--date", default=None)
    p_mile.add_argument("--project", default="")
    p_mile.add_argument("--db", default=None)

    p_sheet = sub.add_parser("sheet", help="printable one-page daily sheet")
    p_sheet.add_argument("--date", default=None)
    p_sheet.add_argument("--project", default="")
    p_sheet.add_argument("--baseline", default=None,
                         help="PC1 baseline xlsx path; auto-found from "
                              "--project when omitted")
    p_sheet.add_argument("--out", default=None)

    p_roll = sub.add_parser("rollup",
                            help="weekly planned vs actual CSV per WBS line")
    p_roll.add_argument("--week-ending", default=None,
                        help="YYYY-MM-DD; default this week's Sunday")
    p_roll.add_argument("--project", default="")
    p_roll.add_argument("--baseline", default=None)
    p_roll.add_argument("--out", default=None)
    p_roll.add_argument("--db", default=None)

    p_recent = sub.add_parser("recent", help="latest entries")
    p_recent.add_argument("--days", type=int, default=7)
    p_recent.add_argument("--db", default=None)

    p_proj = sub.add_parser("project",
                            help="show or set the sticky default project "
                                 "every entry inherits")
    p_proj.add_argument("value", nargs="?", default=None,
                        help="project id to set; omit to show the current "
                             "default")
    p_proj.add_argument("--db", default=None)

    p_init = sub.add_parser("init", help="create the progress_log table")
    p_init.add_argument("--db", default=None)

    args = parser.parse_args(argv)

    if args.command == "log":
        return _print_result(log_production(
            args.person, args.wbs, hours=args.hours, pieces=args.pieces,
            tons=args.tons, sf=args.sf, issues=args.issues,
            entry_date=args.date, project=args.project, db_path=args.db))
    if args.command == "milestone":
        return _print_result(log_milestone(
            args.person, args.wbs, args.credit, hours=args.hours,
            issues=args.issues, entry_date=args.date, project=args.project,
            db_path=args.db))
    if args.command == "sheet":
        return _print_result(build_daily_sheet(
            sheet_date=args.date, out_path=args.out, project=args.project,
            baseline_path=args.baseline))
    if args.command == "rollup":
        roll = weekly_rollup(week_ending=args.week_ending,
                             project=args.project,
                             baseline_path=args.baseline, db_path=args.db)
        if "error" in roll:
            return _print_result(roll)
        out = write_rollup_csv(roll, args.out)
        print(f"week {roll['week_start']} to {roll['week_ending']}: "
              f"{len(roll['rows'])} WBS lines -> {out}")
        for w in roll["warnings"]:
            print(f"  note: {w}")
        return 0
    if args.command == "recent":
        rows = recent_entries(days=args.days, db_path=args.db)
        if not rows:
            print("no entries in window")
            return 0
        for r in rows:
            units = (f"credit {r['pieces_done']:g}"
                     if r["progress_type"] == MILESTONE else
                     f"pieces {r['pieces_done']:g} tons {r['tons_done']:g}")
            print(f"  {r['date']}  {r['person']:<12} {r['wbs_line']:<14} "
                  f"{r['hours']:>6.1f} h  {units}"
                  + (f"  [{r['issues_text']}]" if r["issues_text"] else ""))
        return 0
    if args.command == "project":
        if args.value:
            return _print_result(set_default_project(args.value,
                                                     db_path=args.db))
        current = get_default_project(args.db)
        print(f"default project: {current}" if current else
              "no default project set; set one with: shop_log project "
              "<project-id>")
        return 0
    if args.command == "init":
        path = ensure_schema(args.db)
        print(f"progress_log ready in {path}")
        return 0
    return 1


if __name__ == "__main__":
    sys.exit(main())
