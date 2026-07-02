"""Assembly library v2 applier (Prompt 9): TAKEOFF rows to BOQ and BOM.

Implements P10, P22, P29, F3, F4 against TAKEOFF_SCHEMA_V2.md sections
3.3, 3.4, 9, 10, and 12. Where this code and that document disagree,
the document wins.

Definitions live in takeoff_pipeline/assemblies/*.json, one file per
assembly, version controlled (P29). Waste factors and rounding rules
live INSIDE the definitions. Chat-session overrides are prohibited;
changes are file diffs proposed to Owner (assemblies/README.md).
Definitions carry hours and time, never dollars (P10): a definition
containing a pricing token fails to load, and every emitted cell is
re-scanned before the workbook is saved.

Emission rules:
- Every output row cites a driving TAKEOFF item_id in source_item_id.
- Four streams per P29. Labor is HR; equipment is HR or MIN. Materials
  land on BOQ and BOM; labor, equipment, and subcontract are BOQ-only
  context (section 2: BOM is what the shop orders).
- Openings (F4, 3.4): OPENING child rows over the deck definition's
  deduct_threshold_sf are DEDUCTED from the driving area AND generate
  framed-opening steel adds plus install labor, each citing the
  OPENING row's item_id. At or under the threshold they are ignored
  (waste covers them) and counted in the run report, never silently.
- Measured-class rule (section 9): while measured ANCH or PLATE rows
  exist, the assembly emits a BOQ cross-check row, never an order
  line. Emission format, the refinement validate_takeoff.py
  anticipated: the cross-check row's source_item_id targets the FIRST
  measured row of the class (sorted by item_id) so the section 14
  WARN comparison fires mechanically; the driving rows and the
  class-total verdict ride in formula_ref and notes. With multiple
  measured rows the per-row WARN can be a false positive; the note
  carries the class-total verdict for Ivan either way. When the class
  has NO measured rows, the order line is emitted, flagged in notes.
- Dual-mode scoping (3.3): a LINEAR row carrying a GROUP: token emits
  ONLY its materials stream; the COUNT anchor row's assembly drives
  per-piece labor and equipment. Never both.
- BOM weights are FORMULA cells referencing lb_per_ft (Hard Rule 5).
  lb_per_ft fills from bridge/aisc_validator.py when importable, else
  stays blank for the operator; the skip is reported, never silent.

Pipeline position: step 3 of 13.1, after the operator fills TAKEOFF
lb_per_ft and before validation and the stamp. Stamped files are
immutable (13.4): this module refuses them. Re-running on an unstamped
workbook clears and rewrites the BOQ and BOM data rows; the TAKEOFF
sheet is never written.

Imperial units only. Free tooling (openpyxl). Paths package-relative;
resource_path() only if this is ever promoted into bridge/.
"""

import json
import math
import re
import subprocess
import sys
from pathlib import Path

from takeoff_pipeline.validate_takeoff import (
    BOQ_STREAM_UNITS, CLASSES, MODES)

_PKG = Path(__file__).resolve().parent
DEFAULT_ASM_DIR = _PKG / "assemblies"

DEF_SCHEMA = "ASSEMBLY_DEF_V2"
MEASURED_CLASSES = ("ANCH", "PLATE")
STREAM_OF = {"materials": "MATERIAL", "labor_hours": "LABOR",
             "equipment_time": "EQUIPMENT", "subcontract": "SUBCONTRACT"}
MATERIAL_UNITS = BOQ_STREAM_UNITS["MATERIAL"]
EQUIPMENT_UNITS = BOQ_STREAM_UNITS["EQUIPMENT"]

# P10 / section 12 wall, enforced at definition load AND on every
# emitted cell. Stricter than the validator's data-cell rule on
# purpose: the header token list applies, including the last word.
_PRICING_WORD = re.compile(
    r"(?i)(?<![a-z])(rate|price|cost|margin|markup|amount)(?![a-z])")

_OPENING_PARENT = re.compile(r"OPENING:\s*([A-Z]+-\d+)")
_GROUP_TOKEN = re.compile(r"GROUP:\s*([A-Z]+-\d+)")
_STAMPED_NAME = re.compile(r".+_TAKEOFF_R\d+_[0-9a-f]{12}\.xlsx$")


# -- rounding ----------------------------------------------------------------

def round_up(value, increment) -> float:
    """Round UP to the increment (P29: round UP, sheet and stick
    increments; never down). The epsilon guard keeps a float artifact
    like 30.000000000000004 from buying an extra increment, while a
    real overshoot (30.01 into increment 5) still rounds to 35."""
    if increment <= 0:
        raise ValueError("rounding increment must be positive")
    if value < 0:
        raise ValueError("round_up takes non-negative quantities")
    steps = value / increment
    eps = 1e-9 * max(1.0, abs(steps))
    return round(math.ceil(steps - eps) * increment, 9)


def _clean(qty):
    """Whole quantities write as ints so the sheets read clean."""
    f = float(qty)
    return int(f) if f.is_integer() else f


# -- definition loading ------------------------------------------------------

def _scan_pricing_tokens(text, where) -> None:
    if "$" in text:
        raise ValueError(f"{where}: dollar character violates the "
                         "section 12 pricing wall (P10)")
    m = _PRICING_WORD.search(text)
    if m:
        raise ValueError(f"{where}: pricing token '{m.group(1)}' "
                         "violates the section 12 pricing wall (P10)")


def _require(cond, where, msg) -> None:
    if not cond:
        raise ValueError(f"{where}: {msg}")


def _check_round(rnd, where) -> None:
    _require(isinstance(rnd, dict), where, "round must be an object")
    _require(rnd.get("direction") == "up", where,
             "round.direction must be 'up'; order quantities never "
             "round down (P29)")
    inc = rnd.get("increment")
    _require(isinstance(inc, (int, float)) and inc > 0, where,
             "round.increment must be a positive number")


def _check_material(comp, where, mode) -> None:
    _require(comp.get("component_id"), where, "missing component_id")
    _require(comp.get("description"), where, "missing description")
    _require(comp.get("unit") in MATERIAL_UNITS, where,
             f"material unit must be one of {', '.join(MATERIAL_UNITS)}")
    has_qpu = isinstance(comp.get("qty_per_unit"), (int, float))
    has_pss = isinstance(comp.get("per_sheet_sf"), (int, float))
    _require(has_qpu != has_pss, where,
             "exactly one of qty_per_unit and per_sheet_sf")
    if has_qpu:
        _require(comp["qty_per_unit"] > 0, where,
                 "qty_per_unit must be positive")
    if has_pss:
        _require(comp["per_sheet_sf"] > 0, where,
                 "per_sheet_sf must be positive")
        _require(mode == "AREA", where,
                 "per_sheet_sf only applies to AREA drivers")
        _require(comp.get("unit") == "EA", where,
                 "per_sheet_sf components order whole sheets, unit EA")
    wf = comp.get("waste_factor")
    _require(isinstance(wf, (int, float)) and 0 <= wf < 1, where,
             "waste_factor must be a number in [0, 1)")
    _check_round(comp.get("round"), where)
    ccc = comp.get("cross_check_class")
    _require(ccc is None or ccc in MEASURED_CLASSES, where,
             f"cross_check_class must be null or one of "
             f"{', '.join(MEASURED_CLASSES)}")
    _require(ccc is None or has_qpu, where,
             "cross_check components compare counts; use "
             "qty_per_unit, never per_sheet_sf")
    _require(bool(comp.get("designation"))
             or comp.get("designation_from_driver") is True, where,
             "needs a designation or designation_from_driver true")
    llf = comp.get("length_lf")
    _require(llf is None or (isinstance(llf, (int, float)) and llf > 0),
             where, "length_lf must be null or positive")


def _check_labor(comp, where, qty_key="hours_per_unit") -> None:
    _require(comp.get("component_id"), where, "missing component_id")
    _require(comp.get("description"), where, "missing description")
    v = comp.get(qty_key)
    _require(isinstance(v, (int, float)) and v > 0, where,
             f"{qty_key} must be a positive number")
    _check_round(comp.get("round"), where)


def _check_equipment(comp, where) -> None:
    _require(comp.get("component_id"), where, "missing component_id")
    _require(comp.get("description"), where, "missing description")
    _require(comp.get("unit") in EQUIPMENT_UNITS, where,
             "equipment unit must be HR or MIN")
    v = comp.get("time_per_unit")
    _require(isinstance(v, (int, float)) and v > 0, where,
             "time_per_unit must be a positive number")
    _check_round(comp.get("round"), where)


def _validate_definition(d, path) -> None:
    where = f"assembly definition {path.name}"
    _require(d.get("schema") == DEF_SCHEMA, where,
             f"schema must be '{DEF_SCHEMA}'")
    _require(d.get("assembly_id") == path.stem, where,
             f"assembly_id must equal the file stem '{path.stem}'")
    for key in ("description", "provenance"):
        _require(d.get(key), where, f"missing {key}")
    a = d.get("applies_to") or {}
    _require(a.get("item_class") in CLASSES, where,
             f"applies_to.item_class must be one of {', '.join(CLASSES)}")
    _require(a.get("mode") in MODES, where,
             "applies_to.mode must be COUNT, LINEAR, or AREA")
    for key in ("designation_includes", "designation_excludes"):
        _require(isinstance(a.get(key), list), where,
                 f"applies_to.{key} must be a list")

    streams = d.get("streams") or {}
    _require(set(streams) == set(STREAM_OF), where,
             "streams must carry exactly the four P29 keys: "
             + ", ".join(STREAM_OF))
    for comp in streams["materials"]:
        _check_material(comp, f"{where} materials/"
                        f"{comp.get('component_id', '?')}", a["mode"])
    for comp in streams["labor_hours"]:
        _check_labor(comp, f"{where} labor_hours/"
                     f"{comp.get('component_id', '?')}")
    for comp in streams["equipment_time"]:
        _check_equipment(comp, f"{where} equipment_time/"
                         f"{comp.get('component_id', '?')}")
    sub = streams["subcontract"]
    _require(sub is None or isinstance(sub, list), where,
             "subcontract must be null or a list (nullable per P29)")
    for comp in sub or []:
        _check_material(comp, f"{where} subcontract/"
                        f"{comp.get('component_id', '?')}", a["mode"])
        _require(not comp.get("cross_check_class"), where,
                 "cross_check_class is a materials-stream concept")
        _require(comp.get("per_sheet_sf") is None, where,
                 "per_sheet_sf is a materials-stream concept; "
                 "subcontract scope uses qty_per_unit")

    op = d.get("openings", None)
    _require(op is None or isinstance(op, dict), where,
             "openings must be null or an object")
    if op:
        _require(a["mode"] == "AREA", where,
                 "openings logic only applies to AREA drivers (3.4)")
        thr = op.get("deduct_threshold_sf")
        _require(isinstance(thr, (int, float)) and thr > 0, where,
                 "openings.deduct_threshold_sf must be positive SF")
        _require(op.get("geometry_model") == "square", where,
                 "openings.geometry_model must be 'square' (the only "
                 "model this version implements)")
        for comp in op.get("framed_adds") or []:
            cw = (f"{where} framed_adds/"
                  f"{comp.get('component_id', '?')}")
            _require(comp.get("component_id"), cw, "missing component_id")
            _require(comp.get("description"), cw, "missing description")
            _require(comp.get("designation"), cw, "missing designation")
            _require(comp.get("basis") in ("perimeter", "span"), cw,
                     "basis must be 'perimeter' or 'span'")
            ppo = comp.get("pieces_per_opening")
            _require(isinstance(ppo, (int, float)) and ppo >= 1, cw,
                     "pieces_per_opening must be >= 1")
            _require(comp.get("unit") == "LF", cw,
                     "framed adds are stick steel, unit LF")
            wf = comp.get("waste_factor")
            _require(isinstance(wf, (int, float)) and 0 <= wf < 1, cw,
                     "waste_factor must be a number in [0, 1)")
            _check_round(comp.get("round"), cw)
        for comp in op.get("labor_hours") or []:
            _check_labor(comp, f"{where} openings labor_hours/"
                         f"{comp.get('component_id', '?')}",
                         qty_key="hours_per_opening")


def load_definitions(asm_dir=None) -> dict:
    """All *.json definitions in the folder, validated and token
    scanned. Fails loud on the first malformed or pricing-bearing
    definition; a half-loaded library never applies."""
    asm_dir = Path(asm_dir) if asm_dir else DEFAULT_ASM_DIR
    if not asm_dir.is_dir():
        raise ValueError(f"assembly folder not found: {asm_dir}")
    defs = {}
    for path in sorted(asm_dir.glob("*.json")):
        raw = path.read_text(encoding="utf-8")
        _scan_pricing_tokens(raw, f"assembly definition {path.name}")
        try:
            d = json.loads(raw)
        except ValueError as e:
            raise ValueError(f"assembly definition {path.name}: "
                             f"invalid JSON ({e})") from None
        _validate_definition(d, path)
        defs[d["assembly_id"]] = d
    if not defs:
        raise ValueError(f"no assembly definitions in {asm_dir}")
    return defs


# -- TAKEOFF reading ---------------------------------------------------------

def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _read_takeoff(ws) -> list:
    rows = []
    for r in range(3, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        if all(v is None for v in values):
            continue
        item_id = _text(values[0])
        rows.append({
            "row": r,
            "item_id": item_id,
            "item_class": item_id.split("-")[0] if "-" in item_id else "",
            "designation": _text(values[1]),
            "mode": _text(values[2]).upper(),
            "qty": values[3],
            "unit": _text(values[4]).upper(),
            "notes": _text(values[10]),
        })
    return rows


def _refuse_stamped(ws, path) -> None:
    """13.4: a stamped file is immutable. Anything in row 1 or a
    stamped-pattern filename means stamped; refuse, fail closed."""
    row1 = any(ws.cell(row=1, column=c).value is not None
               for c in range(1, 16))
    if row1 or _STAMPED_NAME.match(Path(path).name):
        raise ValueError(
            "refusing to write: the workbook is stamped and immutable "
            "(13.4). Rerun the pipeline to produce a new export.")


def _match_assembly(row, defs):
    """The single definition applying to a TAKEOFF row, or None.
    OPENING rows never drive an assembly directly (they ride through
    their parent, 3.4). Two matching definitions is a library defect
    and fails loud, never a silent pick."""
    if row["designation"].upper() == "OPENING":
        return None
    matches = []
    for d in defs.values():
        a = d["applies_to"]
        if a["item_class"] != row["item_class"]:
            continue
        if a["mode"] != row["mode"]:
            continue
        desig = row["designation"].upper()
        inc = a["designation_includes"]
        if inc and not any(t.upper() in desig for t in inc):
            continue
        if any(t.upper() in desig for t in a["designation_excludes"]):
            continue
        matches.append(d)
    if len(matches) > 1:
        names = ", ".join(sorted(m["assembly_id"] for m in matches))
        raise ValueError(
            f"TAKEOFF row {row['row']} ({row['item_id']} "
            f"'{row['designation']}') matches more than one assembly "
            f"definition: {names}. Fix the designation_includes/"
            "excludes lists; ambiguity is never resolved silently.")
    return matches[0] if matches else None


# -- emission ----------------------------------------------------------------

def _asm_ref(assembly_id, component_id, detail) -> str:
    return f"ASM:{assembly_id}:{component_id}:{detail}"


def _emit_material(boq, bom, d, comp, source_id, out_qty, detail,
                   driver_designation, note="") -> None:
    """One material component onto BOQ (quantity context) and BOM
    (order line). aisc components get the Hard Rule 5 treatment in
    the BOM writer: lb_per_ft input plus formula weight cells."""
    if out_qty <= 0:
        return
    aid = d["assembly_id"]
    ref = _asm_ref(aid, comp["component_id"], detail)
    designation = (driver_designation
                   if comp.get("designation_from_driver")
                   else comp["designation"])
    boq.append({
        "source": source_id, "assembly_id": aid, "stream": "MATERIAL",
        "description": comp["description"], "qty": _clean(out_qty),
        "unit": comp["unit"], "formula_ref": ref, "notes": note,
    })
    bom.append({
        "source": source_id, "designation": designation,
        "description": comp["description"], "qty": _clean(out_qty),
        "unit": comp["unit"], "length_lf": comp.get("length_lf"),
        "aisc": bool(comp.get("aisc")), "formula_ref": ref,
        "notes": note,
    })


def _emit_time_streams(boq, d, row, driver_qty, detail_base) -> None:
    """labor_hours, equipment_time, and subcontract for one driver
    row. BOQ only: the BOM is what the shop orders."""
    aid = d["assembly_id"]
    for comp in d["streams"]["labor_hours"]:
        qty = round_up(driver_qty * comp["hours_per_unit"],
                       comp["round"]["increment"])
        if qty <= 0:
            continue
        boq.append({
            "source": row["item_id"], "assembly_id": aid,
            "stream": "LABOR", "description": comp["description"],
            "qty": _clean(qty), "unit": "HR",
            "formula_ref": _asm_ref(
                aid, comp["component_id"],
                f"{detail_base} x {comp['hours_per_unit']:g}HR "
                f"roundup {comp['round']['increment']:g}"),
            "notes": "",
        })
    for comp in d["streams"]["equipment_time"]:
        qty = round_up(driver_qty * comp["time_per_unit"],
                       comp["round"]["increment"])
        if qty <= 0:
            continue
        boq.append({
            "source": row["item_id"], "assembly_id": aid,
            "stream": "EQUIPMENT", "description": comp["description"],
            "qty": _clean(qty), "unit": comp["unit"],
            "formula_ref": _asm_ref(
                aid, comp["component_id"],
                f"{detail_base} x {comp['time_per_unit']:g}"
                f"{comp['unit']} roundup "
                f"{comp['round']['increment']:g}"),
            "notes": "",
        })
    for comp in d["streams"]["subcontract"] or []:
        qty = round_up(driver_qty * comp["qty_per_unit"]
                       * (1 + comp["waste_factor"]),
                       comp["round"]["increment"])
        if qty <= 0:
            continue
        boq.append({
            "source": row["item_id"], "assembly_id": aid,
            "stream": "SUBCONTRACT",
            "description": comp["description"], "qty": _clean(qty),
            "unit": comp["unit"],
            "formula_ref": _asm_ref(
                aid, comp["component_id"],
                f"{detail_base} x {comp['qty_per_unit']:g} waste "
                f"{comp['waste_factor']:g} roundup "
                f"{comp['round']['increment']:g}"),
            "notes": "scope context, never an order line",
        })


def _emit_openings(boq, bom, d, parent, openings, report) -> float:
    """F4 deduct-and-add for one deck driver row. Returns the net
    driving area. Over-threshold openings deduct AND add framed steel
    plus install labor, each citing the OPENING row's item_id.
    At-or-under-threshold openings are counted, never silent."""
    op = d["openings"]
    threshold = op["deduct_threshold_sf"]
    qualifying, ignored = [], []
    for o in sorted(openings, key=lambda r: r["item_id"]):
        if not isinstance(o["qty"], (int, float)):
            report["warnings"].append(
                f"{o['item_id']}: OPENING qty '{o['qty']}' is not a "
                "number; not deducted, no adds. Ivan review.")
            continue
        (qualifying if o["qty"] > threshold else ignored).append(o)

    net = float(parent["qty"]) - sum(o["qty"] for o in qualifying)
    if net < 0:
        report["warnings"].append(
            f"{parent['item_id']}: qualifying openings total "
            f"{sum(o['qty'] for o in qualifying):g} SF, more than the "
            f"parent area {parent['qty']:g} SF. Net clamped to 0; the "
            "takeoff rows need Ivan before this BOQ is usable.")
        net = 0.0
    report["openings"]["qualifying"] += len(qualifying)
    report["openings"]["ignored_under_threshold"] += len(ignored)
    if ignored:
        report["openings"]["ignored_detail"].append(
            f"{parent['item_id']}: "
            + ", ".join(f"{o['item_id']} {o['qty']:g}SF" for o in ignored)
            + f" at or under {threshold:g}SF threshold; waste covers")

    aid = d["assembly_id"]
    for o in qualifying:
        side = math.sqrt(float(o["qty"]))
        for comp in op.get("framed_adds") or []:
            basis_len = 4 * side if comp["basis"] == "perimeter" \
                else side
            lf = basis_len * comp["pieces_per_opening"]
            qty = round_up(lf * (1 + comp["waste_factor"]),
                           comp["round"]["increment"])
            detail = (f"{o['item_id']} {o['qty']:g}SF square model "
                      f"side {side:.2f}LF basis {comp['basis']} x "
                      f"{comp['pieces_per_opening']:g} waste "
                      f"{comp['waste_factor']:g} roundup "
                      f"{comp['round']['increment']:g}")
            _emit_material(
                boq, bom, d, comp, o["item_id"], qty, detail,
                comp["designation"],
                note=f"framed-opening add per F4, parent "
                     f"{parent['item_id']}")
        for comp in op.get("labor_hours") or []:
            qty = round_up(comp["hours_per_opening"],
                           comp["round"]["increment"])
            boq.append({
                "source": o["item_id"], "assembly_id": aid,
                "stream": "LABOR", "description": comp["description"],
                "qty": _clean(qty), "unit": "HR",
                "formula_ref": _asm_ref(
                    aid, comp["component_id"],
                    f"{o['item_id']} x "
                    f"{comp['hours_per_opening']:g}HR roundup "
                    f"{comp['round']['increment']:g}"),
                "notes": f"framed-opening install per F4, parent "
                         f"{parent['item_id']}",
            })
    return net


def _finalize_cross_checks(boq, cross_acc, rows, report) -> None:
    """Section 9 cross-check rows, one per accumulated (assembly,
    component). source_item_id targets the first measured row of the
    class so the validator's WARN comparison fires; the class-total
    verdict rides in the notes for Ivan."""
    for (aid, cid), acc in sorted(cross_acc.items()):
        cls = acc["class"]
        measured = sorted(
            (r for r in rows if r["item_class"] == cls),
            key=lambda r: r["item_id"])
        count_rows = [r for r in measured if r["mode"] == "COUNT"
                      and isinstance(r["qty"], (int, float))]
        derived = acc["derived"]
        drivers = "+".join(i for i, _ in acc["drivers"])
        if count_rows:
            total = sum(float(r["qty"]) for r in count_rows)
            ids = ", ".join(f"{r['item_id']} {r['qty']:g} EA"
                            for r in count_rows)
            verdict = ("MATCH" if abs(derived - total) < 1e-9
                       else "MISMATCH, RFI candidate for Ivan")
            note = (f"CROSS-CHECK ONLY, never an order line (section "
                    f"9). Assembly-derived {cls} count {derived:g} vs "
                    f"measured class total {total:g} across {ids}: "
                    f"{verdict}. Measured rows are the system of "
                    "record for ordering (P30).")
        else:
            verdict = "NO COUNT ROWS TO COMPARE"
            note = (f"CROSS-CHECK ONLY, never an order line (section "
                    f"9). Assembly-derived {cls} count {derived:g}; "
                    f"measured {cls} rows exist but none in COUNT "
                    "mode, so no count comparison. Ivan review.")
        report["cross_checks"].append(
            f"{aid}/{cid}: derived {derived:g} {cls}, {verdict}")
        boq.append({
            "source": measured[0]["item_id"], "assembly_id": aid,
            "stream": "MATERIAL",
            "description": f"CROSS-CHECK ONLY, not an order line: "
                           f"assembly-derived {cls} count",
            "qty": _clean(derived), "unit": acc["unit"],
            "formula_ref": _asm_ref(
                aid, cid, f"drivers {drivers} derived {derived:g}"
                          f"{acc['unit']}"),
            "notes": note,
        })


# -- workbook write ----------------------------------------------------------

def _guard_emitted(rows_out, sheet) -> None:
    """Belt and braces: definitions were scanned at load, the emitted
    text is scanned again. A pricing token here is a code defect."""
    for i, row in enumerate(rows_out, start=2):
        for key, value in row.items():
            if isinstance(value, str):
                _scan_pricing_tokens(value, f"{sheet} row {i} {key}")


def _clear_data_rows(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _aisc_filler(report):
    """A lb_per_ft lookup through bridge/aisc_validator.py, the only
    sanctioned weight source (Hard Rule 5). When bridge/ is not
    importable the fill is skipped and REPORTED; the operator fills
    the input cells, exactly like the TAKEOFF sheet flow."""
    cache = {}
    try:
        from bridge.aisc_validator import validate_shape
    except Exception:
        report["aisc_fill"] = (
            "bridge.aisc_validator not importable; BOM lb_per_ft left "
            "blank for the operator (reported, never silent)")
        return lambda designation: None

    def fill(designation):
        if designation not in cache:
            result = validate_shape(designation)
            ok = result.get("valid") and result.get("weight_per_ft")
            cache[designation] = float(result["weight_per_ft"]) \
                if ok else None
            if not ok:
                report["aisc_unresolved"].append(designation)
        return cache[designation]
    return fill


def _write_boq(ws, rows_out) -> None:
    for i, row in enumerate(rows_out, start=1):
        ws.append([f"BOQ-{i:04d}", row["source"], row["assembly_id"],
                   row["stream"], row["description"], row["qty"],
                   row["unit"], row["formula_ref"], row["notes"]])


def _write_bom(ws, rows_out, lb_fill) -> None:
    """Section 10 weight cases. unit LF: weight = qty x lb_per_ft.
    unit EA stick with length_lf: weight = qty x length_lf x
    lb_per_ft. Weight cells are formulas, never typed constants;
    lb_per_ft is the input cell (Hard Rule 5)."""
    for i, row in enumerate(rows_out, start=1):
        r = i + 1
        lb_per_ft = weight = tons = None
        ref = row["formula_ref"]
        if row["aisc"]:
            lb_per_ft = lb_fill(row["designation"])
            if row["unit"] == "LF":
                weight, tons = f"=E{r}*H{r}", f"=I{r}/2000"
            elif row["unit"] == "EA" and row["length_lf"]:
                weight, tons = f"=E{r}*G{r}*H{r}", f"=I{r}/2000"
            if weight:
                ref = (f"{ref}; AISC:bridge/aisc_validator.py:"
                       f"{row['designation']}; DERIVED:weight_lb/2000")
        ws.append([f"BOM-{i:04d}", row["source"], row["designation"],
                   row["description"], row["qty"], row["unit"],
                   row["length_lf"], lb_per_ft, weight, tons, ref,
                   row["notes"]])


def _assemblies_dirty_note(asm_dir):
    """The stamp's F1 cell records the committed hash of assemblies/.
    Uncommitted definition edits would stamp a hash that lies, so the
    dirty state is surfaced here, at apply time."""
    if Path(asm_dir).resolve() != DEFAULT_ASM_DIR.resolve():
        return ("definitions were loaded from a non-default folder; "
                "the F1 stamp records only the committed hash of "
                "takeoff_pipeline/assemblies/, so this output must "
                "not be stamped (13.3)")
    try:
        out = subprocess.run(
            ["git", "status", "--porcelain", "--", str(DEFAULT_ASM_DIR)],
            capture_output=True, text=True, cwd=str(_PKG.parent),
            timeout=10)
        if out.stdout.strip():
            return ("takeoff_pipeline/assemblies/ has uncommitted "
                    "changes; commit them before stamping or the F1 "
                    "assembly-library hash will not describe the "
                    "definitions actually applied (13.3)")
    except Exception:
        return ("could not check assemblies/ git state; verify the "
                "folder is committed before stamping (13.3)")
    return None


# -- apply -------------------------------------------------------------------

def apply_to_workbook(path, asm_dir=None, out_path=None) -> dict:
    """Pipeline step 3 (13.1): read TAKEOFF, apply the definitions,
    rewrite the BOQ and BOM data rows. Returns the run report."""
    import openpyxl

    path = Path(path)
    asm_dir = Path(asm_dir) if asm_dir else DEFAULT_ASM_DIR
    defs = load_definitions(asm_dir)

    wb = openpyxl.load_workbook(str(path))
    for name in ("TAKEOFF", "BOQ", "BOM"):
        if name not in wb.sheetnames:
            raise ValueError(f"workbook is missing the {name} sheet")
    ws = wb["TAKEOFF"]
    _refuse_stamped(ws, path)

    rows = _read_takeoff(ws)
    report = {
        "path": str(path), "definitions": len(defs),
        "takeoff_rows": len(rows), "drivers_applied": 0,
        "unassigned": [], "measured_not_drivers": [],
        "skipped_zero_qty": [], "group_suppressed": [],
        "openings": {"qualifying": 0, "ignored_under_threshold": 0,
                     "ignored_detail": [], "orphaned": []},
        "cross_checks": [], "aisc_unresolved": [],
        "warnings": [],
    }

    by_id = {r["item_id"]: r for r in rows}
    openings_by_parent = {}
    # First pass: every OPENING row, BEFORE parents resolve. Resolving
    # in the same pass made the opening-cites-opening guard depend on
    # sheet order: a parent appearing LATER slipped past and its child
    # vanished with no warning, against the 3.4/F4 never-silent rule.
    opening_ids = {r["item_id"] for r in rows
                   if r["item_class"] == "DECK"
                   and r["designation"].upper() == "OPENING"}
    for r in rows:
        if r["item_id"] not in opening_ids:
            continue
        m = _OPENING_PARENT.search(r["notes"])
        parent = by_id.get(m.group(1)) if m else None
        if not m:
            report["warnings"].append(
                f"{r['item_id']}: OPENING row with no parseable "
                "parent token; not deducted, no adds. Ivan review "
                "(the validator hard-fails this file at step 4)")
        elif parent is None or parent["item_class"] != "DECK" \
                or parent["mode"] != "AREA" \
                or parent["item_id"] in opening_ids:
            report["warnings"].append(
                f"{r['item_id']}: OPENING parent '{m.group(1)}' is "
                "not a deck AREA row; not deducted, no adds. Ivan "
                "review.")
        else:
            openings_by_parent.setdefault(
                parent["item_id"], []).append(r)

    boq, bom = [], []
    cross_acc = {}

    for row in rows:
        if row["item_id"] in opening_ids:
            continue
        d = _match_assembly(row, defs)
        if d is None:
            if row["item_class"] in MEASURED_CLASSES:
                # Expected, not a library gap: measured rows are the
                # system of record (P30); assemblies cross-check
                # against them, they never drive emission (section 9).
                report["measured_not_drivers"].append(
                    f"{row['item_id']} '{row['designation']}': "
                    "measured class, system of record, never an "
                    "assembly driver")
                continue
            report["unassigned"].append(
                f"{row['item_id']} '{row['designation']}' "
                f"{row['mode']}: no assembly definition matches; no "
                "BOQ or BOM rows derived from it")
            for o in openings_by_parent.get(row["item_id"], []):
                report["openings"]["orphaned"].append(
                    f"{o['item_id']}: parent {row['item_id']} is "
                    "unassigned; not deducted, no adds")
            continue
        if not isinstance(row["qty"], (int, float)):
            report["warnings"].append(
                f"{row['item_id']}: qty '{row['qty']}' is not a "
                "number; skipped")
            continue
        if float(row["qty"]) <= 0:
            report["skipped_zero_qty"].append(
                f"{row['item_id']} qty {row['qty']:g} (REMOVED or "
                "zeroed row; history rides, nothing derives)")
            continue

        driver_qty = float(row["qty"])
        if d.get("openings") and row["mode"] == "AREA":
            driver_qty = _emit_openings(
                boq, bom, d, row,
                openings_by_parent.get(row["item_id"], []), report)
        detail_base = f"{row['item_id']} {driver_qty:g}{row['unit']}"

        for comp in d["streams"]["materials"]:
            ccc = comp.get("cross_check_class")
            if ccc and any(r["item_class"] == ccc for r in rows):
                key = (d["assembly_id"], comp["component_id"])
                acc = cross_acc.setdefault(
                    key, {"class": ccc, "unit": comp["unit"],
                          "derived": 0.0, "drivers": []})
                acc["derived"] += driver_qty * comp["qty_per_unit"]
                acc["drivers"].append((row["item_id"], driver_qty))
                continue
            if "per_sheet_sf" in comp \
                    and comp["per_sheet_sf"] is not None:
                qty = round_up(
                    driver_qty * (1 + comp["waste_factor"])
                    / comp["per_sheet_sf"],
                    comp["round"]["increment"])
                detail = (f"{detail_base} net, waste "
                          f"{comp['waste_factor']:g}, "
                          f"{comp['per_sheet_sf']:g}SF per sheet, "
                          f"roundup {comp['round']['increment']:g}")
            else:
                qty = round_up(
                    driver_qty * comp["qty_per_unit"]
                    * (1 + comp["waste_factor"]),
                    comp["round"]["increment"])
                detail = (f"{detail_base} x {comp['qty_per_unit']:g} "
                          f"waste {comp['waste_factor']:g} roundup "
                          f"{comp['round']['increment']:g}")
            note = ""
            if ccc:
                note = (f"no measured {ccc} rows in this takeoff; "
                        "assembly-derived order line, verify against "
                        "the schedule (section 9)")
            _emit_material(boq, bom, d, comp, row["item_id"], qty,
                           detail, row["designation"], note=note)

        if _GROUP_TOKEN.search(row["notes"]):
            report["group_suppressed"].append(
                f"{row['item_id']}: GROUP-linked row, materials "
                "stream only; the COUNT anchor's assembly drives "
                "labor and equipment (3.3)")
        else:
            _emit_time_streams(boq, d, row, driver_qty, detail_base)
        report["drivers_applied"] += 1

    _finalize_cross_checks(boq, cross_acc, rows, report)

    _guard_emitted(boq, "BOQ")
    _guard_emitted(bom, "BOM")

    _clear_data_rows(wb["BOQ"])
    _clear_data_rows(wb["BOM"])
    _write_boq(wb["BOQ"], boq)
    _write_bom(wb["BOM"], bom, _aisc_filler(report))

    target = Path(out_path) if out_path else path
    wb.save(str(target))

    report["out_path"] = str(target)
    report["boq_rows"] = len(boq)
    report["bom_rows"] = len(bom)
    dirty = _assemblies_dirty_note(asm_dir)
    if dirty:
        report["warnings"].append(dirty)
    report["status"] = "APPLIED"
    return report


# -- CLI ---------------------------------------------------------------------

def _flag_value(args, flag):
    if flag not in args:
        return None
    i = args.index(flag)
    if i + 1 >= len(args):
        raise ValueError(f"{flag} needs a value")
    return args[i + 1]


def main() -> int:
    usage = ("usage: py -m takeoff_pipeline.apply_assemblies "
             "<takeoff.xlsx> [--assemblies <dir>] [--out <path>] | "
             "--check-defs [--assemblies <dir>]")
    args = sys.argv[1:]
    try:
        asm_dir = _flag_value(args, "--assemblies")
        if "--check-defs" in args:
            defs = load_definitions(asm_dir)
            print(f"definitions OK: {len(defs)}")
            for aid in sorted(defs):
                a = defs[aid]["applies_to"]
                print(f"  {aid}: {a['item_class']} {a['mode']}")
            return 0
        if not args or args[0].startswith("--"):
            print(usage)
            return 2
        report = apply_to_workbook(args[0], asm_dir=asm_dir,
                                   out_path=_flag_value(args, "--out"))
    except ValueError as e:
        print(e)
        return 1
    for key, value in report.items():
        if isinstance(value, list):
            for line in value:
                print(f"{key}: {line}")
        elif isinstance(value, dict):
            print(f"{key}: {json.dumps(value)}")
        else:
            print(f"{key}: {value}")
    return 0 if report.get("status") == "APPLIED" else 1


if __name__ == "__main__":
    raise SystemExit(main())
